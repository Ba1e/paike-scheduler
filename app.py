from flask import Flask, jsonify, request, send_file, redirect, make_response, has_request_context
from flask_cors import CORS
from werkzeug.exceptions import HTTPException
import copy
import gzip
import io
import json
import os
import time
import re
import hashlib
import ipaddress
import secrets
import shutil
from datetime import datetime, timedelta, timezone
from threading import Lock
from functools import wraps

import sqlite_store

app = Flask(__name__)
ALLOWED_ORIGINS = [o.strip() for o in os.environ.get('SCHED_ALLOWED_ORIGINS', '').split(',') if o.strip()]
if ALLOWED_ORIGINS:
    CORS(app, supports_credentials=True, origins=ALLOWED_ORIGINS)

# ==================== 路径与常量 ====================
SOURCE_DIR = os.path.dirname(__file__)
BASE_DIR = os.environ.get('SCHED_DATA_DIR', SOURCE_DIR)
DEPTS_DIR = os.path.join(BASE_DIR, 'departments')
DEPTS_CONFIG = os.path.join(BASE_DIR, 'departments.json')
USERS_FILE = os.path.join(BASE_DIR, 'users.json')
INVITES_FILE = os.path.join(BASE_DIR, 'invite_codes.json')
SESSIONS_FILE = os.path.join(BASE_DIR, 'sessions.json')
BOOTSTRAP_FILE = os.path.join(BASE_DIR, 'BOOTSTRAP_CODE.txt')

DEFAULT_TERM_ID = 'fy27_summer_autumn'
DEFAULT_TERM_NAME = 'FY27暑秋'
SESSION_COOKIE = 'sched_token'
CSRF_COOKIE = 'sched_csrf'
CSRF_HEADER = 'X-CSRF-Token'
SESSION_TTL = 30 * 86400
COOKIE_SECURE = os.environ.get('SCHED_COOKIE_SECURE', '0') == '1'
TRUST_PROXY_HEADERS = os.environ.get('SCHED_TRUST_PROXY_HEADERS', '0') == '1'
TRUSTED_PROXY_IPS = [
    item.strip()
    for item in os.environ.get('SCHED_TRUSTED_PROXY_IPS', '127.0.0.1,::1').split(',')
    if item.strip()
]
HISTORY_KEEP_FILES = 50
DAILY_BACKUP_KEEP_DAYS = 30
MAX_IMPORT_BYTES = 20 * 1024 * 1024
GZIP_MIN_BYTES = int(os.environ.get('SCHED_GZIP_MIN_BYTES', '4096'))
REQUEST_TIMING_ENABLED = os.environ.get('SCHED_REQUEST_TIMING', '1') == '1'
try:
    SLOW_REQUEST_SECONDS = float(os.environ.get('SCHED_SLOW_REQUEST_SECONDS', '2.0'))
except ValueError:
    SLOW_REQUEST_SECONDS = 2.0
try:
    SLOW_REQUEST_HISTORY_LIMIT = max(0, int(os.environ.get('SCHED_SLOW_REQUEST_HISTORY_LIMIT', '20')))
except ValueError:
    SLOW_REQUEST_HISTORY_LIMIT = 20
SAFE_METHODS = {'GET', 'HEAD', 'OPTIONS'}
CSRF_EXEMPT_PATHS = {
    '/api/auth/login',
    '/api/auth/register',
}
STAFF_ROLES = {'admin', 'jiaowu'}
DEPT_MANAGER_ROLES = {'supervisor', 'regional_manager', 'director'}
EDIT_ROLES = STAFF_ROLES | DEPT_MANAGER_ROLES | {'store_manager'}
VALID_ROLES = {'user', 'admin', 'jiaowu', 'supervisor', 'store_manager', 'regional_manager', 'director'}
SLOT_TIME_MAP = {
    'qingshao': {
        'A': '08:30-10:30', 'B': '10:40-12:40',
        'C': '14:00-16:00', 'D': '16:10-18:10', 'E': '18:30-20:30',
    },
    'default': {
        'A': '08:00-10:00', 'B': '10:20-12:20',
        'C': '13:30-15:30', 'D': '15:50-17:50', 'E': '18:30-20:30',
    },
}
DEFAULT_CAMPUS_CONFIG = {
    'districts': {
        '禅城大区': ['友邦金融中心教学区', '映月湖环宇城教学区', '新兆阳广场教学区'],
        '南海大区': ['新南万教学区', '广佛智城教学区', 'IPARK购物中心教学区'],
        '顺德大区': ['北滘悦然广场教学区', '容桂桂洲大道教学区'],
        '其他': ['富凯广场教学区', '铂顿城教学区', '禅西环宇城教学区'],
    },
    'campus_codes': {},
}
CANONICAL_CAMPUSES = {
    'IPARK购物中心': 'IPARK购物中心教学区',
    'IPARK': 'IPARK购物中心教学区',
    'IP': 'IPARK购物中心教学区',
    '北滘悦然广场': '北滘悦然广场教学区',
    '北滘': '北滘悦然广场教学区',
    '铂顿城': '铂顿城教学区',
    '铂顿': '铂顿城教学区',
    '禅西环宇城': '禅西环宇城教学区',
    '禅西': '禅西环宇城教学区',
    '大良新一城': '大良新一城教学区',
    '大良': '大良新一城教学区',
    '富凯广场': '富凯广场教学区',
    '富凯': '富凯广场教学区',
    '广佛智城': '广佛智城教学区',
    '大沥': '广佛智城教学区',
    '容桂桂洲大道': '容桂桂洲大道教学区',
    '容桂': '容桂桂洲大道教学区',
    '新南万': '新南万教学区',
    '万科': '新南万教学区',
    '南海万科': '新南万教学区',
    '新兆阳广场': '新兆阳广场教学区',
    '新兆阳': '新兆阳广场教学区',
    '映月湖环宇城': '映月湖环宇城教学区',
    '映月湖': '映月湖环宇城教学区',
    '友邦金融中心': '友邦金融中心教学区',
    '友邦': '友邦金融中心教学区',
}
CAMPUS_TRAVEL_MINUTES = {
    '铂顿城教学区': {'铂顿城教学区': 0, '新兆阳广场教学区': 15, '富凯广场教学区': 18, '禅西环宇城教学区': 20, '友邦金融中心教学区': 22, '新南万教学区': 23, '广佛智城教学区': 32, '映月湖环宇城教学区': 25, 'IPARK购物中心教学区': 24, '北滘悦然广场教学区': 42, '大良新一城教学区': 62, '容桂桂洲大道教学区': 68},
    '新兆阳广场教学区': {'铂顿城教学区': 15, '新兆阳广场教学区': 0, '富凯广场教学区': 16, '禅西环宇城教学区': 18, '友邦金融中心教学区': 20, '新南万教学区': 20, '广佛智城教学区': 30, '映月湖环宇城教学区': 22, 'IPARK购物中心教学区': 20, '北滘悦然广场教学区': 40, '大良新一城教学区': 60, '容桂桂洲大道教学区': 65},
    '富凯广场教学区': {'铂顿城教学区': 18, '新兆阳广场教学区': 16, '富凯广场教学区': 0, '禅西环宇城教学区': 22, '友邦金融中心教学区': 25, '新南万教学区': 27, '广佛智城教学区': 35, '映月湖环宇城教学区': 20, 'IPARK购物中心教学区': 26, '北滘悦然广场教学区': 45, '大良新一城教学区': 58, '容桂桂洲大道教学区': 63},
    '禅西环宇城教学区': {'铂顿城教学区': 20, '新兆阳广场教学区': 18, '富凯广场教学区': 22, '禅西环宇城教学区': 0, '友邦金融中心教学区': 23, '新南万教学区': 24, '广佛智城教学区': 28, '映月湖环宇城教学区': 24, 'IPARK购物中心教学区': 23, '北滘悦然广场教学区': 43, '大良新一城教学区': 63, '容桂桂洲大道教学区': 68},
    '友邦金融中心教学区': {'铂顿城教学区': 22, '新兆阳广场教学区': 20, '富凯广场教学区': 25, '禅西环宇城教学区': 23, '友邦金融中心教学区': 0, '新南万教学区': 10, '广佛智城教学区': 22, '映月湖环宇城教学区': 12, 'IPARK购物中心教学区': 11, '北滘悦然广场教学区': 35, '大良新一城教学区': 52, '容桂桂洲大道教学区': 58},
    '新南万教学区': {'铂顿城教学区': 23, '新兆阳广场教学区': 20, '富凯广场教学区': 27, '禅西环宇城教学区': 24, '友邦金融中心教学区': 10, '新南万教学区': 0, '广佛智城教学区': 20, '映月湖环宇城教学区': 11, 'IPARK购物中心教学区': 9, '北滘悦然广场教学区': 33, '大良新一城教学区': 50, '容桂桂洲大道教学区': 55},
    '广佛智城教学区': {'铂顿城教学区': 32, '新兆阳广场教学区': 30, '富凯广场教学区': 35, '禅西环宇城教学区': 28, '友邦金融中心教学区': 22, '新南万教学区': 20, '广佛智城教学区': 0, '映月湖环宇城教学区': 25, 'IPARK购物中心教学区': 20, '北滘悦然广场教学区': 38, '大良新一城教学区': 55, '容桂桂洲大道教学区': 62},
    '映月湖环宇城教学区': {'铂顿城教学区': 25, '新兆阳广场教学区': 22, '富凯广场教学区': 20, '禅西环宇城教学区': 24, '友邦金融中心教学区': 12, '新南万教学区': 11, '广佛智城教学区': 25, '映月湖环宇城教学区': 0, 'IPARK购物中心教学区': 13, '北滘悦然广场教学区': 34, '大良新一城教学区': 53, '容桂桂洲大道教学区': 59},
    'IPARK购物中心教学区': {'铂顿城教学区': 24, '新兆阳广场教学区': 20, '富凯广场教学区': 26, '禅西环宇城教学区': 23, '友邦金融中心教学区': 11, '新南万教学区': 9, '广佛智城教学区': 20, '映月湖环宇城教学区': 13, 'IPARK购物中心教学区': 0, '北滘悦然广场教学区': 32, '大良新一城教学区': 50, '容桂桂洲大道教学区': 55},
    '北滘悦然广场教学区': {'铂顿城教学区': 42, '新兆阳广场教学区': 40, '富凯广场教学区': 45, '禅西环宇城教学区': 43, '友邦金融中心教学区': 35, '新南万教学区': 33, '广佛智城教学区': 38, '映月湖环宇城教学区': 34, 'IPARK购物中心教学区': 32, '北滘悦然广场教学区': 0, '大良新一城教学区': 28, '容桂桂洲大道教学区': 35},
    '大良新一城教学区': {'铂顿城教学区': 62, '新兆阳广场教学区': 60, '富凯广场教学区': 58, '禅西环宇城教学区': 63, '友邦金融中心教学区': 52, '新南万教学区': 50, '广佛智城教学区': 55, '映月湖环宇城教学区': 53, 'IPARK购物中心教学区': 50, '北滘悦然广场教学区': 28, '大良新一城教学区': 0, '容桂桂洲大道教学区': 22},
    '容桂桂洲大道教学区': {'铂顿城教学区': 68, '新兆阳广场教学区': 65, '富凯广场教学区': 63, '禅西环宇城教学区': 68, '友邦金融中心教学区': 58, '新南万教学区': 55, '广佛智城教学区': 62, '映月湖环宇城教学区': 59, 'IPARK购物中心教学区': 55, '北滘悦然广场教学区': 35, '大良新一城教学区': 22, '容桂桂洲大道教学区': 0},
}
SHARED_CLASSROOM_DEPTS = {'gaozhi', 'qingshao'}
DEPT_LABELS = {
    'gaozhi': '高中班级部',
    'qingshao': '青少',
    'yiduiyi': '一对一',
}
FIELD_LABELS_PY = {
    'teacher': '授课教师',
    'slot': '时段',
    'timeRange': '上课时间',
    'room': '教室',
    'period': '期数',
    'classType': '班型',
    'campus': '校区',
    'name': '班级名称',
    'create': '新增',
    'delete': '删除',
    'lifecycle_status': '班级状态',
    'currentCount': '当前人数',
    'merge_sources': '合并来源',
    'merged_into_code': '合并至班级',
    'room_occupancy_notice': '教室占用提醒',
}
COURSE_PATCH_FIELDS = {'teacher', 'slot', 'timeRange', 'room', 'period', 'classType'}
COURSE_UPDATE_RETRY_LIMIT = 3
GENERATE_MODES = {
    'spring_to_summer_autumn': {
        'remove_graduating': True,
        'target_seasons': [
            {'season': '暑假', 'periods': ['1期', '2期', '3期'], 'day': '每天'},
            {'season': '秋季', 'periods': ['周五', '周六', '周日'], 'day_from_period': True},
        ],
        'fy_increment': 1,
    },
    'autumn_to_winter_spring': {
        'remove_graduating': False,
        'target_seasons': [
            {'season': '寒假', 'periods': ['1期', '2期'], 'day': '每天'},
            {'season': '春季', 'periods': ['周五', '周六', '周日'], 'day_from_period': True},
        ],
        'fy_increment': 0,
    },
}

os.makedirs(DEPTS_DIR, exist_ok=True)

data_lock = Lock()
auth_lock = Lock()
workflow_lock = Lock()
conflict_status_lock = Lock()
config_lock = Lock()
login_attempt_lock = Lock()
login_attempts = {}
slow_request_lock = Lock()
recent_slow_requests = []
cross_dept_room_index_lock = Lock()
cross_dept_room_index_cache = {}
conflict_suggestions_cache_lock = Lock()
conflict_suggestions_cache = {}
conflict_summary_cache = {}
CONFLICT_SUGGESTIONS_CACHE_MAX = max(1, int(os.environ.get('SCHED_CONFLICT_CACHE_MAX', '24')))
PRESENCE_TTL = 60
MOOD_LEVELS = [
    {'id': 'hang', 'label': '夯', 'tone': '今天排课很顺，能笑着收工'},
    {'id': 'top', 'label': '顶级', 'tone': '整体稳住，只剩几处微调'},
    {'id': 'human', 'label': '人上人', 'tone': '有点压力，但还能掌控'},
    {'id': 'npc', 'label': 'NPC', 'tone': '重复操作变多，脑子开始排队'},
    {'id': 'down', 'label': '拉完了', 'tone': '今天建议先救火，再谈优雅'},
]
MOOD_LEVEL_IDS = {level['id'] for level in MOOD_LEVELS}
MOOD_NOTE_MAX_LENGTH = 42

REQUIRED_SQLITE_STORE_FUNCTIONS = [
    'get_document',
    'set_document',
    'update_document',
    'update_documents',
    'has_document',
    'get_term_version',
    'touch_term_version',
    'set_term_document_and_touch_version',
    'update_presence',
    'clear_presence',
]


def file_fingerprint(path):
    try:
        with open(path, 'rb') as f:
            digest = hashlib.sha1(f.read()).hexdigest()
        st = os.stat(path)
        return {
            'path': path,
            'sha1': digest[:12],
            'size': st.st_size,
            'mtime': time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(st.st_mtime)),
        }
    except OSError as exc:
        return {'path': path, 'error': str(exc)}


def runtime_diagnostics():
    sqlite_functions = {
        name: callable(getattr(sqlite_store, name, None))
        for name in REQUIRED_SQLITE_STORE_FUNCTIONS
    }
    missing_sqlite_functions = [name for name, ok in sqlite_functions.items() if not ok]
    db_parent = os.path.dirname(sqlite_store.DB_PATH) or '.'
    issues = []
    if missing_sqlite_functions:
        issues.append({
            'severity': 'critical',
            'code': 'sqlite_store_mismatch',
            'message': 'sqlite_store.py 与 app.py 版本不匹配，缺少保存所需函数',
            'details': missing_sqlite_functions,
        })
    if not os.path.isdir(db_parent):
        issues.append({
            'severity': 'critical',
            'code': 'database_parent_missing',
            'message': '数据库目录不存在',
            'details': db_parent,
        })
    return {
        'ok': not issues,
        'issues': issues,
        'sqlite_store': {
            'path': getattr(sqlite_store, '__file__', ''),
            'db_path': sqlite_store.DB_PATH,
            'functions': sqlite_functions,
        },
        'code': {
            'app': file_fingerprint(__file__),
            'sqlite_store': file_fingerprint(getattr(sqlite_store, '__file__', '')),
        },
    }


def assert_runtime_dependencies():
    diagnostics = runtime_diagnostics()
    critical = [item for item in diagnostics['issues'] if item.get('severity') == 'critical']
    if critical:
        messages = '; '.join(f"{item['code']}: {item['message']}" for item in critical)
        raise RuntimeError(f'排课系统启动自检失败：{messages}')


if os.environ.get('SCHED_SKIP_STARTUP_CHECK', '0') != '1':
    assert_runtime_dependencies()


# ==================== 数据读写 ====================
def document_namespace_for_path(path):
    full = os.path.abspath(path)
    if full == os.path.abspath(DEPTS_CONFIG):
        return 'config:departments'
    if full == os.path.abspath(USERS_FILE):
        return 'auth:users'
    if full == os.path.abspath(INVITES_FILE):
        return 'auth:invites'
    if full == os.path.abspath(SESSIONS_FILE):
        return 'auth:sessions'

    rel = os.path.relpath(full, BASE_DIR).replace(os.sep, '/')
    m = re.match(r'^departments/([^/]+)/terms\.json$', rel)
    if m:
        return f'dept:{m.group(1)}:terms'
    m = re.match(r'^departments/([^/]+)/resources/(teachers|classrooms|campus_config)\.json$', rel)
    if m:
        return f'dept:{m.group(1)}:{m.group(2)}'
    m = re.match(r'^departments/([^/]+)/terms/([^/]+)/(data|data_original|changelog|metadata|workflow|history_meta|conflict_status)\.json$', rel)
    if m:
        kind_map = {'data_original': 'original'}
        kind = kind_map.get(m.group(3), m.group(3))
        return f'term:{m.group(1)}:{m.group(2)}:{kind}'
    return None


def load_json(path, default):
    namespace = document_namespace_for_path(path)
    if namespace:
        missing = object()
        value = sqlite_store.get_document(namespace, missing)
        if value is not missing:
            return default if value is None else value
    if os.path.exists(path):
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        if namespace:
            sqlite_store.set_document(namespace, data)
        return data
    return default

def save_json(path, data):
    namespace = document_namespace_for_path(path)
    if namespace:
        sqlite_store.set_document(namespace, data)
    write_json_file(path, data)


def write_json_file(path, data):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    tmp = f"{path}.{os.getpid()}.{time.time_ns()}.tmp"
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)


class ConfigUpdateError(Exception):
    def __init__(self, message, status_code=400):
        self.message = message
        self.status_code = status_code


def update_json_document_atomic(path, default, updater):
    namespace = document_namespace_for_path(path)
    with config_lock:
        if not namespace:
            current = load_json(path, default)
            new_data, result = updater(current)
            save_json(path, new_data)
            return result

        def wrapped(current):
            new_data, result = updater(current)
            return new_data, {'data': new_data, 'result': result}, True

        payload = sqlite_store.update_document(namespace, default, wrapped)
        write_json_file(path, payload['data'])
        return payload['result']


def update_json_documents_atomic(defaults_by_path, updater):
    namespaces = {path: document_namespace_for_path(path) for path in defaults_by_path}
    with config_lock:
        if not all(namespaces.values()):
            current = {path: load_json(path, default) for path, default in defaults_by_path.items()}
            new_values, result = updater(current)
            for path, data in (new_values or {}).items():
                save_json(path, data)
            return result

        defaults_by_namespace = {
            namespaces[path]: default
            for path, default in defaults_by_path.items()
        }

        def wrapped(values):
            current_by_path = {
                path: values.get(namespaces[path])
                for path in defaults_by_path
            }
            new_by_path, result = updater(current_by_path)
            new_by_namespace = {
                namespaces[path]: data
                for path, data in (new_by_path or {}).items()
            }
            return new_by_namespace, {
                'data': new_by_path,
                'result': result,
            }

        payload = sqlite_store.update_documents(defaults_by_namespace, wrapped)
        for path, data in (payload.get('data') or {}).items():
            write_json_file(path, data)
        return payload['result']


# ==================== 密码与邀请码 ====================
def hash_password(password, salt=None):
    if salt is None:
        salt = secrets.token_hex(16)
    h = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt.encode('utf-8'), 200000)
    return salt, h.hex()

def verify_password(password, salt, expected_hash):
    _, h = hash_password(password, salt)
    return secrets.compare_digest(h, expected_hash)


def set_user_password(user, password):
    salt, ph = hash_password(password)
    user['salt'] = salt
    user['password_hash'] = ph
    user.pop('password', None)
    return user

INVITE_ALPHABET = '23456789ABCDEFGHJKLMNPQRSTUVWXYZ'
def gen_invite_code(length=12):
    return ''.join(secrets.choice(INVITE_ALPHABET) for _ in range(length))


# ==================== 用户/邀请码/Session ====================
def load_users(): return load_json(USERS_FILE, [])
def save_users(u): save_json(USERS_FILE, u)
def load_invites(): return load_json(INVITES_FILE, [])
def save_invites(i): save_json(INVITES_FILE, i)

class AuthUpdateError(Exception):
    def __init__(self, message, status_code=400):
        self.message = message
        self.status_code = status_code


def update_users_atomic(updater):
    namespace = document_namespace_for_path(USERS_FILE)
    with auth_lock:
        if not namespace:
            users = load_users()
            new_users, result = updater(users)
            save_users(new_users)
            return result

        def wrapped(users):
            new_users, result = updater(users)
            return new_users, {'users': new_users, 'result': result}, True

        payload = sqlite_store.update_document(namespace, [], wrapped)
        write_json_file(USERS_FILE, payload['users'])
        return payload['result']


def update_invites_atomic(updater):
    namespace = document_namespace_for_path(INVITES_FILE)
    with auth_lock:
        if not namespace:
            invites = load_invites()
            new_invites, result = updater(invites)
            save_invites(new_invites)
            return result

        def wrapped(invites):
            new_invites, result = updater(invites)
            return new_invites, {'invites': new_invites, 'result': result}, True

        payload = sqlite_store.update_document(namespace, [], wrapped)
        write_json_file(INVITES_FILE, payload['invites'])
        return payload['result']


def update_users_and_invites_atomic(updater):
    users_ns = document_namespace_for_path(USERS_FILE)
    invites_ns = document_namespace_for_path(INVITES_FILE)
    with auth_lock:
        if not users_ns or not invites_ns:
            users = load_users()
            invites = load_invites()
            new_users, new_invites, result = updater(users, invites)
            save_users(new_users)
            save_invites(new_invites)
            return result

        def wrapped(values):
            new_users, new_invites, result = updater(values.get(users_ns) or [], values.get(invites_ns) or [])
            return {
                users_ns: new_users,
                invites_ns: new_invites,
            }, {
                'users': new_users,
                'invites': new_invites,
                'result': result,
            }

        payload = sqlite_store.update_documents({users_ns: [], invites_ns: []}, wrapped)
        write_json_file(USERS_FILE, payload['users'])
        write_json_file(INVITES_FILE, payload['invites'])
        return payload['result']

def mood_today():
    return (datetime.now(timezone.utc) + timedelta(hours=8)).strftime('%Y-%m-%d')

def mood_namespace(date_text=None):
    return f"shared:mood:{date_text or mood_today()}"

def mood_level_by_id():
    return {level['id']: level for level in MOOD_LEVELS}

def mood_role_label(role):
    return {
        'admin': '管理员',
        'jiaowu': '教务',
        'director': '总监',
        'supervisor': '教学主管',
        'regional_manager': '大区经理',
        'store_manager': '店长',
        'user': '普通用户',
    }.get(role or '', role or '')

def mood_user_profile(user, dept_names=None):
    dept_names = dept_names or {d['id']: d.get('name', d['id']) for d in load_depts_config()}
    dept_id = user.get('dept_id') or ''
    return {
        'user_id': user.get('id') or user.get('email') or '',
        'name': user.get('name') or user.get('email') or '未知用户',
        'role': user.get('role') or 'user',
        'role_label': mood_role_label(user.get('role') or 'user'),
        'campus': user.get('campus') or '',
        'dept_id': dept_id,
        'dept_name': dept_names.get(dept_id, ''),
    }

def normalize_mood_board(raw, date_text=None):
    date_text = date_text or mood_today()
    board = raw if isinstance(raw, dict) else {}
    entries = board.get('entries') if isinstance(board.get('entries'), dict) else {}
    clean_entries = {}
    for user_id, item in entries.items():
        if not isinstance(item, dict):
            continue
        level = item.get('level')
        if level not in MOOD_LEVEL_IDS:
            continue
        clean_entries[str(user_id)] = {
            'user_id': str(user_id),
            'level': level,
            'note': str(item.get('note') or '')[:MOOD_NOTE_MAX_LENGTH],
            'updated_at': item.get('updated_at') or '',
            'name': item.get('name') or '',
            'role': item.get('role') or '',
            'role_label': item.get('role_label') or mood_role_label(item.get('role') or ''),
            'campus': item.get('campus') or '',
            'dept_id': item.get('dept_id') or '',
            'dept_name': item.get('dept_name') or '',
        }
    return {'date': date_text, 'entries': clean_entries}

def mood_board_payload(board, viewer):
    date_text = board.get('date') or mood_today()
    level_map = mood_level_by_id()
    level_rank = {level['id']: index for index, level in enumerate(MOOD_LEVELS)}
    users_by_id = {str(u.get('id') or u.get('email') or ''): u for u in load_users()}
    dept_names = {d['id']: d.get('name', d['id']) for d in load_depts_config()}
    viewer_id = str(viewer.get('id') or viewer.get('email') or '')
    entries = []
    for user_id, item in (board.get('entries') or {}).items():
        if item.get('level') not in level_map:
            continue
        user = users_by_id.get(str(user_id))
        profile = mood_user_profile(user, dept_names) if user else {
            'user_id': str(user_id),
            'name': item.get('name') or '已离开系统的用户',
            'role': item.get('role') or '',
            'role_label': item.get('role_label') or mood_role_label(item.get('role') or ''),
            'campus': item.get('campus') or '',
            'dept_id': item.get('dept_id') or '',
            'dept_name': item.get('dept_name') or '',
        }
        entries.append({
            **profile,
            'level': item.get('level'),
            'level_label': level_map[item.get('level')]['label'],
            'note': item.get('note') or '',
            'updated_at': item.get('updated_at') or '',
            'is_me': str(user_id) == viewer_id,
        })
    entries.sort(key=lambda x: (level_rank.get(x['level'], 99), x.get('updated_at') or '', x.get('name') or ''))
    groups = [
        {
            **level,
            'entries': [entry for entry in entries if entry.get('level') == level['id']],
        }
        for level in MOOD_LEVELS
    ]
    viewer_profile = mood_user_profile(viewer, dept_names) if viewer else None
    return {
        'date': date_text,
        'levels': MOOD_LEVELS,
        'groups': groups,
        'entries': entries,
        'me': next((entry for entry in entries if entry.get('is_me')), None),
        'viewer': viewer_profile,
        'server_time': time.strftime('%Y-%m-%d %H:%M:%S'),
    }

def update_today_mood_for_user(user, level, note):
    date_text = mood_today()
    namespace = mood_namespace(date_text)
    level = str(level or '').strip()
    note = str(note or '').strip()[:MOOD_NOTE_MAX_LENGTH]
    if level not in MOOD_LEVEL_IDS:
        raise AuthUpdateError('请选择一个有效的心情档位', 400)
    profile = mood_user_profile(user)
    user_id = profile['user_id']

    def updater(raw):
        board = normalize_mood_board(raw, date_text)
        entries = dict(board.get('entries') or {})
        entries[user_id] = {
            **profile,
            'level': level,
            'note': note,
            'updated_at': time.strftime('%Y-%m-%d %H:%M:%S'),
        }
        board['entries'] = entries
        return board, mood_board_payload(board, user), True

    return sqlite_store.update_document(namespace, {'date': date_text, 'entries': {}}, updater)

def delete_today_mood_for_user(user):
    date_text = mood_today()
    namespace = mood_namespace(date_text)
    user_id = str(user.get('id') or user.get('email') or '')

    def updater(raw):
        board = normalize_mood_board(raw, date_text)
        entries = dict(board.get('entries') or {})
        entries.pop(user_id, None)
        board['entries'] = entries
        return board, mood_board_payload(board, user), True

    return sqlite_store.update_document(namespace, {'date': date_text, 'entries': {}}, updater)

def load_sessions(): return sqlite_store.list_sessions(ttl=SESSION_TTL)
def save_sessions(s):
    sqlite_store.clear_sessions()
    now = time.time()
    for token, rec in (s or {}).items():
        created_at = float(rec.get('created_at') or now)
        if now - created_at >= SESSION_TTL:
            continue
        sqlite_store.create_session(
            token,
            rec.get('user_id', ''),
            rec.get('csrf_token') or secrets.token_urlsafe(32),
            created_at=created_at,
        )

def find_user_by_email(email):
    email = email.lower().strip()
    for u in load_users():
        if u['email'] == email:
            return u
    return None

def find_user_by_id(uid):
    for u in load_users():
        if u['id'] == uid:
            return u
    return None

def create_session(user_id):
    token = secrets.token_urlsafe(32)
    csrf_token = secrets.token_urlsafe(32)
    sqlite_store.create_session(token, user_id, csrf_token, ttl=SESSION_TTL)
    return token

def destroy_session(token):
    sqlite_store.delete_session(token)

def current_user():
    token = request.cookies.get(SESSION_COOKIE)
    if not token:
        return None
    rec = sqlite_store.get_session(token, ttl=SESSION_TTL)
    if not rec:
        legacy = load_json(SESSIONS_FILE, {}).get(token)
        if not legacy or time.time() - float(legacy.get('created_at') or 0) > SESSION_TTL:
            return None
        rec = {
            'user_id': legacy.get('user_id', ''),
            'created_at': float(legacy.get('created_at') or time.time()),
            'csrf_token': legacy.get('csrf_token') or secrets.token_urlsafe(32),
        }
        sqlite_store.create_session(token, rec['user_id'], rec['csrf_token'], created_at=rec['created_at'], ttl=SESSION_TTL)
    request.session_token = token
    request.session_record = rec
    return find_user_by_id(rec['user_id'])

def current_csrf_token(load_if_missing=True):
    rec = getattr(request, 'session_record', None)
    if not rec and load_if_missing:
        current_user()
        rec = getattr(request, 'session_record', None)
    return rec.get('csrf_token') if rec else None

def can_access_dept_page(user, dept_id):
    if not user:
        return False
    if user.get('role') in STAFF_ROLES:
        return True
    return user.get('dept_id') == dept_id

def record_login_attempt(ip, success):
    now = time.time()
    with login_attempt_lock:
        attempts = login_attempts.get(ip, [])
        attempts = [t for t in attempts if now - t < 60]
        if success:
            attempts = []
        else:
            attempts.append(now)
        login_attempts[ip] = attempts
        return len(attempts)

def count_login_attempts(ip):
    now = time.time()
    with login_attempt_lock:
        attempts = login_attempts.get(ip, [])
        attempts = [t for t in attempts if now - t < 60]
        login_attempts[ip] = attempts
        return len(attempts)

def is_trusted_proxy(remote_addr):
    if not remote_addr:
        return False
    try:
        remote_ip = ipaddress.ip_address(remote_addr)
    except ValueError:
        return False
    for item in TRUSTED_PROXY_IPS:
        try:
            if '/' in item:
                if remote_ip in ipaddress.ip_network(item, strict=False):
                    return True
            elif remote_ip == ipaddress.ip_address(item):
                return True
        except ValueError:
            continue
    return False

def client_ip_for_rate_limit():
    remote_addr = request.remote_addr or ''
    if TRUST_PROXY_HEADERS and is_trusted_proxy(remote_addr):
        forwarded = (request.headers.get('X-Forwarded-For') or '').split(',')[0].strip()
        if forwarded:
            return forwarded
    return remote_addr or 'unknown'


def record_slow_request(duration_ms, status_code):
    if not SLOW_REQUEST_HISTORY_LIMIT:
        return
    item = {
        'time': time.strftime('%Y-%m-%d %H:%M:%S'),
        'method': request.method,
        'path': request.path,
        'status': status_code,
        'duration_ms': duration_ms,
    }
    with slow_request_lock:
        recent_slow_requests.append(item)
        del recent_slow_requests[:-SLOW_REQUEST_HISTORY_LIMIT]


def recent_slow_request_snapshot():
    with slow_request_lock:
        return list(recent_slow_requests)


@app.before_request
def mark_request_start():
    if REQUEST_TIMING_ENABLED:
        request._schedule_started_at = time.perf_counter()
    return None


@app.before_request
def csrf_protect():
    if request.method in SAFE_METHODS or request.path in CSRF_EXEMPT_PATHS:
        return None
    if not (request.path.startswith('/api/') or '/api/' in request.path):
        return None
    token = request.cookies.get(SESSION_COOKIE)
    if not token:
        return None
    u = current_user()
    if not u:
        return None
    sent = request.headers.get(CSRF_HEADER)
    expected = current_csrf_token()
    if not sent or not expected or not secrets.compare_digest(sent, expected):
        return jsonify({'error': 'csrf failed'}), 403
    return None

@app.after_request
def set_csrf_cookie(resp):
    started = getattr(request, '_schedule_started_at', None)
    if started is not None:
        duration_ms = int((time.perf_counter() - started) * 1000)
        resp.headers['X-Response-Time-ms'] = str(duration_ms)
        if duration_ms >= int(SLOW_REQUEST_SECONDS * 1000):
            record_slow_request(duration_ms, resp.status_code)
            app.logger.warning(
                'Slow request %s %s -> %s in %sms',
                request.method,
                request.path,
                resp.status_code,
                duration_ms,
            )
    token = current_csrf_token(load_if_missing=False)
    if token:
        resp.set_cookie(CSRF_COOKIE, token, max_age=SESSION_TTL, httponly=False, samesite='Lax', secure=COOKIE_SECURE)
    if should_gzip_response(resp):
        raw = resp.get_data()
        compressed = gzip.compress(raw)
        if len(compressed) < len(raw):
            resp.set_data(compressed)
            resp.headers['Content-Encoding'] = 'gzip'
            resp.headers['Content-Length'] = str(len(compressed))
            resp.headers['Vary'] = add_vary_header(resp.headers.get('Vary'), 'Accept-Encoding')
    return resp

def add_vary_header(current, value):
    parts = [p.strip() for p in (current or '').split(',') if p.strip()]
    if not any(p.lower() == value.lower() for p in parts):
        parts.append(value)
    return ', '.join(parts)

def should_gzip_response(resp):
    if request.method not in {'GET', 'HEAD'}:
        return False
    if 'gzip' not in (request.headers.get('Accept-Encoding') or '').lower():
        return False
    if resp.direct_passthrough or resp.status_code < 200 or resp.status_code >= 300:
        return False
    if resp.headers.get('Content-Encoding'):
        return False
    mimetype = (resp.mimetype or '').lower()
    if mimetype not in {'application/json', 'text/html', 'text/css', 'application/javascript', 'text/javascript'}:
        return False
    try:
        size = len(resp.get_data())
    except RuntimeError:
        return False
    return size >= GZIP_MIN_BYTES


@app.errorhandler(Exception)
def handle_unexpected_error(exc):
    if isinstance(exc, HTTPException):
        if request.path.startswith('/api/') or '/api/' in request.path:
            return jsonify({
                'error': exc.description,
                'code': exc.name.lower().replace(' ', '_'),
                'path': request.path,
            }), exc.code
        return exc
    app.logger.exception('Unhandled exception: %s', exc)
    if request.path.startswith('/api/') or '/api/' in request.path:
        return jsonify({
            'error': '服务器内部错误，请联系管理员检查日志',
            'code': 'internal_error',
            'path': request.path,
        }), 500
    raise exc


# ==================== 装饰器 ====================
def require_auth(f):
    @wraps(f)
    def wrap(*a, **kw):
        u = current_user()
        if not u:
            if request.path.startswith('/api/') or '/api/' in request.path:
                return jsonify({'error': 'unauthorized'}), 401
            return redirect('/auth')
        request.user = u
        return f(*a, **kw)
    return wrap

def require_admin(f):
    @wraps(f)
    def wrap(*a, **kw):
        u = current_user()
        if not u:
            return jsonify({'error': 'unauthorized'}), 401
        if u.get('role') != 'admin':
            return jsonify({'error': 'forbidden'}), 403
        request.user = u
        return f(*a, **kw)
    return wrap

def require_staff(f):
    @wraps(f)
    def wrap(*a, **kw):
        u = current_user()
        if not u:
            return jsonify({'error': 'unauthorized'}), 401
        if u.get('role') not in STAFF_ROLES:
            return jsonify({'error': 'forbidden'}), 403
        request.user = u
        return f(*a, **kw)
    return wrap

def check_dept_access(dept_id):
    """返回 (user, error_response)，二选一为 None"""
    u = current_user()
    if not u:
        return None, (jsonify({'error': 'unauthorized'}), 401)
    if u.get('role') not in STAFF_ROLES and u.get('dept_id') != dept_id:
        return None, (jsonify({'error': '无权访问该部门'}), 403)
    return u, None

def get_district_campuses(user, dept_id):
    district = user.get('district')
    if not district:
        return []
    return get_campus_config(dept_id).get('districts', {}).get(district, [])

def can_edit_course(user, course, dept_id):
    role = user.get('role')
    if role in STAFF_ROLES:
        return True
    if role in DEPT_MANAGER_ROLES:
        return user.get('dept_id') == dept_id
    if role == 'store_manager':
        return user.get('dept_id') == dept_id and course.get('campus') == user.get('campus')
    return False

def can_create_course(user, campus, dept_id):
    role = user.get('role')
    if role in STAFF_ROLES:
        return True
    if role in DEPT_MANAGER_ROLES:
        return user.get('dept_id') == dept_id
    if role == 'store_manager':
        return user.get('dept_id') == dept_id and campus == user.get('campus')
    return False

def check_workflow_permission(dept_id, term_id, user):
    wf = load_workflow(dept_id, term_id)
    status = wf.get('status', 'draft')
    role = user.get('role')
    allowed_roles = {
        'draft': STAFF_ROLES,
        'scheduling': EDIT_ROLES,
        'reviewing': STAFF_ROLES | DEPT_MANAGER_ROLES,
        'confirmed': STAFF_ROLES,
    }
    if role not in allowed_roles.get(status, set()):
        status_label = {
            'draft': '草稿',
            'scheduling': '排课中',
            'reviewing': '审核中',
            'confirmed': '已确认',
        }.get(status, status)
        return jsonify({'error': f'当前流程状态为「{status_label}」，你的角色无法编辑'}), 403
    return None

def get_time_range(dept_id, slot):
    key = 'qingshao' if dept_id == 'qingshao' else 'default'
    return SLOT_TIME_MAP.get(key, {}).get(slot, '')

def course_lifecycle_status(course):
    return course.get('lifecycle_status') or course.get('course_status') or 'active'

def is_active_course(course):
    return course_lifecycle_status(course) not in {'cancelled', 'merged'}

def lifecycle_status_label(course):
    status = course_lifecycle_status(course)
    if status == 'cancelled':
        return '已取消'
    if status == 'merged':
        code = course.get('merged_into_code') or ''
        return f'已合并至{code}' if code else '已合并'
    return '正常'

def active_courses(courses):
    return [c for c in courses if is_active_course(c)]

def course_time_key(course, slot=None):
    return (
        course.get('season'),
        course.get('period'),
        slot if slot is not None else course.get('slot'),
        course.get('day', ''),
    )

class CourseIndex:
    def __init__(self, courses, related_courses=None, dept_id=''):
        self.courses = courses
        self.related_courses = related_courses or []
        self.dept_id = dept_id
        self.active = [c for c in courses if is_active_course(c)]
        self.related_active = [c for c in self.related_courses if is_active_course(c)]
        self.by_teacher_time = {}
        self.by_day_teacher = {}
        self.by_room_time = {}
        self.suite_courses = []
        self.suite_by_context = {}
        self.low_enrollment = []
        for c in self.active:
            teacher = c.get('teacher')
            if teacher:
                self.by_teacher_time.setdefault((teacher, *course_time_key(c)), []).append(c)
                self.by_day_teacher.setdefault((teacher, c.get('season'), c.get('period'), c.get('day', '')), []).append(c)
            if c.get('campus') and c.get('room') and c.get('slot'):
                self.by_room_time.setdefault((c.get('campus'), c.get('room'), *course_time_key(c)), []).append(c)
                key = course_shared_room_key(c)
                if key:
                    self.by_room_time.setdefault((key, *course_time_key(c)), []).append(c)
            if course_subject(c) in SUITE_SUBJECTS:
                self.suite_courses.append(c)
                context = (c.get('campus'), c.get('season'), c.get('period'), c.get('day', ''), course_grade(c, dept_id))
                self.suite_by_context.setdefault(context, []).append(c)
            if low_enrollment_info(c):
                self.low_enrollment.append(c)
        for c in self.related_active:
            key = c.get('room_key') or course_shared_room_key(c)
            if key and c.get('slot'):
                self.by_room_time.setdefault((key, *course_time_key(c)), []).append(c)
            if c.get('campus') and c.get('room') and c.get('slot'):
                self.by_room_time.setdefault((c.get('campus'), c.get('room'), *course_time_key(c)), []).append(c)

def parse_count_value(value):
    text = str(value or '').strip()
    if not text:
        return None
    m = re.search(r'\d+', text)
    return int(m.group(0)) if m else None

def public_course_ref(course):
    return {
        'id': course.get('id'),
        'code': course.get('code', ''),
        'name': course.get('name', ''),
        'currentCount': course.get('currentCount', ''),
        'campus': course.get('campus', ''),
    }

def format_merge_sources_for_export(course):
    sources = course.get('merge_sources') if isinstance(course, dict) else []
    if not isinstance(sources, list):
        return ''
    rows = []
    for idx, source in enumerate(sources, 1):
        if not isinstance(source, dict):
            continue
        name = ' '.join(str(v).strip() for v in [source.get('code', ''), source.get('name', '')] if str(v or '').strip())
        if not name:
            name = f'来源班{idx}'
        parts = [name]
        if source.get('currentCount'):
            parts.append(f"{source.get('currentCount')}人")
        if source.get('merged_at'):
            parts.append(str(source.get('merged_at')))
        if source.get('merged_by'):
            parts.append(f"操作人：{source.get('merged_by')}")
        if source.get('reason'):
            parts.append(f"原因：{source.get('reason')}")
        rows.append(' '.join(parts))
    return '；'.join(rows)

def room_owner_dept(room):
    text = room or ''
    if re.search(r'(素质教室|青少)', text):
        return 'qingshao'
    if re.search(r'(素养教室|学习机教室|中学|高中|中学部|小组|ZV)', text):
        return 'gaozhi'
    return ''

def shared_room_key(campus, room):
    if not campus or not room:
        return ''
    text = re.sub(r'\s+', '', str(room))
    campus_text = re.sub(r'\s+', '', str(campus))
    owner = room_owner_dept(text)
    room_no = ''
    m = re.search(r'(\d{2,4})(?!.*\d)', text)
    if m:
        room_no = m.group(1)
    else:
        room_no = text.replace(campus_text, '')
        room_no = re.sub(r'(教学区|素养|素质|学习机|教室|临时|共用|借用青少|借用中学|借青少|借中学)', '', room_no)
    return f'{campus_text}|{owner}|{room_no}' if owner and room_no else ''

def course_shared_room_key(course):
    return shared_room_key(course.get('campus', ''), course.get('room', ''))

def term_exists(dept_id, term_id):
    return any(t.get('id') == term_id for t in load_terms(dept_id))

def cross_dept_room_index_signature(term_id):
    signature = []
    for dept in sorted(SHARED_CLASSROOM_DEPTS):
        if not dept_exists(dept) or not term_exists(dept, term_id):
            continue
        tv = sqlite_store.get_term_version(dept, term_id)
        signature.append((dept, int(tv.get('version') or 0)))
    return tuple(signature)

def _build_cross_dept_room_index_uncached(term_id):
    room_index = {}
    for dept in SHARED_CLASSROOM_DEPTS:
        if not dept_exists(dept) or not term_exists(dept, term_id):
            continue
        for c in load_term_data(dept, term_id):
            if not is_active_course(c):
                continue
            key = course_shared_room_key(c)
            if not key:
                continue
            entry = room_index.setdefault(key, {
                'key': key,
                'campus': c.get('campus', ''),
                'owner_dept': key.split('|')[1] if len(key.split('|')) >= 3 else '',
                'room_no': key.split('|')[2] if len(key.split('|')) >= 3 else '',
                'depts': set(),
                'rooms_by_dept': {},
            })
            entry['depts'].add(dept)
            entry['rooms_by_dept'].setdefault(dept, set()).add(c.get('room', ''))
    related = {}
    for key, entry in room_index.items():
        if len(entry['depts']) < 2:
            continue
        related[key] = {
            **entry,
            'depts': sorted(entry['depts']),
            'rooms_by_dept': {d: sorted(v) for d, v in entry['rooms_by_dept'].items()},
        }
    return related

def build_cross_dept_room_index(term_id):
    signature = cross_dept_room_index_signature(term_id)
    cache_key = (term_id, signature)
    with cross_dept_room_index_lock:
        cached = cross_dept_room_index_cache.get(cache_key)
        if cached is not None:
            return copy.deepcopy(cached)
    related = _build_cross_dept_room_index_uncached(term_id)
    with cross_dept_room_index_lock:
        stale_keys = [key for key in cross_dept_room_index_cache if key[0] == term_id and key != cache_key]
        for key in stale_keys:
            cross_dept_room_index_cache.pop(key, None)
        cross_dept_room_index_cache[cache_key] = copy.deepcopy(related)
    return related

def is_shared_classroom(dept_id, term_id, campus, room):
    if dept_id not in SHARED_CLASSROOM_DEPTS:
        return False
    key = shared_room_key(campus, room)
    return bool(key and key in build_cross_dept_room_index(term_id))

def same_course_time(course, season, period, slot, day):
    return (
        course.get('season') == season
        and course.get('period') == period
        and course.get('slot') == slot
        and course.get('day', '') == day
    )

def public_course_conflict(course, dept_id=None):
    out = {
        'id': course.get('id'),
        'code': course.get('code', ''),
        'name': course.get('name', ''),
        'teacher': course.get('teacher', ''),
        'campus': course.get('campus', ''),
        'room': course.get('room', ''),
        'season': course.get('season', ''),
        'period': course.get('period', ''),
        'slot': course.get('slot', ''),
        'day': course.get('day', ''),
        'subject': course.get('subject', ''),
        'grade': course.get('grade', ''),
        'classType': course.get('classType', ''),
        'currentCount': course.get('currentCount', ''),
        'lifecycle_status': course_lifecycle_status(course),
    }
    if dept_id:
        out['dept_id'] = dept_id
        out['dept_label'] = DEPT_LABELS.get(dept_id, dept_id)
    return out

def check_single_conflict(courses, teacher, room, campus, season, period, slot, day, exclude_id=None):
    teacher_conflict = None
    room_conflict = None
    for c in courses:
        if not is_active_course(c):
            continue
        if exclude_id is not None and c.get('id') == exclude_id:
            continue
        if same_course_time(c, season, period, slot, day):
            if teacher and c.get('teacher') == teacher:
                teacher_conflict = c
            if room and campus and c.get('room') == room and c.get('campus') == campus:
                room_conflict = c
    return {'teacher_conflict': teacher_conflict, 'room_conflict': room_conflict}

def check_shared_room_conflict(dept_id, term_id, room, campus, season, period, slot, day):
    if dept_id not in SHARED_CLASSROOM_DEPTS:
        return None
    key = shared_room_key(campus, room)
    if not key:
        return None
    for other_dept in sorted(SHARED_CLASSROOM_DEPTS - {dept_id}):
        if not dept_exists(other_dept) or not term_exists(other_dept, term_id):
            continue
        for c in load_term_data(other_dept, term_id):
            if not is_active_course(c):
                continue
            if course_shared_room_key(c) == key and same_course_time(c, season, period, slot, day):
                return public_course_conflict(c, other_dept)
    return None

def check_course_conflict(dept_id, term_id, courses, teacher, room, campus, season, period, slot, day, exclude_id=None):
    result = check_single_conflict(courses, teacher, room, campus, season, period, slot, day, exclude_id=exclude_id)
    result['shared_room_conflict'] = check_shared_room_conflict(
        dept_id, term_id, room, campus, season, period, slot, day
    )
    return result


def should_soft_allow_room_conflict(payload):
    source = str((payload or {}).get('source') or '').strip()
    return source in {'capacity_drag', 'capacity_drag_undo', 'inline_edit', 'batch_edit'}


def has_room_occupancy_notice(conflict_result):
    return bool(
        (conflict_result or {}).get('room_conflict')
        or (conflict_result or {}).get('shared_room_conflict')
    )


def room_occupancy_notice_text(conflict_result):
    parts = []
    room_conflict = (conflict_result or {}).get('room_conflict')
    if room_conflict:
        label = room_conflict.get('code') or room_conflict.get('name') or room_conflict.get('id') or '占用课程'
        parts.append(f"本部门教室占用：{label}")
    shared_conflict = (conflict_result or {}).get('shared_room_conflict')
    if shared_conflict:
        label = shared_conflict.get('code') or shared_conflict.get('name') or shared_conflict.get('id') or '占用课程'
        dept_label = shared_conflict.get('dept_label') or shared_conflict.get('dept_id') or '其他部门'
        parts.append(f"跨部门教室占用：{dept_label} {label}".strip())
    return '；'.join(parts)


def append_room_occupancy_notices(dept_id, term_id, user, action, reason, notices):
    notices = [n for n in (notices or []) if n and n.get('notice')]
    if not notices:
        return
    timestamp = time.strftime('%Y-%m-%d %H:%M:%S')
    changes = []
    for notice in notices:
        course = notice.get('course') or {}
        changes.append({
            'course_id': course.get('id'),
            'code': course.get('code', ''),
            'name': course.get('name', ''),
            'field': 'room_occupancy_notice',
            'from': '',
            'to': notice.get('notice', ''),
        })
    log = load_changelog(dept_id, term_id)
    log.append({
        'time': timestamp,
        'user': user.get('name') or user.get('email'),
        'email': user.get('email'),
        'action': f"{action}：教室占用后台提醒",
        'reason': reason,
        'changes': changes,
    })
    save_changelog(dept_id, term_id, log[-500:])
    app.logger.info(
        'capacity drag room occupancy notice dept=%s term=%s user=%s count=%s notices=%s',
        dept_id,
        term_id,
        user.get('email'),
        len(changes),
        [c.get('to') for c in changes],
    )

def canonical_campus_name(campus):
    text = str(campus or '').strip()
    if not text:
        return ''
    if text in CAMPUS_TRAVEL_MINUTES:
        return text
    if text in CANONICAL_CAMPUSES:
        return CANONICAL_CAMPUSES[text]
    compact = text.replace('教学区', '').replace('购物中心', '').replace('广场', '').strip()
    if compact in CANONICAL_CAMPUSES:
        return CANONICAL_CAMPUSES[compact]
    for alias, full in CANONICAL_CAMPUSES.items():
        if alias and alias in text:
            return full
    return text

def campus_travel_minutes(from_campus, to_campus):
    a = canonical_campus_name(from_campus)
    b = canonical_campus_name(to_campus)
    if not a or not b:
        return None
    if a == b:
        return 0
    return CAMPUS_TRAVEL_MINUTES.get(a, {}).get(b)

def short_campus_label(campus):
    full = canonical_campus_name(campus)
    for alias, value in CANONICAL_CAMPUSES.items():
        if value == full and len(alias) <= 4:
            return alias
    return (campus or '').replace('教学区', '').replace('购物中心', '').replace('广场', '')

def course_subject(course):
    subject = str(course.get('subject') or '').strip()
    name = str(course.get('name') or '')
    if subject:
        return '博文' if subject == '语文' else subject
    for s in ['双语', '益智', '博文', '语文', '科学', '实践', '数学', '英语', '物理', '化学', 'KET', 'PET', 'YLE']:
        if s in name:
            return '博文' if s == '语文' else s
    return ''

SUITE_SUBJECT_ORDER = ['双语', '益智', '博文', '科学', '实践']
SUITE_SUBJECTS = set(SUITE_SUBJECT_ORDER)
LOW_ENROLLMENT_THRESHOLDS = {'素养': 12, '小组': 5}

def ordered_suite_subjects(subjects):
    subject_set = {s for s in subjects if s}
    ordered = [s for s in SUITE_SUBJECT_ORDER if s in subject_set]
    ordered.extend(sorted(subject_set - set(ordered)))
    return ordered

def suite_coordination_cost_label(move_count, affected_count, teacher_count, residual_count, travel_review):
    if residual_count or move_count >= 4 or teacher_count >= 4 or travel_review:
        return '高'
    if move_count >= 3 or affected_count >= 5 or teacher_count >= 3:
        return '中'
    return '低'

def course_grade(course, dept_id):
    raw = str(course.get('grade') or '').strip()
    name = str(course.get('name') or '')
    if dept_id == 'qingshao':
        if raw in {'幼儿园大班', '一年级', '二年级', '三年级', '四年级', '五年级', '六年级'}:
            return raw
        if raw == 'S3' or '大班' in name or '幼儿园' in name:
            return '幼儿园大班'
        m = re.search(r'([一二三四五六])年级', name)
        if m:
            return m.group(1) + '年级'
        m = re.search(r'([0-6])级', name)
        grade_map = {'0': '幼儿园大班', '1': '一年级', '2': '二年级', '3': '三年级', '4': '四年级', '5': '五年级', '6': '六年级'}
        if m:
            return grade_map.get(m.group(1), '')
        code = str(course.get('code') or '')
        if len(code) >= 4:
            return grade_map.get(code[3], '')
        return ''
    return extract_grade_from_course(course, dept_id)

def course_class_band(course):
    subject = course_subject(course)
    if subject not in {'双语', '益智'}:
        return ''
    raw = str(course.get('classType') or course.get('level') or '').strip()
    if raw in {'A', 'B', 'C'}:
        return raw
    name = re.sub(r'IPARK', '', str(course.get('name') or ''))
    m = re.search(r'[A-C]', name)
    return m.group(0) if m else ''

def suite_compatible(a, b, dept_id):
    if a.get('campus') != b.get('campus'):
        return False
    for key in ('season', 'period'):
        if a.get(key) != b.get(key):
            return False
    if a.get('day', '') != b.get('day', ''):
        return False
    if course_grade(a, dept_id) != course_grade(b, dept_id):
        return False
    a_sub, b_sub = course_subject(a), course_subject(b)
    if a_sub in {'双语', '益智'} and b_sub in {'双语', '益智'}:
        a_band, b_band = course_class_band(a), course_class_band(b)
        return not a_band or not b_band or a_band == b_band
    return True

def slot_order_for(dept_id):
    return list(SLOT_TIME_MAP.get('qingshao' if dept_id == 'qingshao' else 'default', {}).keys())

def schedulable_slots_for_course(course, dept_id):
    slots = slot_order_for(dept_id)
    period = str(course.get('period') or '')
    day = str(course.get('day') or '')
    if '周日' in {period, day}:
        return [s for s in slots if s in {'A', 'B'}]
    return slots

def schedulable_slots_for_period_day(dept_id, period='', day=''):
    return schedulable_slots_for_course({'period': period, 'day': day}, dept_id)

def adjacent_slots(slot, dept_id):
    order = slot_order_for(dept_id)
    if slot not in order:
        return []
    idx = order.index(slot)
    out = []
    if idx > 0:
        out.append(order[idx - 1])
    if idx + 1 < len(order):
        out.append(order[idx + 1])
    return out

def course_label(course):
    return course.get('code') or course.get('name') or str(course.get('id') or '')

def public_conflict_group(group_type, group, dept_id, suggestions=None, cross=False):
    first = group[0] if group else {}
    if group_type == 'teacher':
        label = first.get('teacher', '')
        audience = '主管'
    else:
        label = f"{first.get('room', '')} · {short_campus_label(first.get('campus', ''))}".strip(' ·')
        audience = '店长'
    return {
        'type': group_type,
        'label': label,
        'teacher': first.get('teacher', ''),
        'room': first.get('room', ''),
        'campus': first.get('campus', ''),
        'season': first.get('season', ''),
        'period': first.get('period', ''),
        'slot': first.get('slot', ''),
        'day': first.get('day', ''),
        'audience': audience,
        'cross': cross,
        'classes': [public_course_conflict(c, c.get('dept_id')) for c in group],
        'suggestions': suggestions or [],
    }

def suite_context_courses(courses, course, dept_id):
    index = courses if isinstance(courses, CourseIndex) else None
    if index:
        context = (
            course.get('campus'),
            course.get('season'),
            course.get('period'),
            course.get('day', ''),
            course_grade(course, dept_id),
        )
        return index.suite_by_context.get(context, [])
    return [c for c in active_courses(courses) if course_subject(c) in SUITE_SUBJECTS]

def suite_slot_coverage(courses, course, dept_id, slot, ignore_ids=None):
    ignore_ids = set(ignore_ids or [])
    source_courses = suite_context_courses(courses, course, dept_id)
    moved = dict(course, slot=slot)
    subject = course_subject(moved)
    if subject not in SUITE_SUBJECTS:
        return {'score': 0, 'subjects': [], 'same_slot': [], 'nearby': []}
    same_slot = [
        c for c in source_courses
        if c.get('id') != course.get('id')
        and c.get('id') not in ignore_ids
        and c.get('slot') == slot
        and suite_compatible(moved, c, dept_id)
    ]
    nearby = [
        c for c in source_courses
        if c.get('id') != course.get('id')
        and c.get('id') not in ignore_ids
        and c.get('slot') in adjacent_slots(slot, dept_id)
        and suite_compatible(moved, c, dept_id)
    ]
    subjects = sorted({subject, *(course_subject(c) for c in nearby if course_subject(c) != subject)})
    return {
        'score': len(subjects) if subjects else 1,
        'subjects': subjects or [subject],
        'same_slot': same_slot,
        'nearby': nearby,
    }

def suite_move_note(courses, course, dept_id, target_slot, ignore_ids=None):
    ignore_ids = set(ignore_ids or [])
    subject = course_subject(course)
    if subject not in SUITE_SUBJECTS:
        return {'level': 'neutral', 'score': 0, 'subjects': [], 'text': '非核心套班科目，主要检查教师和教室占用。'}
    target = suite_slot_coverage(courses, course, dept_id, target_slot, ignore_ids=ignore_ids)
    current = suite_slot_coverage(courses, course, dept_id, course.get('slot'), ignore_ids=ignore_ids)
    if target['same_slot']:
        names = '、'.join(course_subject(c) + course_grade(c, dept_id) + course_class_band(c) for c in target['same_slot'][:3])
        return {'level': 'bad', 'score': 0, 'subjects': [], 'text': f'目标时段已有同校区同年级套班课：{names}，可能逼迫家长二选一。'}
    if target['score'] > current['score']:
        subjects = [s for s in target['subjects'] if s != subject]
        return {
            'level': 'good',
            'score': target['score'],
            'subjects': target['subjects'],
            'score_delta': target['score'] - current['score'],
            'text': f'相邻时段可衔接：{"、".join(subjects)}；套班覆盖{"、".join(target["subjects"])}共{target["score"]}科，比当前组合更完整。',
        }
    if target['score'] == current['score'] and target['score'] > 1:
        subjects = [s for s in target['subjects'] if s != subject]
        return {
            'level': 'good',
            'score': target['score'],
            'subjects': target['subjects'],
            'score_delta': 0,
            'text': f'相邻时段可衔接：{"、".join(subjects)}；套班覆盖{"、".join(target["subjects"])}共{target["score"]}科，结构保留度较好。',
        }
    if target['score'] < current['score']:
        return {
            'level': 'warn',
            'score': target['score'],
            'subjects': target['subjects'],
            'score_delta': target['score'] - current['score'],
            'text': f'调整后套班覆盖{"、".join(target["subjects"])}共{target["score"]}科，少于当前{"、".join(current["subjects"])}共{current["score"]}科，建议优先保留更多科目组合。',
        }
    return {'level': 'warn', 'score': 1, 'subjects': [subject], 'score_delta': 0, 'text': '未找到相邻时段的双语/益智/博文/科学/实践套班课，建议主管复核招生搭配。'}

def course_lesson_kind(course):
    text = ' '.join(str(course.get(k) or '') for k in ('classKind', 'class_type', 'desc', 'name', 'room', 'capacity'))
    capacity = parse_count_value(course.get('capacity'))
    if '小组' in text or capacity == 6:
        return '小组'
    if '素养' in text or capacity == 20:
        return '素养'
    return ''

def low_enrollment_info(course):
    kind = course_lesson_kind(course)
    threshold = LOW_ENROLLMENT_THRESHOLDS.get(kind)
    count = parse_count_value(course.get('currentCount'))
    if not kind or threshold is None or count is None or count >= threshold:
        return None
    return {'kind': kind, 'threshold': threshold, 'count': count}

def suite_release_note(courses, course, dept_id):
    subject = course_subject(course)
    if subject not in SUITE_SUBJECTS:
        return {'level': 'good', 'text': '释放班级不是核心套班科目，取消/合并对套班结构影响相对小。'}
    coverage = suite_slot_coverage(courses, course, dept_id, course.get('slot'), ignore_ids={course.get('id')})
    partners = coverage.get('nearby') or []
    if not partners:
        return {'level': 'good', 'text': f'{course_label(course)} 当前未识别到相邻套班课，取消/合并对套班完整性影响相对可控。'}
    labels = '、'.join(f"{course_subject(c)}{course_grade(c, dept_id)}{course_class_band(c)}{c.get('slot', '')}段" for c in partners[:4])
    return {'level': 'warn', 'text': f'{course_label(course)} 与 {labels} 存在套班关联；若取消/合并，需同步复核这些班的续报和搭配。'}

def release_merge_band(course):
    band = course_class_band(course)
    if band:
        return band
    raw = str(course.get('classType') or course.get('level') or '').strip()
    return raw if raw in {'A', 'B', 'C'} else ''

def low_enrollment_merge_targets(courses, source, dept_id, limit=3, exclude_ids=None):
    exclude_ids = set(exclude_ids or [])
    index = courses if isinstance(courses, CourseIndex) else None
    source_courses = index.active if index else active_courses(courses)
    source_count = parse_count_value(source.get('currentCount')) or 0
    source_subject = course_subject(source)
    source_grade = course_grade(source, dept_id)
    source_band = release_merge_band(source)
    source_kind = course_lesson_kind(source)
    candidates = []
    for target in source_courses:
        if target.get('id') == source.get('id'):
            continue
        if target.get('id') in exclude_ids:
            continue
        if target.get('campus') != source.get('campus'):
            continue
        if target.get('season') != source.get('season'):
            continue
        if course_subject(target) != source_subject:
            continue
        if course_grade(target, dept_id) != source_grade:
            continue
        target_band = release_merge_band(target)
        if source_band and target_band and source_band != target_band:
            continue
        target_count = parse_count_value(target.get('currentCount'))
        after_count = (target_count or 0) + source_count
        same_period = target.get('period') == source.get('period')
        same_slot = same_period and target.get('slot') == source.get('slot') and target.get('day', '') == source.get('day', '')
        adjacent = same_period and target.get('slot') in adjacent_slots(source.get('slot'), dept_id)
        suite = suite_release_note(courses, target, dept_id)
        score = 50
        notes = []
        if same_slot:
            score += 16
            notes.append('同时间段合并，学生沟通成本最低')
        elif adjacent:
            score += 8
            notes.append('相邻时段，可作为次优合并目标')
        elif same_period:
            score += 4
            notes.append('同一期数但需调整上课时间')
        else:
            score -= 8
            notes.append('不同期数，合并前需确认学生时间')
        if source_band and target_band == source_band:
            score += 8
            notes.append(f'同{source_band}班型')
        elif source_band and not target_band:
            score -= 3
            notes.append('目标班班型未明确')
        if after_count:
            notes.append(f'合并后约{after_count}人')
        target_low = low_enrollment_info(target)
        if target_low and after_count >= target_low['threshold']:
            score += 6
            notes.append(f'可同时把目标{target_low["kind"]}班提升到保留线以上')
        if suite['level'] == 'warn':
            score -= 4
            notes.append('目标班有套班关联，需复核留存搭配')
        candidates.append({
            'course': public_course_conflict(target, dept_id),
            'target_count': target_count,
            'source_count': source_count,
            'after_count': after_count,
            'same_period': same_period,
            'same_slot': same_slot,
            'same_band': bool(source_band and target_band == source_band),
            'score': max(0, min(100, score)),
            'notes': notes[:4],
            'suite_level': suite['level'],
            'kind': source_kind,
        })
    candidates.sort(key=lambda item: (
        not item['same_slot'],
        -item['score'],
        course_label(item['course']),
    ))
    return candidates[:limit]

def low_enrollment_release_evaluation(courses, source, representative, dept_id, low_info, release_suite, target_suite, merge_targets, travel_profile):
    source_count = low_info.get('count') or 0
    threshold = low_info.get('threshold') or 0
    gap = max(0, threshold - source_count)
    best_merge = merge_targets[0] if merge_targets else None
    target_subjects = target_suite.get('subjects') or []
    checks = []

    release_score = 50
    if gap >= 4:
        release_score += 18
    elif gap >= 2:
        release_score += 10
    else:
        release_score += 4
    if release_suite.get('level') == 'warn':
        release_score -= 8
    if travel_profile.get('level') == 'warn':
        release_score -= 6
    if not merge_targets:
        release_score -= 12

    merge_score = 0
    if best_merge:
        merge_score = int(best_merge.get('score') or 0)
        if best_merge.get('same_slot'):
            merge_score += 8
        elif best_merge.get('same_period'):
            merge_score += 3
        if best_merge.get('same_band'):
            merge_score += 6
        merge_score = max(0, min(100, merge_score))

    withdrawal_risk = '高'
    if best_merge and best_merge.get('same_slot') and best_merge.get('same_band'):
        withdrawal_risk = '中'
    if best_merge and best_merge.get('same_slot') and gap >= 4 and release_suite.get('level') != 'warn':
        withdrawal_risk = '中'
    if not best_merge:
        withdrawal_risk = '高'

    suite_risk = '中' if release_suite.get('level') == 'warn' else '低'
    execution_level = '需主管复核'
    if best_merge and best_merge.get('same_slot') and release_suite.get('level') != 'warn' and travel_profile.get('level') == 'good':
        execution_level = '相对可执行'
    if not best_merge or withdrawal_risk == '高':
        execution_level = '谨慎兜底'

    checks.append({
        'item': '释放教师价值',
        'level': 'good' if gap >= 4 else 'warn',
        'detail': f"{course_label(source)} 当前{source_count}/{threshold}人，释放后可承接 {course_label(representative)}",
    })
    checks.append({
        'item': '合并去向',
        'level': 'good' if best_merge and best_merge.get('same_slot') else ('warn' if best_merge else 'bad'),
        'detail': (
            f"优先并入 {course_label(best_merge.get('course') or {})}，合并后约{best_merge.get('after_count')}人"
            if best_merge else '未找到同校区同年级同科目的明确合并目标'
        ),
    })
    checks.append({
        'item': '套班影响',
        'level': 'warn' if release_suite.get('level') == 'warn' else 'good',
        'detail': release_suite.get('text', ''),
    })
    checks.append({
        'item': '退班/结转风险',
        'level': 'bad' if withdrawal_risk == '高' else 'warn',
        'detail': '取消/合并会改变学生班级归属，需店长确认招生集中和家长接受度',
    })

    review_items = [
        '这是最终兜底方案，应排在调时段、换老师、多科联动协调之后',
        '执行前主管需复核释放教师是否确实能承接目标班',
        '店长需确认被取消/合并班学生去向，避免因合班导致退班',
    ]
    if target_subjects:
        review_items.append(f"目标班当前套班覆盖：{'、'.join(target_subjects)}")
    if release_suite.get('level') == 'warn':
        review_items.append('被释放班本身有套班关联，需同步评估相关科目的留存搭配')

    science_score = max(0, min(100, int((release_score * 0.45) + (merge_score * 0.35) + (len(target_subjects) * 5) - (18 if withdrawal_risk == '高' else 8))))
    return {
        'science_score': science_score,
        'release_score': max(0, min(100, release_score)),
        'merge_score': merge_score,
        'withdrawal_risk': withdrawal_risk,
        'suite_risk': suite_risk,
        'execution_level': execution_level,
        'low_gap': gap,
        'checks': checks,
        'review_items': review_items[:5],
        'summary': f"释放班低于保留线{gap}人；合并目标{'已识别' if best_merge else '未明确'}；退班/结转风险{withdrawal_risk}。",
    }

def is_room_available_at(dept_id, term_id, courses, related_courses, campus, room, season, period, slot, day, exclude_ids=None):
    exclude_ids = set(exclude_ids or [])
    key = shared_room_key(campus, room)
    index = courses if isinstance(courses, CourseIndex) else None
    if index:
        checks = []
        if campus and room:
            checks.extend(index.by_room_time.get((campus, room, season, period, slot, day), []))
        if key:
            checks.extend(index.by_room_time.get((key, season, period, slot, day), []))
        for c in checks:
            if c.get('id') in exclude_ids and not c.get('related'):
                continue
            return False
        return True
    for c in list(courses) + list(related_courses):
        if not is_active_course(c):
            continue
        if c.get('id') in exclude_ids and not c.get('related'):
            continue
        if not same_course_time(c, season, period, slot, day):
            continue
        if c.get('campus') == campus and c.get('room') == room:
            return False
        if key and (c.get('room_key') or course_shared_room_key(c)) == key:
            return False
    return True

def teacher_busy_at(courses, teacher, season, period, slot, day, exclude_ids=None):
    exclude_ids = set(exclude_ids or [])
    if isinstance(courses, CourseIndex):
        return any(
            c.get('id') not in exclude_ids
            for c in courses.by_teacher_time.get((teacher, season, period, slot, day), [])
        )
    for c in courses:
        if not is_active_course(c):
            continue
        if c.get('id') in exclude_ids:
            continue
        if c.get('teacher') == teacher and same_course_time(c, season, period, slot, day):
            return True
    return False

def adjacent_teacher_travel(courses, teacher, dept_id, target_campus, season, period, slot, day):
    travels = []
    source_courses = courses.by_day_teacher.get((teacher, season, period, day), []) if isinstance(courses, CourseIndex) else active_courses(courses)
    for near_slot in adjacent_slots(slot, dept_id):
        for c in source_courses:
            if c.get('teacher') != teacher or not same_course_time(c, season, period, near_slot, day):
                continue
            minutes = campus_travel_minutes(c.get('campus'), target_campus)
            if minutes is not None:
                travels.append({
                    'minutes': minutes,
                    'from_campus': c.get('campus', ''),
                    'slot': near_slot,
                    'course': course_label(c),
                })
    if not travels:
        return {'minutes': 0, 'text': '相邻时段未排课，无跑校压力。'}
    worst = max(travels, key=lambda x: x['minutes'])
    return {
        'minutes': worst['minutes'],
        'text': f"相邻{worst['slot']}段在{short_campus_label(worst['from_campus'])}，车程约{worst['minutes']}分钟。",
    }

def slot_half(slot, dept_id):
    if slot in {'A', 'B'}:
        return '上午'
    if slot in {'C', 'D', 'E'}:
        return '下午/晚上'
    return ''

def compact_campus_sequence(campus_slots):
    sequence = []
    for item in campus_slots:
        campus = canonical_campus_name(item.get('campus'))
        if not campus:
            continue
        if not sequence or sequence[-1]['campus'] != campus:
            sequence.append({'campus': campus, 'slot': item.get('slot'), 'course': item.get('course')})
    return sequence

def travel_route_text(sequence, proposed=None):
    if not sequence:
        return '未识别'
    pieces = []
    for item in sequence:
        slot = item.get('slot') or ''
        campus = short_campus_label(item.get('campus'))
        pieces.append(f"{slot}段@{campus}" if slot else campus)
    route = '→'.join(pieces)
    if len(sequence) == 1 and proposed:
        return f"{route}（仅识别到建议课）"
    return route

def travel_transition_level(minutes):
    if minutes is None:
        return 'unknown'
    if minutes > 30:
        return 'bad'
    if minutes > 20:
        return 'warn'
    return 'good'

def has_returning_campus_pattern(sequence):
    seen = set()
    for item in sequence:
        campus = item['campus']
        if campus in seen:
            return True
        seen.add(campus)
    return False

def route_segment_payload(item, dept_id):
    slot = item.get('slot') or ''
    return {
        'slot': slot,
        'half': slot_half(slot, dept_id),
        'campus': short_campus_label(item.get('campus')),
        'course': item.get('course', ''),
    }

def half_route_payload(campus_slots, dept_id):
    out = []
    for half in ['上午', '下午/晚上']:
        half_items = [item for item in campus_slots if slot_half(item.get('slot'), dept_id) == half]
        if not half_items:
            continue
        sequence = compact_campus_sequence(half_items)
        campuses = [short_campus_label(item.get('campus')) for item in sequence if item.get('campus')]
        out.append({
            'half': half,
            'campus_count': len({item.get('campus') for item in sequence if item.get('campus')}),
            'returning': has_returning_campus_pattern(sequence),
            'route': travel_route_text(sequence),
            'segments': [route_segment_payload(item, dept_id) for item in sequence],
            'campuses': campuses,
        })
    return out

def travel_profile_public_fields(profile):
    return {
        'travel_level': profile.get('level'),
        'travel_minutes': profile.get('max_minutes', 0),
        'travel_route': profile.get('route', ''),
        'travel_segments': profile.get('route_segments', []),
        'travel_transitions': profile.get('transitions', []),
        'travel_halves': profile.get('half_routes', []),
        'travel_flags': profile.get('flags', []),
        'travel_text': profile.get('text', ''),
    }

def teacher_day_travel_profile(courses, teacher, dept_id, season, period, day, proposed=None, exclude_ids=None):
    exclude_ids = set(exclude_ids or [])
    order = slot_order_for(dept_id)
    order_idx = {slot: idx for idx, slot in enumerate(order)}
    campus_slots = []
    source_courses = courses.by_day_teacher.get((teacher, season, period, day), []) if isinstance(courses, CourseIndex) else active_courses(courses)
    for c in source_courses:
        if c.get('id') in exclude_ids:
            continue
        if c.get('season') != season or c.get('period') != period or c.get('day', '') != day:
            continue
        if c.get('slot') not in order_idx:
            continue
        campus_slots.append({'slot': c.get('slot'), 'campus': c.get('campus'), 'course': course_label(c)})
    if proposed and proposed.get('slot') in order_idx:
        campus_slots.append(proposed)
    campus_slots.sort(key=lambda item: order_idx.get(item.get('slot'), 999))
    sequence = compact_campus_sequence(campus_slots)
    unique_campuses = {item['campus'] for item in sequence}
    half_map = {}
    for item in campus_slots:
        half = slot_half(item.get('slot'), dept_id)
        campus = canonical_campus_name(item.get('campus'))
        if half and campus:
            half_map.setdefault(half, set()).add(campus)
    transitions = []
    for prev, cur in zip(sequence, sequence[1:]):
        minutes = campus_travel_minutes(prev['campus'], cur['campus'])
        transitions.append({
            'from': prev['campus'],
            'to': cur['campus'],
            'from_label': short_campus_label(prev['campus']),
            'to_label': short_campus_label(cur['campus']),
            'from_slot': prev.get('slot'),
            'to_slot': cur.get('slot'),
            'minutes': minutes,
            'level': travel_transition_level(minutes),
        })
    returning = has_returning_campus_pattern(sequence)
    multi_half = {half: campuses for half, campuses in half_map.items() if len(campuses) > 1}
    max_minutes = max([t['minutes'] or 0 for t in transitions] or [0])
    route = travel_route_text(sequence, proposed=proposed)
    flags = []
    if len(unique_campuses) >= 3 or returning:
        level = 'bad'
        text = f"教师当天校区动线为 {route}，存在多校区/折返跑课风险。"
        if len(unique_campuses) >= 3:
            flags.append({'level': 'bad', 'label': '一天多校区', 'detail': f'当天涉及{len(unique_campuses)}个校区，教师精力消耗高'})
        if returning:
            flags.append({'level': 'bad', 'label': '折返跑课', 'detail': '出现 A-B-A 式校区折返，不建议作为常规方案'})
    elif max_minutes > 30:
        level = 'bad'
        text = f"教师当天校区动线为 {route}；相邻校区车程最高约{max_minutes}分钟，超过可接受跑校范围。"
        flags.append({'level': 'bad', 'label': '车程超限', 'detail': f'相邻时段最高约{max_minutes}分钟，超过30分钟上限'})
    elif multi_half:
        level = 'warn'
        halves = '、'.join(multi_half.keys())
        text = f"教师当天校区动线为 {route}；{halves}内跨校区，建议主管评估体力消耗和课间交通。"
        flags.append({'level': 'warn', 'label': '半天跨校区', 'detail': f'{halves}内跨校区，需确认课间衔接'})
    elif max_minutes > 20:
        level = 'warn'
        text = f"教师当天校区动线为 {route}；相邻校区车程最高约{max_minutes}分钟，必要时需微调上课时间。"
        flags.append({'level': 'warn', 'label': '需微调时间', 'detail': f'相邻时段最高约{max_minutes}分钟，可考虑推迟约10分钟'})
    else:
        level = 'good'
        text = f"教师当天动线 {route}，未形成明显折返跑课。"
        flags.append({'level': 'good', 'label': '动线可接受', 'detail': '未发现明显折返或远距离跑校'})
    return {
        'level': level,
        'text': text,
        'campus_count': len(unique_campuses),
        'returning': returning,
        'max_minutes': max_minutes,
        'route': route,
        'route_segments': [
            route_segment_payload(item, dept_id)
            for item in sequence
        ],
        'half_routes': half_route_payload(campus_slots, dept_id),
        'transitions': transitions,
        'flags': flags,
        'sequence': sequence,
    }

def suggestion_sort_key(suggestion):
    risk_rank = {'低': 0, '中': 1, '需复核': 2, '需协调': 3}
    travel_rank = {'good': 0, 'warn': 1, 'bad': 2}
    category_rank = {
        'teacher_substitute': 0,
        'teacher_time': 1,
        'coordinated_swap': 2,
        'suite_reflow': 3,
        'suite_coordination': 4,
        'low_enrollment_release': 5,
        'room_swap': 0,
    }
    return (
        category_rank.get(suggestion.get('category'), 5),
        risk_rank.get(suggestion.get('risk'), 9),
        travel_rank.get(suggestion.get('travel_level'), 9),
        -int(suggestion.get('priority_score', 0) or 0),
        suggestion.get('travel_minutes', 0),
        -suggestion.get('suite_delta', 0),
        -suggestion.get('suite_score', 0),
        str(suggestion.get('title') or ''),
    )

def score_suggestion(suggestion):
    score = 50
    reasons = []
    tradeoffs = []
    risk = suggestion.get('risk')
    category = suggestion.get('category')
    travel_level = suggestion.get('travel_level')
    travel_minutes = int(suggestion.get('travel_minutes') or suggestion.get('route_travel_minutes') or 0)
    suite_score = int(suggestion.get('suite_score') or 0)
    suite_delta = int(suggestion.get('suite_delta') or 0)

    if risk == '低':
        score += 18
        reasons.append('低风险，可直接作为优先方案复核')
    elif risk == '中':
        score += 6
        tradeoffs.append('存在中等风险，需确认教师产能或家长沟通影响')
    elif risk == '需复核':
        score -= 8
        tradeoffs.append('需要主管复核后执行')
    elif risk == '需协调':
        score -= 15
        tradeoffs.append('需要线下协调，不建议直接应用')

    if category == 'teacher_substitute':
        score += 14
        reasons.append('不改变班级上课时间，家长沟通成本最低')
    elif category == 'teacher_time':
        score += 10
        reasons.append('只调整单门课时间，执行成本较低')
    elif category == 'coordinated_swap':
        score += 6
        tradeoffs.append('涉及两门课联动，需要同时确认两位老师')
    elif category == 'suite_reflow':
        score += 3
        tradeoffs.append('涉及多门套班重排，需要主管确认完整方案')
    elif category == 'low_enrollment_release':
        score -= 4
        tradeoffs.append('依赖取消/合并低人数班，作为兜底方案')
    elif category == 'room_swap':
        score += 12
        reasons.append('只调整教室，不改变上课时间')

    if travel_level == 'good':
        score += 10
        reasons.append('教师动线压力低')
    elif travel_level == 'warn':
        score -= 4
        tradeoffs.append('跨校区耗时接近上限，需确认课间衔接')
    elif travel_level == 'bad':
        score -= 25
        tradeoffs.append('教师动线疲劳风险高')
    if travel_minutes:
        if travel_minutes <= 20:
            score += 6
            reasons.append(f'跨校区约{travel_minutes}分钟，在课间可控范围内')
        elif travel_minutes <= 30:
            score -= 3
            tradeoffs.append(f'跨校区约{travel_minutes}分钟，可能需要微调上下课时间')
        else:
            score -= 18
            tradeoffs.append(f'跨校区约{travel_minutes}分钟，不推荐作为常规方案')

    if suite_score >= 3:
        score += 10
        reasons.append(f'调整后套班覆盖{suite_score}科，留存搭配较完整')
    elif suite_score == 2:
        score += 5
        reasons.append('调整后仍保留两科套班搭配')
    elif suite_score == 1 and category not in {'room_swap', 'teacher_substitute'}:
        tradeoffs.append('套班覆盖较弱，需确认招生/留存影响')
    if suite_delta > 0:
        score += min(8, suite_delta * 4)
        reasons.append(f'套班完整度提升{suite_delta}科')
    elif suite_delta < 0:
        score += max(-10, suite_delta * 5)
        tradeoffs.append(f'套班完整度下降{abs(suite_delta)}科')

    plan_len = len(suggestion.get('plan') or [])
    if plan_len >= 3:
        score -= 4
        tradeoffs.append(f'涉及{plan_len}步联动调整，执行前需逐项核对')
    elif plan_len == 2:
        tradeoffs.append('涉及2步联动调整')

    if suggestion.get('release_count') is not None:
        score -= 2
        tradeoffs.append(f"低人数班{suggestion.get('release_count')}人需先确认取消/合并")

    return {
        'score': max(0, min(100, score)),
        'reasons': reasons[:4],
        'tradeoffs': tradeoffs[:4],
    }

def attach_suggestion_scores(suggestions):
    for suggestion in suggestions:
        scored = score_suggestion(suggestion)
        suggestion['priority_score'] = scored['score']
        suggestion['score_reasons'] = scored['reasons']
        suggestion['tradeoffs'] = scored['tradeoffs']
    return suggestions

def course_swap_suggestions(dept_id, term_id, courses, related_courses, group):
    suggestions = []
    slots = slot_order_for(dept_id)
    group_ids = {c.get('id') for c in group}
    index = courses if isinstance(courses, CourseIndex) else None
    source_courses = index.suite_courses if index else [c for c in active_courses(courses) if course_subject(c) in SUITE_SUBJECTS]
    for c in group:
        c_allowed_slots = set(schedulable_slots_for_course(c, dept_id))
        candidates = [
            x for x in source_courses
            if x.get('id') != c.get('id')
            and x.get('id') not in group_ids
            and x.get('slot') in slots
            and x.get('slot') in c_allowed_slots
            and c.get('slot') in schedulable_slots_for_course(x, dept_id)
            and suite_compatible(c, x, dept_id)
        ]
        candidates.sort(key=lambda x: (
            0 if course_subject(x) != course_subject(c) else 1,
            slots.index(x.get('slot')) if x.get('slot') in slots else 99,
            course_label(x),
        ))
        for other in candidates[:8]:
            c_target_slot = other.get('slot')
            other_target_slot = c.get('slot')
            exclude_ids = {c.get('id'), other.get('id')}
            if teacher_busy_at(courses, c.get('teacher'), c.get('season'), c.get('period'), c_target_slot, c.get('day', ''), exclude_ids=exclude_ids):
                continue
            if teacher_busy_at(courses, other.get('teacher'), other.get('season'), other.get('period'), other_target_slot, other.get('day', ''), exclude_ids=exclude_ids):
                continue
            if not is_room_available_at(dept_id, term_id, courses, related_courses, c.get('campus'), c.get('room'), c.get('season'), c.get('period'), c_target_slot, c.get('day', ''), exclude_ids=exclude_ids):
                continue
            if not is_room_available_at(dept_id, term_id, courses, related_courses, other.get('campus'), other.get('room'), other.get('season'), other.get('period'), other_target_slot, other.get('day', ''), exclude_ids=exclude_ids):
                continue
            c_suite = suite_move_note(courses, c, dept_id, c_target_slot, ignore_ids=exclude_ids)
            other_suite = suite_move_note(courses, other, dept_id, other_target_slot, ignore_ids=exclude_ids)
            if c_suite['level'] == 'bad' or other_suite['level'] == 'bad':
                continue
            c_travel = teacher_day_travel_profile(
                courses, c.get('teacher'), dept_id, c.get('season'), c.get('period'), c.get('day', ''),
                proposed={'slot': c_target_slot, 'campus': c.get('campus'), 'course': course_label(c)},
                exclude_ids={c.get('id')},
            )
            other_travel = teacher_day_travel_profile(
                courses, other.get('teacher'), dept_id, other.get('season'), other.get('period'), other.get('day', ''),
                proposed={'slot': other_target_slot, 'campus': other.get('campus'), 'course': course_label(other)},
                exclude_ids={other.get('id')},
            )
            if c_travel['level'] == 'bad' or other_travel['level'] == 'bad':
                continue
            suggestions.append({
                'category': 'coordinated_swap',
                'audience': '主管',
                'risk': '中' if 'warn' in {c_suite['level'], other_suite['level'], c_travel['level'], other_travel['level']} else '低',
                'title': f"联动交换 {course_label(c)} 与 {course_label(other)} 的时段",
                'detail': (
                    f"{course_label(c)} 调到 {c_target_slot}段，{course_label(other)} 调到 {other_target_slot}段；"
                    f"两位老师目标时段均空闲，教室可用。{c_suite['text']} {other_suite['text']} "
                    f"{c_travel['text']} {other_travel['text']}"
                ),
                'course_id': c.get('id'),
                'swap_with_id': other.get('id'),
                'target_slot': c_target_slot,
                'swap_target_slot': other_target_slot,
                'suite_score': c_suite.get('score', 0) + other_suite.get('score', 0),
                'suite_delta': c_suite.get('score_delta', 0) + other_suite.get('score_delta', 0),
                'travel_level': 'warn' if 'warn' in {c_travel['level'], other_travel['level']} else 'good',
                'travel_minutes': max(c_travel.get('max_minutes', 0), other_travel.get('max_minutes', 0)),
                'travel_routes': [
                    {'teacher': c.get('teacher', ''), **travel_profile_public_fields(c_travel)},
                    {'teacher': other.get('teacher', ''), **travel_profile_public_fields(other_travel)},
                ],
            })
    attach_suggestion_scores(suggestions)
    suggestions.sort(key=suggestion_sort_key)
    return suggestions[:3]

def virtual_courses_with_moves(courses, moves):
    source_courses = courses.courses if isinstance(courses, CourseIndex) else courses
    moved = [dict(c) for c in source_courses]
    move_map = {m['id']: m for m in moves if m.get('id') is not None}
    for c in moved:
        move = move_map.get(c.get('id'))
        if not move:
            continue
        c['slot'] = move['to_slot']
        c['timeRange'] = get_time_range(move.get('dept_id', ''), move['to_slot']) or c.get('timeRange', '')
    if isinstance(courses, CourseIndex):
        return CourseIndex(moved, courses.related_courses, courses.dept_id)
    return moved

def suite_same_slot_blockers(courses, course, dept_id, target_slot, ignore_ids=None):
    ignore_ids = set(ignore_ids or [])
    source_courses = suite_context_courses(courses, course, dept_id)
    moved = dict(course, slot=target_slot)
    return [
        x for x in source_courses
        if x.get('id') != course.get('id')
        and x.get('id') not in ignore_ids
        and x.get('slot') == target_slot
        and suite_compatible(moved, x, dept_id)
    ]

def suite_relocation_slot_order(dept_id, preferred_slot, fallback_slot):
    out = []
    for slot in adjacent_slots(preferred_slot, dept_id) + [fallback_slot] + slot_order_for(dept_id):
        if slot and slot not in out:
            out.append(slot)
    return out

def find_suite_relocation(dept_id, term_id, courses, related_courses, blocker, target_slot, fallback_slot, planned_moves, blocker_ids):
    blocked_slots = {target_slot}
    existing_move_ids = {m['id'] for m in planned_moves}
    for candidate_slot in suite_relocation_slot_order(dept_id, target_slot, fallback_slot):
        if candidate_slot in blocked_slots or candidate_slot == blocker.get('slot'):
            continue
        if candidate_slot not in schedulable_slots_for_course(blocker, dept_id):
            continue
        trial_moves = planned_moves + [{'id': blocker.get('id'), 'to_slot': candidate_slot, 'dept_id': dept_id}]
        trial_courses = virtual_courses_with_moves(courses, trial_moves)
        exclude_ids = set(existing_move_ids) | {blocker.get('id')}
        if teacher_busy_at(
            trial_courses,
            blocker.get('teacher'),
            blocker.get('season'),
            blocker.get('period'),
            candidate_slot,
            blocker.get('day', ''),
            exclude_ids=exclude_ids,
        ):
            continue
        if not is_room_available_at(
            dept_id,
            term_id,
            trial_courses,
            related_courses,
            blocker.get('campus'),
            blocker.get('room'),
            blocker.get('season'),
            blocker.get('period'),
            candidate_slot,
            blocker.get('day', ''),
            exclude_ids={blocker.get('id')},
        ):
            continue
        suite = suite_move_note(trial_courses, blocker, dept_id, candidate_slot, ignore_ids={blocker.get('id')})
        if suite['level'] == 'bad':
            continue
        travel = teacher_day_travel_profile(
            trial_courses,
            blocker.get('teacher'),
            dept_id,
            blocker.get('season'),
            blocker.get('period'),
            blocker.get('day', ''),
            proposed={'slot': candidate_slot, 'campus': blocker.get('campus'), 'course': course_label(blocker)},
            exclude_ids={blocker.get('id')},
        )
        if travel['level'] == 'bad':
            continue
        return {'slot': candidate_slot, 'suite': suite, 'travel': travel}
    return None

def suite_reflow_validation(dept_id, term_id, courses, related_courses, focus_course, planned_moves):
    original_index = courses if isinstance(courses, CourseIndex) else CourseIndex(courses, related_courses, dept_id)
    final_courses = virtual_courses_with_moves(courses, planned_moves)
    final_index = final_courses if isinstance(final_courses, CourseIndex) else CourseIndex(final_courses, related_courses, dept_id)
    move_map = {m.get('id'): m for m in planned_moves if m.get('id') is not None}
    affected_slots = set()
    for move in planned_moves:
        if move.get('from_slot'):
            affected_slots.add(move.get('from_slot'))
        if move.get('to_slot'):
            affected_slots.add(move.get('to_slot'))
    for slot in list(affected_slots):
        affected_slots.update(adjacent_slots(slot, dept_id))
    context = (
        focus_course.get('campus'),
        focus_course.get('season'),
        focus_course.get('period'),
        focus_course.get('day', ''),
        course_grade(focus_course, dept_id),
    )
    order = slot_order_for(dept_id)
    order_idx = {slot: idx for idx, slot in enumerate(order)}

    def context_subset(index):
        return [
            c for c in index.suite_by_context.get(context, [])
            if c.get('id') in move_map or c.get('slot') in affected_slots
        ]

    def slot_snapshot(context_courses):
        out = []
        for slot in order:
            slot_courses = [c for c in context_courses if c.get('slot') == slot]
            if not slot_courses:
                continue
            out.append({
                'slot': slot,
                'subjects': sorted({course_subject(c) for c in slot_courses if course_subject(c)}),
                'courses': [
                    {
                        'course_id': c.get('id'),
                        'course_label': course_label(c),
                        'subject': course_subject(c),
                        'teacher': c.get('teacher', ''),
                        'band': course_class_band(c),
                        'moved': c.get('id') in move_map,
                        'original_slot': move_map.get(c.get('id'), {}).get('from_slot', c.get('slot')),
                        'final_slot': c.get('slot'),
                    }
                    for c in sorted(slot_courses, key=lambda x: (course_subject(x), course_label(x)))
                ],
            })
        return out

    def same_slot_conflicts(context_courses):
        out = []
        for slot in order:
            slot_courses = [c for c in context_courses if c.get('slot') == slot]
            for i, left in enumerate(slot_courses):
                for right in slot_courses[i + 1:]:
                    if suite_compatible(left, right, dept_id):
                        out.append({
                            'slot': slot,
                            'left': public_course_conflict(left, dept_id),
                            'right': public_course_conflict(right, dept_id),
                            'reason': '同校区同年级同班型套班课仍在同一时段，可能造成学生二选一',
                        })
        return out

    original_context_courses = context_subset(original_index)
    final_context_courses = context_subset(final_index)
    affected = []
    for c in sorted(final_context_courses, key=lambda x: (order_idx.get(x.get('slot'), 99), course_subject(x), course_label(x))):
        move = move_map.get(c.get('id'), {})
        ref = public_course_conflict(c, dept_id)
        ref.update({
            'original_slot': move.get('from_slot', c.get('slot')),
            'final_slot': move.get('to_slot', c.get('slot')),
            'moved': c.get('id') in move_map,
            'band': course_class_band(c),
        })
        affected.append(ref)

    original_subjects = ordered_suite_subjects({course_subject(c) for c in original_context_courses if course_subject(c)})
    covered_subjects = ordered_suite_subjects({course_subject(c) for c in final_context_courses if course_subject(c)})
    original_residual = same_slot_conflicts(original_context_courses)
    residual = same_slot_conflicts(final_context_courses)
    hard_checks = []
    final_by_id = {c.get('id'): c for c in final_index.active}
    for move in planned_moves:
        moved_course = final_by_id.get(move.get('id')) or move.get('course') or {}
        teacher_clear = not teacher_busy_at(
            final_index,
            moved_course.get('teacher'),
            moved_course.get('season'),
            moved_course.get('period'),
            moved_course.get('slot'),
            moved_course.get('day', ''),
            exclude_ids={moved_course.get('id')},
        )
        room_clear = is_room_available_at(
            dept_id,
            term_id,
            final_index,
            related_courses,
            moved_course.get('campus'),
            moved_course.get('room'),
            moved_course.get('season'),
            moved_course.get('period'),
            moved_course.get('slot'),
            moved_course.get('day', ''),
            exclude_ids={moved_course.get('id')},
        )
        hard_checks.append({
            'course_id': moved_course.get('id'),
            'course_label': course_label(moved_course),
            'slot': moved_course.get('slot', ''),
            'teacher_clear': teacher_clear,
            'room_clear': room_clear,
            'passed': bool(teacher_clear and room_clear),
        })

    warn_points = [
        m for m in planned_moves
        if m.get('suite', {}).get('level') == 'warn' or m.get('travel', {}).get('level') == 'warn'
    ]
    teacher_impacts = []
    teacher_keys = []
    for move in planned_moves:
        course = move.get('course') or {}
        key = (
            course.get('teacher'),
            course.get('season'),
            course.get('period'),
            course.get('day', ''),
        )
        if key[0] and key not in teacher_keys:
            teacher_keys.append(key)
    for teacher, season, period, day in teacher_keys:
        day_courses = final_index.by_day_teacher.get((teacher, season, period, day), [])
        travel = teacher_day_travel_profile(final_index, teacher, dept_id, season, period, day)
        teacher_impacts.append({
            'teacher': teacher,
            'season': season,
            'period': period,
            'day': day,
            'travel_level': travel.get('level'),
            'travel_minutes': travel.get('max_minutes', 0),
            'travel_route': travel.get('route', ''),
            'travel_segments': travel.get('route_segments', []),
            'travel_transitions': travel.get('transitions', []),
            'travel_halves': travel.get('half_routes', []),
            'travel_flags': travel.get('flags', []),
            'campus_count': travel.get('campus_count', 0),
            'returning': travel.get('returning', False),
            'moved_course_ids': [
                m.get('id') for m in planned_moves
                if (m.get('course') or {}).get('teacher') == teacher
                and (m.get('course') or {}).get('season') == season
                and (m.get('course') or {}).get('period') == period
                and (m.get('course') or {}).get('day', '') == day
            ],
            'courses': [
                {
                    'course_id': c.get('id'),
                    'course_label': course_label(c),
                    'subject': course_subject(c),
                    'campus': c.get('campus', ''),
                    'slot': c.get('slot', ''),
                    'moved': c.get('id') in move_map,
                    'original_slot': move_map.get(c.get('id'), {}).get('from_slot', c.get('slot')),
                }
                for c in sorted(day_courses, key=lambda x: (order_idx.get(x.get('slot'), 99), course_label(x)))
            ],
        })

    failed_hard_checks = [x for x in hard_checks if not x['passed']]
    travel_review = any(t.get('travel_level') in {'warn', 'bad'} for t in teacher_impacts)
    residual_count = len(residual)
    coverage_delta = len(covered_subjects) - len(original_subjects)
    move_count = len(planned_moves)
    teacher_count = len(teacher_impacts)
    communication_cost = suite_coordination_cost_label(
        move_count,
        len(affected),
        teacher_count,
        residual_count,
        travel_review,
    )
    subject_coverage_level = '强' if len(covered_subjects) >= 4 else ('中' if len(covered_subjects) >= 3 else '弱')
    missing_bands = [
        course_label(c) for c in final_context_courses
        if course_subject(c) in {'双语', '益智'} and not course_class_band(c)
    ]
    review_items = []
    if residual_count:
        review_items.append(f'仍有{residual_count}个同段套班冲突，需要主管重新协调时段')
    if coverage_delta < 0:
        review_items.append(f'套班覆盖科目减少{abs(coverage_delta)}科，需评估续报/留存影响')
    if failed_hard_checks:
        review_items.append('存在教师或教室硬约束未通过，方案不能直接执行')
    if travel_review:
        review_items.append('部分老师动线接近或超过跑校上限，需复核体力消耗和课间衔接')
    if len(planned_moves) >= 3:
        review_items.append(f'涉及{len(planned_moves)}门课联动，执行前需逐项确认老师、教室、家长沟通')
    if len(covered_subjects) < 3:
        review_items.append('套班覆盖不足3科，建议继续寻找同校区同年级其他核心科目形成更完整组合')
    if missing_bands:
        review_items.append(f"{'、'.join(missing_bands[:3])} 班型未明确，双语/益智搭班需人工确认")
    if not review_items:
        review_items.append('系统未发现硬性阻断，主管重点确认执行沟通即可')

    science_score = 100
    science_score -= residual_count * 20
    science_score -= len(failed_hard_checks) * 30
    science_score -= sum(18 if t.get('travel_level') == 'bad' else 8 for t in teacher_impacts if t.get('travel_level') in {'warn', 'bad'})
    science_score -= max(0, -coverage_delta) * 10
    science_score -= max(0, len(planned_moves) - 2) * 4
    science_score -= {'低': 0, '中': 5, '高': 12}.get(communication_cost, 0)
    if len(covered_subjects) >= 4:
        science_score += 6
    elif len(covered_subjects) < 3:
        science_score -= 8
    science_score = max(0, min(100, science_score))
    summary_tail = '未发现残余同段套班冲突' if residual_count == 0 else f'仍有{residual_count}个同段冲突需要人工协调'
    confidence = '高' if science_score >= 85 and not failed_hard_checks and not residual_count else ('中' if science_score >= 65 else '需人工重排')
    execution_level = '可优先复核'
    if failed_hard_checks or residual_count:
        execution_level = '需重新推演'
    elif communication_cost == '高' or travel_review:
        execution_level = '需主管深度复核'
    elif communication_cost == '中' or warn_points:
        execution_level = '需主管复核'
    coordination_summary = (
        f"覆盖{len(covered_subjects)}科（{'、'.join(covered_subjects) or '未识别'}），"
        f"联动{move_count}门课、影响{len(affected)}个班、{teacher_count}位老师；"
        f"沟通成本{communication_cost}，执行等级：{execution_level}。"
    )
    return {
        'context': {
            'campus': focus_course.get('campus', ''),
            'season': focus_course.get('season', ''),
            'period': focus_course.get('period', ''),
            'day': focus_course.get('day', ''),
            'grade': course_grade(focus_course, dept_id),
        },
        'affected_count': len(affected),
        'affected_courses': affected,
        'original_subjects': original_subjects,
        'covered_subjects': covered_subjects,
        'coverage_score': len(covered_subjects),
        'coverage_delta': coverage_delta,
        'subject_coverage_level': subject_coverage_level,
        'coordination_summary': coordination_summary,
        'communication_cost': communication_cost,
        'execution_level': execution_level,
        'coordination_kpis': {
            'move_count': move_count,
            'affected_count': len(affected),
            'teacher_count': teacher_count,
            'covered_subject_count': len(covered_subjects),
            'residual_conflict_count': residual_count,
        },
        'before_slots': slot_snapshot(original_context_courses),
        'slots': slot_snapshot(final_context_courses),
        'teacher_impacts': teacher_impacts,
        'hard_checks': hard_checks,
        'system_checks': [
            {'item': '教师时间冲突', 'passed': all(x['teacher_clear'] for x in hard_checks)},
            {'item': '教室占用冲突', 'passed': all(x['room_clear'] for x in hard_checks)},
            {'item': '套班同段冲突', 'passed': residual_count == 0, 'before_count': len(original_residual), 'after_count': residual_count},
            {'item': '教师跑校动线', 'passed': not travel_review},
        ],
        'residual_conflict_count': residual_count,
        'residual_conflicts': residual[:5],
        'manual_review': bool(residual_count or warn_points or travel_review or failed_hard_checks),
        'review_items': review_items,
        'science_score': science_score,
        'confidence': confidence,
        'summary': f"系统已验证{len(affected)}门受影响班级、{len(teacher_impacts)}位老师动线，调整后覆盖{'、'.join(covered_subjects) or '未识别'}；{summary_tail}。",
    }

def suite_reflow_suggestions(dept_id, term_id, courses, related_courses, group):
    suggestions = []
    slots = slot_order_for(dept_id)
    group_ids = {c.get('id') for c in group}
    for c in group:
        if course_subject(c) not in SUITE_SUBJECTS:
            continue
        for target_slot in schedulable_slots_for_course(c, dept_id):
            if target_slot == c.get('slot'):
                continue
            if teacher_busy_at(courses, c.get('teacher'), c.get('season'), c.get('period'), target_slot, c.get('day', ''), exclude_ids={c.get('id')}):
                continue
            if not is_room_available_at(dept_id, term_id, courses, related_courses, c.get('campus'), c.get('room'), c.get('season'), c.get('period'), target_slot, c.get('day', ''), exclude_ids={c.get('id')}):
                continue
            blockers = suite_same_slot_blockers(courses, c, dept_id, target_slot, ignore_ids=group_ids)
            if not blockers or len(blockers) > 3:
                continue
            blockers.sort(key=lambda x: (
                0 if course_subject(x) != course_subject(c) else 1,
                slots.index(x.get('slot')) if x.get('slot') in slots else 99,
                course_label(x),
            ))
            planned_moves = [{'id': c.get('id'), 'from_slot': c.get('slot'), 'to_slot': target_slot, 'dept_id': dept_id, 'course': c}]
            blocker_ids = {b.get('id') for b in blockers}
            move_notes = []
            feasible = True
            for blocker in blockers:
                relocation = find_suite_relocation(
                    dept_id,
                    term_id,
                    courses,
                    related_courses,
                    blocker,
                    target_slot,
                    c.get('slot'),
                    planned_moves,
                    blocker_ids,
                )
                if not relocation:
                    feasible = False
                    break
                planned_moves.append({
                    'id': blocker.get('id'),
                    'from_slot': blocker.get('slot'),
                    'to_slot': relocation['slot'],
                    'dept_id': dept_id,
                    'course': blocker,
                    'suite': relocation['suite'],
                    'travel': relocation['travel'],
                })
                move_notes.append(f"{course_label(blocker)} 从 {blocker.get('slot')}段 调到 {relocation['slot']}段")
            if not feasible:
                continue
            final_courses = virtual_courses_with_moves(courses, planned_moves)
            target_suite = suite_move_note(final_courses, c, dept_id, target_slot)
            target_travel = teacher_day_travel_profile(
                final_courses,
                c.get('teacher'),
                dept_id,
                c.get('season'),
                c.get('period'),
                c.get('day', ''),
                proposed={'slot': target_slot, 'campus': c.get('campus'), 'course': course_label(c)},
                exclude_ids={c.get('id')},
            )
            if target_suite['level'] == 'bad' or target_travel['level'] == 'bad':
                continue
            move_summary = [f"{course_label(c)} 从 {c.get('slot')}段 调到 {target_slot}段"] + move_notes
            validation = suite_reflow_validation(dept_id, term_id, courses, related_courses, c, planned_moves)
            warn_count = sum(1 for m in planned_moves[1:] if m.get('suite', {}).get('level') == 'warn' or m.get('travel', {}).get('level') == 'warn')
            suggestions.append({
                'category': 'suite_reflow',
                'audience': '主管',
                'risk': '需复核',
                'title': f"套班联动重排：{course_label(c)} 调到 {target_slot}段，并同步调整 {len(blockers)} 门套班课",
                'detail': (
                    '；'.join(move_summary) + '。'
                    f"目标教师该时段空闲，原教室可用。{target_suite['text']} {target_travel['text']} "
                    f"{validation['summary']}{warn_count} 个联动点需要主管复核教师产能或套班保留度。"
                ),
                'course_id': c.get('id'),
                'target_slot': target_slot,
                'plan': [
                    {
                        'course_id': move['id'],
                        'course_label': course_label(move['course']),
                        'from_slot': move.get('from_slot'),
                        'to_slot': move.get('to_slot'),
                        'teacher': move['course'].get('teacher', ''),
                        'subject': course_subject(move['course']),
                    }
                    for move in planned_moves
                ],
                'suite_score': target_suite.get('score', 0),
                'suite_delta': target_suite.get('score_delta', 0),
                'suite_validation': validation,
                'affected_count': validation['affected_count'],
                'covered_subjects': validation['covered_subjects'],
                'residual_conflict_count': validation['residual_conflict_count'],
                'travel_level': 'warn' if warn_count or target_travel['level'] == 'warn' else 'good',
                'travel_minutes': target_travel.get('max_minutes', 0),
                'travel_route': target_travel.get('route', ''),
                'travel_segments': target_travel.get('route_segments', []),
                'travel_transitions': target_travel.get('transitions', []),
                'travel_halves': target_travel.get('half_routes', []),
                'travel_flags': target_travel.get('flags', []),
            })
    attach_suggestion_scores(suggestions)
    suggestions.sort(key=suggestion_sort_key)
    return suggestions[:3]

def low_enrollment_release_suggestions(dept_id, courses, group, teacher_subjects):
    if not group:
        return []
    representative = group[0]
    wanted_subject = course_subject(representative)
    if not wanted_subject:
        return []
    group_ids = {c.get('id') for c in group}
    suggestions = []
    seen = set()
    source_courses = courses.low_enrollment if isinstance(courses, CourseIndex) else active_courses(courses)
    for busy_course in source_courses:
        if busy_course.get('id') in group_ids:
            continue
        teacher = busy_course.get('teacher')
        if not teacher or teacher == representative.get('teacher'):
            continue
        if wanted_subject not in teacher_subjects.get(teacher, set()):
            continue
        if not same_course_time(
            busy_course,
            representative.get('season'),
            representative.get('period'),
            representative.get('slot'),
            representative.get('day', ''),
        ):
            continue
        low_info = low_enrollment_info(busy_course)
        if not low_info:
            continue
        key = (teacher, busy_course.get('id'), representative.get('id'))
        if key in seen:
            continue
        seen.add(key)
        travel_profile = teacher_day_travel_profile(
            courses,
            teacher,
            dept_id,
            representative.get('season'),
            representative.get('period'),
            representative.get('day', ''),
            proposed={'slot': representative.get('slot'), 'campus': representative.get('campus'), 'course': course_label(representative)},
            exclude_ids={busy_course.get('id')},
        )
        if travel_profile['level'] == 'bad':
            continue
        release_suite = suite_release_note(courses, busy_course, dept_id)
        target_suite = suite_slot_coverage(courses, representative, dept_id, representative.get('slot'), ignore_ids={representative.get('id')})
        target_subjects = '、'.join(target_suite.get('subjects') or [])
        merge_targets = low_enrollment_merge_targets(courses, busy_course, dept_id, exclude_ids=group_ids)
        detail_parts = [
            f"{course_label(busy_course)} 为{low_info['kind']}班，当前{low_info['count']}人，低于建议保留线{low_info['threshold']}人；这是最终兜底方案，取消/合并可能带来退班和结转风险。",
            f"释放后 {teacher} 可在原时段承接 {course_label(representative)}，减少目标班调课和家长沟通成本。",
            release_suite['text'],
            travel_profile['text'],
        ]
        if merge_targets:
            best_merge = merge_targets[0]
            merge_course = best_merge.get('course') or {}
            detail_parts.append(
                f"若必须合并，优先复核并入同校区同年级的 {course_label(merge_course)}，合并后约{best_merge['after_count']}人；仍需店长确认招生集中和家长接受度。"
            )
        else:
            detail_parts.append('未识别到同校区同年级同科目的明确合并目标，若执行取消需单独评估学生去向。')
        if target_subjects:
            detail_parts.append(f"目标班当前套班覆盖：{target_subjects}，方案不改变目标班上课时间，仍需主管复核同年级科目搭配。")
        evaluation = low_enrollment_release_evaluation(
            courses,
            busy_course,
            representative,
            dept_id,
            low_info,
            release_suite,
            target_suite,
            merge_targets,
            travel_profile,
        )
        suggestions.append({
            'category': 'low_enrollment_release',
            'audience': '主管',
            'risk': '需复核',
            'title': f"兜底方案：评估取消/合并 {course_label(busy_course)}，释放 {teacher}",
            'detail': ' '.join(detail_parts),
            'course_id': representative.get('id'),
            'teacher': teacher,
            'release_course': public_course_conflict(busy_course, dept_id),
            'release_kind': low_info['kind'],
            'release_count': low_info['count'],
            'release_threshold': low_info['threshold'],
            **travel_profile_public_fields(travel_profile),
            'suite_level': release_suite['level'],
            'suite_score': target_suite.get('score', 0),
            'merge_targets': merge_targets,
            'merge_target_count': len(merge_targets),
            'release_evaluation': evaluation,
            'release_science_score': evaluation.get('science_score', 0),
            'withdrawal_risk': evaluation.get('withdrawal_risk', '高'),
        })
    attach_suggestion_scores(suggestions)
    suggestions.sort(key=lambda s: (
        suggestion_sort_key(s),
        -int(s.get('release_science_score') or 0),
        {'中': 0, '高': 1}.get(s.get('withdrawal_risk'), 2),
    ))
    return suggestions[:3]

def teacher_conflict_suggestions(dept_id, term_id, courses, group, related_courses=None):
    suggestions = []
    related_courses = related_courses if related_courses is not None else related_room_courses_for_dept(dept_id, term_id)['courses']
    index = courses if isinstance(courses, CourseIndex) else CourseIndex(courses, related_courses, dept_id)
    slots = slot_order_for(dept_id)

    time_candidates = []
    for c in group:
        for target_slot in schedulable_slots_for_course(c, dept_id):
            if target_slot == c.get('slot'):
                continue
            if teacher_busy_at(index, c.get('teacher'), c.get('season'), c.get('period'), target_slot, c.get('day', ''), exclude_ids={c.get('id')}):
                continue
            if not is_room_available_at(dept_id, term_id, index, related_courses, c.get('campus'), c.get('room'), c.get('season'), c.get('period'), target_slot, c.get('day', ''), exclude_ids={c.get('id')}):
                continue
            suite = suite_move_note(index, c, dept_id, target_slot)
            if suite['level'] == 'bad':
                continue
            travel_profile = teacher_day_travel_profile(
                index, c.get('teacher'), dept_id, c.get('season'), c.get('period'), c.get('day', ''),
                proposed={'slot': target_slot, 'campus': c.get('campus'), 'course': course_label(c)},
                exclude_ids={c.get('id')},
            )
            if travel_profile['level'] == 'bad':
                continue
            risk = '低' if suite['level'] == 'good' and travel_profile['level'] == 'good' else '中'
            time_candidates.append({
                'category': 'teacher_time',
                'audience': '主管',
                'risk': risk,
                'title': f"调整 {course_label(c)} 到 {target_slot}段",
                'detail': f"同教师该时段空闲，原教室可用。{suite['text']} {travel_profile['text']}",
                'course_id': c.get('id'),
                'target_slot': target_slot,
                'target_time': get_time_range(dept_id, target_slot),
                **travel_profile_public_fields(travel_profile),
                'suite_score': suite.get('score', 0),
                'suite_delta': suite.get('score_delta', 0),
            })
    attach_suggestion_scores(time_candidates)
    time_candidates.sort(key=lambda s: (
        suggestion_sort_key(s),
        -s.get('suite_delta', 0),
        -s.get('suite_score', 0),
        slots.index(s.get('target_slot')) if s.get('target_slot') in slots else 99,
    ))
    suggestions.extend(time_candidates[:3])

    teacher_subjects = {}
    for c in index.active:
        t = c.get('teacher')
        s = course_subject(c)
        if t and s:
            teacher_subjects.setdefault(t, set()).add(s)
    for t in load_teachers(dept_id):
        name = t.get('name')
        subject = t.get('subject')
        if name and subject:
            teacher_subjects.setdefault(name, set()).add(subject)

    representative = group[0]
    wanted_subject = course_subject(representative)
    candidates = []
    for teacher, subjects in teacher_subjects.items():
        if not teacher or teacher == representative.get('teacher') or wanted_subject not in subjects:
            continue
        if teacher_busy_at(courses, teacher, representative.get('season'), representative.get('period'), representative.get('slot'), representative.get('day', '')):
            continue
        travel = adjacent_teacher_travel(courses, teacher, dept_id, representative.get('campus'), representative.get('season'), representative.get('period'), representative.get('slot'), representative.get('day', ''))
        if travel['minutes'] > 30:
            continue
        day_profile = teacher_day_travel_profile(
            courses, teacher, dept_id, representative.get('season'), representative.get('period'), representative.get('day', ''),
            proposed={'slot': representative.get('slot'), 'campus': representative.get('campus'), 'course': course_label(representative)},
        )
        if day_profile['level'] == 'bad':
            continue
        candidates.append((travel['minutes'], teacher, travel, day_profile))
    for minutes, teacher, travel, day_profile in sorted(candidates, key=lambda x: (x[0], x[1]))[:3]:
        suggestions.append({
            'category': 'teacher_substitute',
            'audience': '主管',
            'risk': '低' if minutes <= 20 and day_profile['level'] == 'good' else '中',
            'title': f"可考虑由 {teacher} 代上当前时段",
            'detail': f"同科目老师当前时段空闲，避免改班级时间。{travel['text']} {day_profile['text']}" + (' 20-30分钟需评估是否微调上课时间。' if minutes > 20 else ''),
            'course_id': representative.get('id'),
            'teacher': teacher,
            **travel_profile_public_fields(day_profile),
            'travel_minutes': minutes,
            'route_travel_minutes': day_profile.get('max_minutes', 0),
        })

    suggestions.extend(course_swap_suggestions(dept_id, term_id, index, related_courses, group))
    suggestions.extend(suite_reflow_suggestions(dept_id, term_id, index, related_courses, group))
    suggestions.extend(low_enrollment_release_suggestions(dept_id, index, group, teacher_subjects))

    suite_courses = []
    for c in group:
        related = [
            x for x in suite_context_courses(index, c, dept_id)
            if x.get('id') != c.get('id')
            and suite_compatible(c, x, dept_id)
        ]
        if related:
            suite_courses.append(f"{course_label(c)} 关联 {len(related)} 个同校区同年级套班课")
    if suite_courses:
        suggestions.append({
            'category': 'suite_coordination',
            'audience': '主管',
            'risk': '需复核',
            'title': '联动检查套班完整性',
            'detail': '；'.join(suite_courses[:3]) + '。如果单点调课会造成同年级科目撞时段，应同步调整相关双语/益智/博文/科学/实践课。',
        })
    attach_suggestion_scores(suggestions)
    suggestions.sort(key=suggestion_sort_key)
    return suggestions[:8]

def room_capacity_fit(room_meta, group):
    _ = room_meta
    counted_courses = []
    unknown_counts = []
    for c in group:
        count = parse_count_value(c.get('currentCount'))
        if count is None:
            unknown_counts.append(c)
        else:
            counted_courses.append((count, c))
    sorted_courses = sorted(counted_courses, key=lambda item: (item[0], course_label(item[1])))
    recommended = sorted_courses[0][1] if sorted_courses else (unknown_counts[0] if unknown_counts else None)
    label = course_label(recommended) if recommended else ''
    risk = '低' if recommended and not unknown_counts else '需复核'
    note_parts = []
    if label:
        note_parts.append(f"建议优先移动人数较少班级：{label}")
    if unknown_counts:
        note_parts.append('部分班级人数未填，店长需现场确认移动对象')
    return {
        'best_count': sorted_courses[0][0] if sorted_courses else None,
        'fit_count': len(counted_courses) + len(unknown_counts),
        'course_id': recommended.get('id') if recommended else None,
        'course_label': course_label(recommended) if recommended else '',
        'recommended_course': public_course_conflict(recommended) if recommended else None,
        'unknown_count': len(unknown_counts),
        'risk': risk,
        'note': '；'.join(note_parts) or '可任选一门冲突课程调整到该空教室',
    }

def room_solution_checks(campus, season, period, slot, day, room, fit, shared_key=''):
    checks = [
        {'item': '同校区同时间空闲', 'passed': True, 'detail': f"{short_campus_label(campus)} {season} {period} {day} {slot}段未占用"},
        {'item': '移动对象', 'passed': bool(fit.get('course_id')), 'detail': fit.get('note', '')},
    ]
    if shared_key:
        checks.append({'item': '跨部门共用标记', 'passed': True, 'detail': '该教室已识别为可跨部门共用/借用教室'})
    else:
        checks.append({'item': '跨部门共用标记', 'passed': True, 'detail': '未识别为跨部门共用教室'})
    if fit.get('unknown_count'):
        checks.append({'item': '人数完整性', 'passed': False, 'detail': f"{fit.get('unknown_count')}个班人数未填，店长需现场确认"})
    return checks

def room_solution_steps(course_label_text, room, campus, shared_key=''):
    steps = [
        f"先在教室空挡表核对 {short_campus_label(campus)} {room} 当前时段确实空闲",
        f"将 {course_label_text or '适合调整的班级'} 改到 {room}，不改变上课时间",
        '现场同步教务/前台更新教室指引，避免学生走错教室',
    ]
    if shared_key:
        steps.insert(1, '若为跨部门共用教室，先和对方部门确认借用口径')
    return steps

def room_conflict_suggestions(dept_id, term_id, courses, group, related_courses):
    first = group[0]
    campus = first.get('campus', '')
    season, period, slot, day = first.get('season'), first.get('period'), first.get('slot'), first.get('day', '')
    rooms = {}
    for r in load_classrooms(dept_id):
        if r.get('campus') == campus and r.get('name'):
            rooms[r['name']] = r
    for c in courses:
        if c.get('campus') == campus and c.get('room'):
            rooms.setdefault(c['room'], {'name': c['room'], 'campus': campus, 'capacity': c.get('capacity', '')})
    current_rooms = {c.get('room') for c in group}
    candidates = []
    shared_index = build_cross_dept_room_index(term_id) if dept_id in SHARED_CLASSROOM_DEPTS else {}
    for room in sorted(rooms):
        if room in current_rooms:
            continue
        if not is_room_available_at(dept_id, term_id, courses, related_courses, campus, room, season, period, slot, day):
            continue
        meta = rooms[room]
        fit = room_capacity_fit(meta, group)
        if not fit:
            continue
        shared_key = shared_room_key(campus, room)
        shared_meta = shared_index.get(shared_key) if shared_key else None
        course_label_text = fit['course_label'] or '可调整班级'
        candidates.append((fit['risk'] != '低', fit.get('best_count') is None, fit.get('best_count') or 999, room, {
            'category': 'room_swap',
            'audience': '店长',
            'risk': fit['risk'],
            'title': f"将 {course_label_text} 换到 {room}",
            'detail': f"{short_campus_label(campus)} {season} {period} {day} {slot}段空闲；{fit['note']}。",
            'room': room,
            'campus': campus,
            'course_id': fit['course_id'],
            'course_label': fit['course_label'],
            'recommended_course': fit.get('recommended_course'),
            'best_count': fit.get('best_count'),
            'fit_count': fit['fit_count'],
            'unknown_count': fit.get('unknown_count', 0),
            'room_plan': {
                'campus': campus,
                'campus_label': short_campus_label(campus),
                'room': room,
                'room_short': room.replace(campus, '').replace('教学区', '').replace('教室', '').strip() or room,
                'season': season,
                'period': period,
                'day': day,
                'slot': slot,
                'time_label': get_time_range(dept_id, slot),
                'shared': bool(shared_meta),
                'shared_depts': shared_meta.get('depts', []) if shared_meta else [],
                'shared_rooms': shared_meta.get('rooms_by_dept', {}) if shared_meta else {},
            },
            'room_checks': room_solution_checks(campus, season, period, slot, day, room, fit, shared_key if shared_meta else ''),
            'room_steps': room_solution_steps(course_label_text, room, campus, shared_key if shared_meta else ''),
        }))
    candidates.sort(key=lambda item: item[:4])
    suggestions = [item[4] for item in candidates[:6]]
    if not suggestions:
        suggestions.append({
            'category': 'room_swap',
            'audience': '店长',
            'risk': '需协调',
            'title': '当前校区同时间段未找到空教室',
            'detail': '建议店长先核实临时教室、借用教室或现场教室命名是否已录入系统。',
            'campus': campus,
            'room_plan': {
                'campus': campus,
                'campus_label': short_campus_label(campus),
                'season': season,
                'period': period,
                'day': day,
                'slot': slot,
                'time_label': get_time_range(dept_id, slot),
            },
            'room_steps': [
                '先到教室空挡表复核是否有未录入或命名不一致的临时教室',
                '如需借用跨部门教室，先确认对方部门同时间段没有排课',
                '若仍无教室，店长需协调换教室、换校区或保留人工处理记录',
            ],
        })
    attach_suggestion_scores(suggestions)
    return suggestions

def include_related_room_course(current_dept, current_keys, other_course):
    key = course_shared_room_key(other_course)
    if not key:
        return False
    if key in current_keys:
        return True
    return room_owner_dept(other_course.get('room', '')) == current_dept

def related_room_courses_for_dept(dept_id, term_id):
    if dept_id not in SHARED_CLASSROOM_DEPTS:
        return {'courses': [], 'rooms': {}}
    own_courses = active_courses(load_term_data(dept_id, term_id))
    own_keys = {course_shared_room_key(c) for c in own_courses if course_shared_room_key(c)}
    cross_index = build_cross_dept_room_index(term_id)
    rooms = {}
    out = []
    for key, info in cross_index.items():
        rooms[key] = {
            'key': key,
            'label': '共用',
            'type': 'shared',
            'depts': info.get('depts', []),
            'rooms_by_dept': info.get('rooms_by_dept', {}),
        }
    for other_dept in sorted(SHARED_CLASSROOM_DEPTS - {dept_id}):
        if not dept_exists(other_dept) or not term_exists(other_dept, term_id):
            continue
        for c in active_courses(load_term_data(other_dept, term_id)):
            key = course_shared_room_key(c)
            if not include_related_room_course(dept_id, own_keys, c):
                continue
            owner = room_owner_dept(c.get('room', ''))
            label = '共用' if key in cross_index else ('借用' if owner != dept_id else '本部门教室')
            rooms.setdefault(key, {
                'key': key,
                'label': label,
                'type': 'shared' if label == '共用' else 'borrowed',
                'depts': sorted({dept_id, other_dept}),
                'rooms_by_dept': {},
            })
            item = public_course_conflict(c, other_dept)
            item['room_key'] = key
            item['related'] = True
            out.append(item)
    return {'courses': out, 'rooms': rooms}

def conflict_suggestions_cache_key(dept_id, term_id):
    own_version = int(sqlite_store.get_term_version(dept_id, term_id).get('version') or 0)
    related_signature = tuple(cross_dept_room_index_signature(term_id)) if dept_id in SHARED_CLASSROOM_DEPTS else ()
    return (dept_id, term_id, own_version, related_signature)

def get_conflict_suggestions_cache(cache_key):
    with conflict_suggestions_cache_lock:
        cached = conflict_suggestions_cache.get(cache_key)
        if cached is not None:
            conflict_suggestions_cache.pop(cache_key, None)
            conflict_suggestions_cache[cache_key] = cached
            return copy.deepcopy(cached)
    return None

def get_conflict_summary_cache(cache_key):
    with conflict_suggestions_cache_lock:
        cached = conflict_summary_cache.get(cache_key)
        if cached is not None:
            conflict_summary_cache.pop(cache_key, None)
            conflict_summary_cache[cache_key] = cached
            return copy.deepcopy(cached)
    return None

def set_conflict_suggestions_cache(cache_key, payload):
    dept_id, term_id = cache_key[0], cache_key[1]
    with conflict_suggestions_cache_lock:
        stale_keys = [
            key for key in conflict_suggestions_cache
            if (key[0] == dept_id and key[1] == term_id and key != cache_key)
        ]
        for key in stale_keys:
            conflict_suggestions_cache.pop(key, None)
        conflict_suggestions_cache.pop(cache_key, None)
        conflict_suggestions_cache[cache_key] = copy.deepcopy(payload)
        while len(conflict_suggestions_cache) > CONFLICT_SUGGESTIONS_CACHE_MAX:
            oldest_key = next(iter(conflict_suggestions_cache))
            conflict_suggestions_cache.pop(oldest_key, None)

def set_conflict_summary_cache(cache_key, payload):
    dept_id, term_id = cache_key[0], cache_key[1]
    with conflict_suggestions_cache_lock:
        stale_keys = [
            key for key in conflict_summary_cache
            if (key[0] == dept_id and key[1] == term_id and key != cache_key)
        ]
        for key in stale_keys:
            conflict_summary_cache.pop(key, None)
        conflict_summary_cache.pop(cache_key, None)
        conflict_summary_cache[cache_key] = copy.deepcopy(payload)
        while len(conflict_summary_cache) > CONFLICT_SUGGESTIONS_CACHE_MAX:
            oldest_key = next(iter(conflict_summary_cache))
            conflict_summary_cache.pop(oldest_key, None)

def build_conflict_groups_payload(dept_id, term_id, include_suggestions):
    courses = load_term_data(dept_id, term_id)
    related = related_room_courses_for_dept(dept_id, term_id)
    course_index = CourseIndex(courses, related['courses'], dept_id) if include_suggestions else None
    teacher_map = {}
    room_map = {}
    for c in active_courses(courses):
        if c.get('teacher') and c.get('slot'):
            key = f"{c['teacher']}|{c['season']}|{c['period']}|{c['slot']}|{c.get('day','')}"
            teacher_map.setdefault(key, []).append(c)
        if c.get('room') and c.get('campus') and c.get('slot'):
            room_key = course_shared_room_key(c) or f"{c['room']}|{c['campus']}"
            key = f"{room_key}|{c['season']}|{c['period']}|{c['slot']}|{c.get('day','')}"
            room_map.setdefault(key, []).append(c)
    for c in related['courses']:
        if c.get('room') and c.get('campus') and c.get('slot'):
            room_key = c.get('room_key') or course_shared_room_key(c) or f"{c['room']}|{c['campus']}"
            key = f"{room_key}|{c['season']}|{c['period']}|{c['slot']}|{c.get('day','')}"
            room_map.setdefault(key, []).append(c)
    teacher_groups = []
    for group in teacher_map.values():
        if len(group) <= 1:
            continue
        teacher_groups.append(public_conflict_group(
            'teacher',
            group,
            dept_id,
            suggestions=teacher_conflict_suggestions(dept_id, term_id, course_index, group, related['courses']) if include_suggestions else [],
        ))
    room_groups = []
    for group in room_map.values():
        if len(group) <= 1:
            continue
        cross = any(c.get('related') or c.get('dept_label') for c in group)
        room_groups.append(public_conflict_group(
            'room',
            group,
            dept_id,
            suggestions=room_conflict_suggestions(dept_id, term_id, courses, group, related['courses']) if include_suggestions else [],
            cross=cross,
        ))
    return {
        'teacher': teacher_groups,
        'room': room_groups,
    }


# ==================== 部门 / 学期目录 ====================
def load_depts_config(): return load_json(DEPTS_CONFIG, [])
def save_depts_config(d): save_json(DEPTS_CONFIG, d)

def get_dept_dir(dept_id):
    p = os.path.join(DEPTS_DIR, dept_id)
    os.makedirs(p, exist_ok=True)
    os.makedirs(os.path.join(p, 'terms'), exist_ok=True)
    return p

def get_resources_dir(dept_id):
    p = os.path.join(get_dept_dir(dept_id), 'resources')
    os.makedirs(p, exist_ok=True)
    return p

def teachers_file(dept_id): return os.path.join(get_resources_dir(dept_id), 'teachers.json')
def classrooms_file(dept_id): return os.path.join(get_resources_dir(dept_id), 'classrooms.json')
def campus_config_file(dept_id): return os.path.join(get_resources_dir(dept_id), 'campus_config.json')

def load_teachers(dept_id): return load_json(teachers_file(dept_id), [])
def save_teachers(dept_id, teachers): save_json(teachers_file(dept_id), teachers)
def load_classrooms(dept_id): return load_json(classrooms_file(dept_id), [])
def save_classrooms(dept_id, classrooms): save_json(classrooms_file(dept_id), classrooms)

def get_campus_config(dept_id):
    config = load_json(campus_config_file(dept_id), None)
    if config is None:
        return DEFAULT_CAMPUS_CONFIG
    return {
        'districts': config.get('districts', {}),
        'campus_codes': config.get('campus_codes', {}),
    }

def save_campus_config(dept_id, config):
    save_json(campus_config_file(dept_id), config)

def dept_terms_file(dept_id): return os.path.join(get_dept_dir(dept_id), 'terms.json')
def load_terms(dept_id): return load_json(dept_terms_file(dept_id), [])
def save_terms(dept_id, terms): save_json(dept_terms_file(dept_id), terms)

def get_term_dir(dept_id, term_id):
    p = os.path.join(get_dept_dir(dept_id), 'terms', term_id)
    os.makedirs(p, exist_ok=True)
    os.makedirs(os.path.join(p, 'history'), exist_ok=True)
    return p

def term_data_file(dept_id, term_id): return os.path.join(get_term_dir(dept_id, term_id), 'data.json')
def term_original_file(dept_id, term_id): return os.path.join(get_term_dir(dept_id, term_id), 'data_original.json')
def term_history_dir(dept_id, term_id): return os.path.join(get_term_dir(dept_id, term_id), 'history')
def term_changelog_file(dept_id, term_id): return os.path.join(get_term_dir(dept_id, term_id), 'changelog.json')
def term_metadata_file(dept_id, term_id): return os.path.join(get_term_dir(dept_id, term_id), 'metadata.json')
def term_workflow_file(dept_id, term_id): return os.path.join(get_term_dir(dept_id, term_id), 'workflow.json')
def term_history_meta_file(dept_id, term_id): return os.path.join(get_term_dir(dept_id, term_id), 'history_meta.json')
def term_daily_backup_dir(dept_id, term_id):
    p = os.path.join(get_term_dir(dept_id, term_id), 'daily_backups')
    os.makedirs(p, exist_ok=True)
    return p
def term_conflict_status_file(dept_id, term_id): return os.path.join(get_term_dir(dept_id, term_id), 'conflict_status.json')
def load_term_data(dept_id, term_id): return load_json(term_data_file(dept_id, term_id), [])
def load_term_original(dept_id, term_id): return load_json(term_original_file(dept_id, term_id), [])
def load_changelog(dept_id, term_id): return load_json(term_changelog_file(dept_id, term_id), [])
def save_changelog(dept_id, term_id, log): save_json(term_changelog_file(dept_id, term_id), log)
def load_history_meta(dept_id, term_id): return load_json(term_history_meta_file(dept_id, term_id), {})
def save_history_meta(dept_id, term_id, meta): save_json(term_history_meta_file(dept_id, term_id), meta)
def load_workflow(dept_id, term_id):
    return load_json(term_workflow_file(dept_id, term_id), {'status': 'draft', 'updated_at': '', 'updated_by': '', 'history': []})
def save_workflow(dept_id, term_id, wf): save_json(term_workflow_file(dept_id, term_id), wf)
def workflow_namespace(dept_id, term_id): return document_namespace_for_path(term_workflow_file(dept_id, term_id))

VALID_WORKFLOW_STATUSES = {'draft', 'scheduling', 'reviewing', 'confirmed'}
WORKFLOW_TRANSITIONS = {
    'draft': {'scheduling'},
    'scheduling': {'reviewing'},
    'reviewing': {'confirmed', 'scheduling'},
    'confirmed': {'scheduling'},
}

class WorkflowUpdateError(Exception):
    def __init__(self, message, status_code=400):
        self.message = message
        self.status_code = status_code


def validate_workflow_transition(user, current, new_status):
    if new_status not in VALID_WORKFLOW_STATUSES:
        raise WorkflowUpdateError('状态不正确', 400)
    role = user.get('role')
    if role in STAFF_ROLES:
        return
    if new_status != current and new_status not in WORKFLOW_TRANSITIONS.get(current, set()):
        raise WorkflowUpdateError(f'不允许从 {current} 流转到 {new_status}', 400)
    allowed = False
    if current == 'scheduling' and new_status == 'reviewing':
        allowed = role in {'store_manager'} | DEPT_MANAGER_ROLES
    elif current == 'reviewing' and new_status in {'confirmed', 'scheduling'}:
        allowed = role in DEPT_MANAGER_ROLES | {'director'}
    elif new_status == current:
        allowed = True
    if not allowed:
        raise WorkflowUpdateError('无权流转当前状态', 403)


def update_workflow_atomic(dept_id, term_id, user, new_status):
    with workflow_lock:
        namespace = workflow_namespace(dept_id, term_id)
        default = {'status': 'draft', 'updated_at': '', 'updated_by': '', 'history': []}
        if not namespace:
            wf = load_workflow(dept_id, term_id)
            current = wf.get('status') or 'draft'
            validate_workflow_transition(user, current, new_status)
            if new_status == current:
                return {'ok': True, 'changed': False, 'workflow': wf}
            now = time.strftime('%Y-%m-%d %H:%M:%S')
            wf.update({'status': new_status, 'updated_at': now, 'updated_by': user.get('name') or user.get('email')})
            history = wf.get('history') or []
            history.append({'status': new_status, 'at': now, 'by': user.get('email')})
            wf['history'] = history[-100:]
            save_workflow(dept_id, term_id, wf)
            return {'ok': True, 'changed': True, 'workflow': wf}

        def updater(wf):
            current = wf.get('status') or 'draft'
            validate_workflow_transition(user, current, new_status)
            if new_status == current:
                return wf, {'ok': True, 'changed': False, 'workflow': wf}, False
            now = time.strftime('%Y-%m-%d %H:%M:%S')
            next_wf = dict(wf)
            next_wf.update({'status': new_status, 'updated_at': now, 'updated_by': user.get('name') or user.get('email')})
            history = list(next_wf.get('history') or [])
            history.append({'status': new_status, 'at': now, 'by': user.get('email')})
            next_wf['history'] = history[-100:]
            return next_wf, {'ok': True, 'changed': True, 'workflow': next_wf}, True

        result = sqlite_store.update_document(namespace, default, updater)
        if result.get('changed'):
            write_json_file(term_workflow_file(dept_id, term_id), result['workflow'])
        return result
def load_conflict_status(dept_id, term_id): return load_json(term_conflict_status_file(dept_id, term_id), {})
def save_conflict_status(dept_id, term_id, data): save_json(term_conflict_status_file(dept_id, term_id), data)

def update_conflict_status_atomic(dept_id, term_id, key, status, user):
    namespace = document_namespace_for_path(term_conflict_status_file(dept_id, term_id))
    now = time.strftime('%Y-%m-%d %H:%M:%S')
    item = {
        'status': status,
        'updated_at': now,
        'updated_by': user.get('name') or user.get('email'),
    }
    with conflict_status_lock:
        if not namespace:
            data = load_conflict_status(dept_id, term_id)
            data[key] = item
            save_conflict_status(dept_id, term_id, data)
            return item

        def updater(data):
            next_data = dict(data or {})
            next_data[key] = item
            return next_data, {'data': next_data, 'item': item}, True

        result = sqlite_store.update_document(namespace, {}, updater)
        write_json_file(term_conflict_status_file(dept_id, term_id), result['data'])
        return result['item']

def ensure_daily_backup(dept_id, term_id):
    backup_dir = term_daily_backup_dir(dept_id, term_id)
    today = time.strftime('%Y%m%d')
    path = os.path.join(backup_dir, f'{today}.json')
    if not os.path.exists(path):
        save_json(path, load_term_data(dept_id, term_id))
    files = [
        os.path.join(backup_dir, fn) for fn in os.listdir(backup_dir)
        if fn.endswith('.json') and os.path.isfile(os.path.join(backup_dir, fn))
    ]
    files.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    for old in files[DAILY_BACKUP_KEEP_DAYS:]:
        try:
            os.remove(old)
        except OSError:
            pass
    return path

def latest_daily_backup_info(dept_id, term_id):
    backup_dir = term_daily_backup_dir(dept_id, term_id)
    if not os.path.isdir(backup_dir):
        return None
    files = [
        os.path.join(backup_dir, fn) for fn in os.listdir(backup_dir)
        if fn.endswith('.json') and os.path.isfile(os.path.join(backup_dir, fn))
    ]
    if not files:
        return None
    latest = max(files, key=os.path.getmtime)
    return {
        'filename': os.path.basename(latest),
        'saved_at': time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(os.path.getmtime(latest))),
        'size': os.path.getsize(latest),
    }

def history_file_count(dept_id, term_id):
    hist_dir = term_history_dir(dept_id, term_id)
    if not os.path.isdir(hist_dir):
        return 0
    return sum(
        1 for fn in os.listdir(hist_dir)
        if fn.endswith('.json') and os.path.isfile(os.path.join(hist_dir, fn))
    )

def file_signature(path):
    namespace = document_namespace_for_path(path)
    if namespace and namespace.startswith('term:'):
        parts = namespace.split(':')
        if len(parts) == 4 and parts[3] in {'data', 'original'}:
            tv = sqlite_store.get_term_version(parts[1], parts[2])
            return f"sqlite:{tv['version']}"
    try:
        st = os.stat(path)
    except OSError:
        return 'missing'
    return f"{int(st.st_mtime_ns)}:{st.st_size}"

def build_term_metadata(dept_id, term_id, user=None):
    data_file = term_data_file(dept_id, term_id)
    original_file = term_original_file(dept_id, term_id)
    term_version = sqlite_store.get_term_version(dept_id, term_id)
    data_sig = f"sqlite:{term_version['version']}"
    original_sig = data_sig
    data_mtime = os.path.getmtime(data_file) if os.path.exists(data_file) else 0
    original_mtime = os.path.getmtime(original_file) if os.path.exists(original_file) else 0
    updated_ts = max(data_mtime, original_mtime)
    meta = {
        'count': len(load_json(data_file, [])),
        'original_count': len(load_json(original_file, [])),
        'data_version': data_sig,
        'original_version': original_sig,
        'version': data_sig,
        'updated_at': term_version.get('updated_at') or time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(updated_ts or time.time())),
    }
    meta['updated_by'] = (user.get('name') or user.get('email')) if user else term_version.get('updated_by', '')
    return meta

def save_term_metadata(dept_id, term_id, user=None):
    meta = build_term_metadata(dept_id, term_id, user=user)
    save_json(term_metadata_file(dept_id, term_id), meta)
    return meta

def load_term_metadata(dept_id, term_id):
    meta = load_json(term_metadata_file(dept_id, term_id), {})
    term_version = sqlite_store.get_term_version(dept_id, term_id)
    current_version = f"sqlite:{term_version['version']}"
    if meta.get('version') != current_version:
        meta = build_term_metadata(dept_id, term_id)
        save_json(term_metadata_file(dept_id, term_id), meta)
    return meta

def load_term_version_metadata(dept_id, term_id):
    term_version = sqlite_store.get_term_version(dept_id, term_id)
    version = f"sqlite:{term_version['version']}"
    return {
        'data_version': version,
        'original_version': version,
        'version': version,
        'updated_at': term_version.get('updated_at') or '',
        'updated_by': term_version.get('updated_by') or '',
    }

def term_version_conflict(dept_id, term_id):
    expected = request.headers.get('X-Data-Version')
    if not expected:
        current = f"sqlite:{sqlite_store.get_term_version(dept_id, term_id)['version']}"
        return jsonify({
            'error': '缺少数据版本，请刷新页面后再操作',
            'code': 'missing_data_version',
            'current_version': current,
        }), 428
    current = f"sqlite:{sqlite_store.get_term_version(dept_id, term_id)['version']}"
    if expected != current:
        return jsonify({
            'error': '数据已被其他人修改，请刷新后再操作',
            'code': 'version_conflict',
            'current_version': current,
        }), 409
    return None

class TermVersionConflict(Exception):
    def __init__(self, current_version):
        self.current_version = current_version


def parse_sqlite_version(value):
    if not value:
        return None
    m = re.match(r'^sqlite:(\d+)$', str(value))
    return int(m.group(1)) if m else None


def current_request_term_version():
    if not has_request_context():
        return None
    return parse_sqlite_version(request.headers.get('X-Data-Version'))


def term_version_conflict_payload(current_version):
    current = f"sqlite:{current_version}"
    return {
        'error': '数据已被其他人修改，请刷新后再操作',
        'code': 'version_conflict',
        'current_version': current,
    }


def term_version_conflict_response(exc):
    return jsonify(term_version_conflict_payload(exc.current_version)), 409


def current_term_version_number(dept_id, term_id):
    return int(sqlite_store.get_term_version(dept_id, term_id).get('version') or 0)


def term_version_header(version):
    return f"sqlite:{int(version)}"


def missing_data_version_problem(dept_id, term_id):
    if request.headers.get('X-Data-Version'):
        return None
    return {
        'error': '缺少数据版本，请刷新页面后再操作',
        'code': 'missing_data_version',
        'current_version': term_version_header(current_term_version_number(dept_id, term_id)),
    }, 428


def is_request_version_stale(current_version):
    return request.headers.get('X-Data-Version') != term_version_header(current_version)


def comparable_field_value(value):
    return '' if value is None else str(value)


def field_values_equal(left, right):
    return comparable_field_value(left) == comparable_field_value(right)


def course_patch_fields(payload):
    if not isinstance(payload, dict):
        return {}
    return {k: v for k, v in payload.items() if k in COURSE_PATCH_FIELDS}


def course_patch_base_fields(payload):
    if not isinstance(payload, dict):
        return {}
    base = payload.get('_base')
    if base is None:
        base = payload.get('base_fields')
    if base is None:
        base = payload.get('base')
    return base if isinstance(base, dict) else {}


def course_label_for_message(course, fallback=''):
    return ' '.join(
        str(part)
        for part in [course.get('code'), course.get('name')]
        if part
    ) or str(fallback or course.get('id') or '')


def field_level_conflict_problem(dept_id, term_id, course, fields, base_fields, current_version):
    missing = missing_data_version_problem(dept_id, term_id)
    if missing:
        return missing
    if not is_request_version_stale(current_version):
        return None
    for field, desired_value in fields.items():
        actual_value = course.get(field, '')
        if field_values_equal(actual_value, desired_value):
            continue
        if field not in base_fields:
            return {
                'error': '数据已更新，缺少字段基线，无法安全合并，请刷新后重试',
                'code': 'version_conflict',
                'current_version': term_version_header(current_version),
                'course_id': course.get('id'),
                'course_label': course_label_for_message(course),
                'field': field,
                'field_label': FIELD_LABELS_PY.get(field, field),
            }, 409
        base_value = base_fields.get(field, '')
        if not field_values_equal(actual_value, base_value):
            label = FIELD_LABELS_PY.get(field, field)
            return {
                'error': f'{label}已被其他人修改，请刷新后核对',
                'code': 'field_conflict',
                'current_version': term_version_header(current_version),
                'course_id': course.get('id'),
                'course_label': course_label_for_message(course),
                'field': field,
                'field_label': label,
                'base_value': comparable_field_value(base_value),
                'current_value': comparable_field_value(actual_value),
                'attempted_value': comparable_field_value(desired_value),
            }, 409
    return None


def jsonify_with_term_version(payload, dept_id, term_id, status=200):
    resp = make_response(jsonify(payload), status)
    resp.headers['X-Data-Version'] = f"sqlite:{sqlite_store.get_term_version(dept_id, term_id)['version']}"
    return resp

def data_api_etag(*parts):
    raw = '|'.join(str(p) for p in parts)
    return hashlib.sha1(raw.encode('utf-8')).hexdigest()

def stable_json_hash(value):
    raw = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(',', ':'))
    return hashlib.sha1(raw.encode('utf-8')).hexdigest()

def conditional_json_response(payload, etag, extra_headers=None):
    headers = {'Cache-Control': 'private, no-cache'}
    if extra_headers:
        headers.update(extra_headers)
    if request.if_none_match.contains(etag):
        resp = make_response('', 304)
    else:
        resp = make_response(jsonify(payload))
    resp.set_etag(etag)
    for key, value in headers.items():
        resp.headers[key] = value
    return resp

def conditional_not_modified_response(etag, extra_headers=None):
    if not request.if_none_match.contains(etag):
        return None
    headers = {'Cache-Control': 'private, no-cache'}
    if extra_headers:
        headers.update(extra_headers)
    resp = make_response('', 304)
    resp.set_etag(etag)
    for key, value in headers.items():
        resp.headers[key] = value
    return resp

def find_course_index(courses, course_id):
    for i, c in enumerate(courses):
        if c.get('id') == course_id:
            return i
    if 0 <= course_id < len(courses) and courses[course_id].get('id') in (None, course_id):
        return course_id
    return None

def unique_history_path(hist_dir, ts):
    base = os.path.join(hist_dir, f'data_{ts}.json')
    if not os.path.exists(base):
        return base
    i = 1
    while True:
        candidate = os.path.join(hist_dir, f'data_{ts}_{i}.json')
        if not os.path.exists(candidate):
            return candidate
        i += 1

def prune_history(hist_dir, keep=HISTORY_KEEP_FILES):
    files = [
        os.path.join(hist_dir, fn)
        for fn in os.listdir(hist_dir)
        if fn.endswith('.json') and os.path.isfile(os.path.join(hist_dir, fn))
    ]
    files.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    for old in files[keep:]:
        try:
            os.remove(old)
        except OSError:
            pass

def save_term_data(dept_id, term_id, courses, user=None, old_courses=None, action='保存排课', reason='', expected_version=None):
    meta = None
    backup_name = None
    before_version = sqlite_store.get_term_version(dept_id, term_id).get('version')
    previous_courses = None
    diffs = []
    with data_lock:
        ensure_daily_backup(dept_id, term_id)
        ts = time.strftime('%Y%m%d_%H%M%S')
        data_file = term_data_file(dept_id, term_id)
        hist_dir = term_history_dir(dept_id, term_id)
        backup = unique_history_path(hist_dir, ts)
        previous_courses = old_courses if old_courses is not None else load_term_data(dept_id, term_id)
        diffs = compute_diffs(previous_courses, courses)
        data_namespace = document_namespace_for_path(data_file)
        expected_version = current_request_term_version() if expected_version is None else expected_version
        if data_namespace:
            result = sqlite_store.set_term_document_and_touch_version(
                data_namespace,
                courses,
                dept_id,
                term_id,
                user=user,
                expected_version=expected_version,
            )
            if not result.get('ok'):
                current = result.get('current') or {}
                raise TermVersionConflict(current.get('version'))
            write_json_file(data_file, courses)
        else:
            save_json(data_file, courses)
            sqlite_store.touch_term_version(dept_id, term_id, user=user)
        save_json(backup, previous_courses)
        backup_name = os.path.basename(backup)
        if user:
            history_meta = load_history_meta(dept_id, term_id)
            current_version = sqlite_store.get_term_version(dept_id, term_id).get('version')
            summary = summarize_course_changes(previous_courses, courses, limit=0)
            history_meta[backup_name] = {
                'saved_at': time.strftime('%Y-%m-%d %H:%M:%S'),
                'user': user.get('name') or user.get('email'),
                'email': user.get('email', ''),
                'action': action,
                'reason': reason,
                'backup_kind': 'before_change',
                'before_version': before_version,
                'after_version': current_version,
                'diff_summary': {
                    'added_count': summary.get('added_count', 0),
                    'removed_count': summary.get('removed_count', 0),
                    'changed_count': len(diffs),
                    'fields': summarize_diff_fields(diffs),
                },
                'class_count': len(previous_courses),
            }
            save_history_meta(dept_id, term_id, history_meta)
        prune_history(hist_dir)
        meta = save_term_metadata(dept_id, term_id, user=user)
    if user and old_courses is not None:
        if diffs:
            log = load_changelog(dept_id, term_id)
            log.append({
                'time': time.strftime('%Y-%m-%d %H:%M:%S'),
                'user': user.get('name') or user.get('email'),
                'email': user.get('email'),
                'action': action,
                'reason': reason,
                'changes': diffs,
            })
            if len(log) > 500:
                log = log[-500:]
            save_changelog(dept_id, term_id, log)
    return meta


def save_term_data_or_conflict(*args, **kwargs):
    try:
        save_term_data(*args, **kwargs)
        return None
    except TermVersionConflict as exc:
        return term_version_conflict_response(exc)


def compute_diffs(old_courses, new_courses):
    diffs = []
    old_map = {c.get('id', i): c for i, c in enumerate(old_courses)}
    new_map = {c.get('id', i): c for i, c in enumerate(new_courses)}
    track_fields = [
        'teacher', 'slot', 'timeRange', 'room', 'period', 'classType',
        'lifecycle_status', 'currentCount', 'merged_into_code',
    ]
    for cid, new_c in new_map.items():
        old_c = old_map.get(cid)
        if not old_c:
            continue
        for field in track_fields:
            ov = old_c.get(field, '')
            nv = new_c.get(field, '')
            if ov != nv:
                diffs.append({
                    'course_id': cid,
                    'code': new_c.get('code', ''),
                    'name': new_c.get('name', ''),
                    'field': field,
                    'from': ov,
                    'to': nv,
                })
    return diffs


def summarize_diff_fields(diffs, limit=4):
    field_counts = {}
    for d in diffs or []:
        label = FIELD_LABELS_PY.get(d.get('field', ''), d.get('field', ''))
        field_counts[label] = field_counts.get(label, 0) + 1
    ordered = sorted(field_counts.items(), key=lambda item: (-item[1], item[0]))
    return [{'field': field, 'count': count} for field, count in ordered[:limit]]


def course_key(course, fallback):
    code = course.get('code')
    if code:
        return f"code:{code}"
    return f"id:{course.get('id', fallback)}"

def summarize_course_changes(old_courses, new_courses, limit=30):
    old_map = {course_key(c, i): c for i, c in enumerate(old_courses)}
    new_map = {course_key(c, i): c for i, c in enumerate(new_courses)}
    old_keys = set(old_map)
    new_keys = set(new_map)
    fields = ['teacher', 'period', 'slot', 'room', 'classType', 'campus', 'name']
    added = [new_map[k] for k in sorted(new_keys - old_keys)]
    removed = [old_map[k] for k in sorted(old_keys - new_keys)]
    changed = []
    for k in sorted(old_keys & new_keys):
        old = old_map[k]
        new = new_map[k]
        diffs = []
        for field in fields:
            if old.get(field, '') != new.get(field, ''):
                diffs.append({'field': field, 'from': old.get(field, ''), 'to': new.get(field, '')})
        if diffs:
            changed.append({
                'code': new.get('code') or old.get('code') or '',
                'name': new.get('name') or old.get('name') or '',
                'diffs': diffs,
            })
    return {
        'added_count': len(added),
        'removed_count': len(removed),
        'changed_count': len(changed),
        'added': [{'code': c.get('code', ''), 'name': c.get('name', ''), 'campus': c.get('campus', '')} for c in added[:limit]],
        'removed': [{'code': c.get('code', ''), 'name': c.get('name', ''), 'campus': c.get('campus', '')} for c in removed[:limit]],
        'changed': changed[:limit],
    }

def is_inserted_course(course):
    return course.get('desc') == '插空新增' or course.get('created_by_action') == 'insert_course'

def mark_course_lifecycle(course, status, user, reason=''):
    now = time.strftime('%Y-%m-%d %H:%M:%S')
    if status == 'active':
        for key in [
            'lifecycle_status', 'lifecycle_reason', 'lifecycle_at', 'lifecycle_by',
            'merged_into_id', 'merged_into_code', 'merged_into_name', 'merged_count_added',
        ]:
            course.pop(key, None)
        return
    course['lifecycle_status'] = status
    course['lifecycle_reason'] = reason
    course['lifecycle_at'] = now
    course['lifecycle_by'] = user.get('email') if user else ''

def merge_course_counts(target, source):
    target_count = parse_count_value(target.get('currentCount'))
    source_count = parse_count_value(source.get('currentCount'))
    if target_count is None or source_count is None:
        return False
    target['currentCount'] = str(target_count + source_count)
    return True

def unmerge_course_counts(target, source):
    target_count = parse_count_value(target.get('currentCount'))
    source_count = parse_count_value(source.get('currentCount'))
    if target_count is None or source_count is None:
        return False
    target['currentCount'] = str(max(0, target_count - source_count))
    return True

def extract_grade_from_course(c, dept_id):
    if dept_id == 'qingshao':
        return ''
    grade = c.get('grade', '')
    if grade in ('初三', '高二', '高一', '初二', '初一'):
        return grade
    name = c.get('name', '')
    if '高二' in name: return '高二'
    if '初三' in name or '高一预科' in name: return '初三'
    if '高一' in name: return '高一'
    if '初二' in name or '高一预备' in name: return '初二'
    if '初一' in name or '高一准备' in name: return '初一'
    code = c.get('code', '')
    grade_map = {'1': '初一', '2': '初二', '3': '初三', '4': '高一', '5': '高二'}
    if len(code) >= 8 and code[1:3] == 'ZV':
        return grade_map.get(code[0], '')
    if len(code) >= 6 and code[0:2].isdigit() and code[3] == 'Z':
        return grade_map.get(code[4], '')
    return ''

def extract_source_fy(courses):
    for c in courses:
        code = str(c.get('code', '') or '')
        if len(code) >= 5 and code[1:3] == 'ZV':
            return code[3:5]
        if len(code) >= 2 and code[0:2].isdigit():
            return code[0:2]
    return '27'

def build_generated_courses(source_courses, dept_id, term_id, mode):
    from code_generator import transform_code, extract_target_fy
    config = GENERATE_MODES[mode]
    graduating_grades = {'初三', '高二'}
    filtered = []
    removed_count = 0
    for c in source_courses:
        if config['remove_graduating'] and extract_grade_from_course(c, dept_id) in graduating_grades:
            removed_count += 1
            continue
        filtered.append(c)
    base_fy = extract_source_fy(filtered)
    target_fy = str(int(base_fy) + config['fy_increment']) if base_fy.isdigit() else extract_target_fy(term_id)
    generated = []
    season_counts = {}
    warnings = []
    sample_map = {}
    for season_config in config['target_seasons']:
        season = season_config['season']
        season_counts[season] = 0
        for c in filtered:
            code = c.get('code', '')
            new_code = transform_code(code, dept_id, target_fy, season)
            if code and new_code == code:
                warnings.append(f'班级编码无法转换：{code}')
            if code and len(sample_map) < 3:
                sample_map.setdefault(code, {})[season] = new_code
            for period in season_config['periods']:
                gc = dict(c)
                gc['season'] = season
                gc['period'] = period
                gc['code'] = new_code
                gc['day'] = period if season_config.get('day_from_period') else season_config.get('day', '每天')
                gc.pop('_originalTeacher', None)
                gc.pop('_originalPeriod', None)
                generated.append(gc)
                season_counts[season] += 1
    for i, c in enumerate(generated):
        c['id'] = i
    sample_code_transform = [{'from': k, 'to': v} for k, v in sample_map.items()]
    campus_counts = {}
    for c in generated:
        campus = c.get('campus') or '未填写校区'
        campus_counts[campus] = campus_counts.get(campus, 0) + 1
    return generated, {
        'source_count': len(source_courses),
        'removed_graduating': removed_count,
        'remaining_after_filter': len(filtered),
        'season_counts': season_counts,
        'campus_counts': campus_counts,
        'total_output': len(generated),
        'sample_code_transform': sample_code_transform,
        'warnings': sorted(set(warnings))[:20],
    }

def safe_history_path(dept_id, term_id, filename):
    if filename != os.path.basename(filename) or not filename.endswith('.json'):
        return None
    return os.path.join(term_history_dir(dept_id, term_id), filename)

def dept_exists(dept_id):
    return any(d['id'] == dept_id for d in load_depts_config())

def term_exists(dept_id, term_id):
    return any(t['id'] == term_id for t in load_terms(dept_id))

def yiduiyi_monthly_term_name(dt=None):
    dt = dt or datetime.now()
    fy = dt.year - 2000 + (1 if dt.month >= 6 else 0)
    quarter = ((dt.month - 6) % 12) // 3 + 1
    return f'FY{fy:02d}Q{quarter}·{dt.month}月'

def make_term_id_from_terms(dept_id, terms):
    if dept_id == 'yiduiyi':
        now = datetime.now()
        fy = now.year - 2000 + (1 if now.month >= 6 else 0)
        quarter = ((now.month - 6) % 12) // 3 + 1
        base = f'fy{fy:02d}_q{quarter}_{now.month:02d}'
    else:
        base = 'term_' + time.strftime('%Y%m%d%H%M%S')
    existing = {t.get('id') for t in terms}
    if base not in existing:
        return base
    suffix = 2
    while f'{base}_{suffix}' in existing:
        suffix += 1
    return f'{base}_{suffix}'

def make_term_id(dept_id):
    return make_term_id_from_terms(dept_id, load_terms(dept_id))

def get_default_term(dept_id):
    terms = load_terms(dept_id)
    for t in terms:
        if t.get('is_default') and not t.get('archived'):
            return t.get('id')
    for t in terms:
        if not t.get('archived'):
            return t.get('id')
    return terms[0]['id'] if terms else None


# ==================== 数据迁移 ====================
def migrate_legacy():
    # 1. 根目录历史数据 → gaozhi
    OLD_DATA = os.path.join(BASE_DIR, 'data.json')
    OLD_ORIGINAL = os.path.join(BASE_DIR, 'data_original.json')
    OLD_HISTORY = os.path.join(BASE_DIR, 'history')
    gz_flat = os.path.join(DEPTS_DIR, 'gaozhi', 'data.json')
    gz_termed = os.path.join(DEPTS_DIR, 'gaozhi', 'terms', DEFAULT_TERM_ID, 'data.json')
    if os.path.exists(OLD_DATA) and not os.path.exists(gz_flat) and not os.path.exists(gz_termed):
        gz = get_dept_dir('gaozhi')
        shutil.copy2(OLD_DATA, os.path.join(gz, 'data.json'))
        if os.path.exists(OLD_ORIGINAL):
            shutil.copy2(OLD_ORIGINAL, os.path.join(gz, 'data_original.json'))
        if os.path.exists(OLD_HISTORY):
            os.makedirs(os.path.join(gz, 'history'), exist_ok=True)
            for fn in os.listdir(OLD_HISTORY):
                dst = os.path.join(gz, 'history', fn)
                if not os.path.exists(dst):
                    shutil.copy2(os.path.join(OLD_HISTORY, fn), dst)
        if not os.path.exists(DEPTS_CONFIG):
            save_depts_config([{'id': 'gaozhi', 'name': '高中班级部', 'description': 'FY27排课'}])

    # 2. dept 下扁平结构 → terms/<DEFAULT_TERM_ID>/
    for d in load_depts_config():
        dept_id = d['id']
        dept_root = get_dept_dir(dept_id)
        flat_data = os.path.join(dept_root, 'data.json')
        terms_json = dept_terms_file(dept_id)
        if os.path.exists(flat_data) and not os.path.exists(terms_json):
            term_dir = get_term_dir(dept_id, DEFAULT_TERM_ID)
            for fn in ['data.json', 'data_original.json']:
                src = os.path.join(dept_root, fn)
                if os.path.exists(src):
                    dst = os.path.join(term_dir, fn)
                    if not os.path.exists(dst):
                        shutil.move(src, dst)
            old_hist = os.path.join(dept_root, 'history')
            new_hist = os.path.join(term_dir, 'history')
            if os.path.isdir(old_hist):
                os.makedirs(new_hist, exist_ok=True)
                for fn in os.listdir(old_hist):
                    dst = os.path.join(new_hist, fn)
                    if not os.path.exists(dst):
                        shutil.move(os.path.join(old_hist, fn), dst)
                try: os.rmdir(old_hist)
                except OSError: pass
            save_terms(dept_id, [{
                'id': DEFAULT_TERM_ID, 'name': DEFAULT_TERM_NAME,
                'description': '历史数据自动迁移批次',
                'created_at': time.strftime('%Y-%m-%d %H:%M:%S'),
            }])

def bootstrap_admin():
    """没有用户时，生成一次性 admin 邀请码并打印到控制台。"""
    if load_users():
        return
    existing = [c for c in load_invites() if c.get('role') == 'admin' and not c.get('used_by')]
    if existing:
        code = existing[0]['code']
    else:
        code = gen_invite_code(12)
        invites = load_invites()
        invites.append({
            'code': code, 'dept_id': None, 'role': 'admin',
            'created_at': time.strftime('%Y-%m-%d %H:%M:%S'),
            'created_by': 'system', 'used_by': None, 'used_at': None,
        })
        save_invites(invites)
    with open(BOOTSTRAP_FILE, 'w', encoding='utf-8') as f:
        f.write(f'初始管理员邀请码：{code}\n请在 /auth 注册时使用，注册成功后此码作废。\n')
    print('=' * 60)
    print(f'  初始管理员邀请码: {code}')
    print(f'  请访问 /auth 用此邀请码注册首个管理员账号')
    print('=' * 60)

migrate_legacy()
sqlite_store.init_db()
bootstrap_admin()


# ==================== 静态入口 ====================
def serve_project_file(filename, mimetype, cache_seconds=0):
    path = os.path.join(SOURCE_DIR, filename)
    stat = os.stat(path)
    etag = hashlib.sha1(f"{filename}:{stat.st_mtime_ns}:{stat.st_size}".encode('utf-8')).hexdigest()
    cache_control = f'private, max-age={cache_seconds}, must-revalidate' if cache_seconds else 'no-cache'
    if request.if_none_match.contains(etag):
        resp = make_response('', 304)
        resp.set_etag(etag)
        resp.headers['Cache-Control'] = cache_control
        return resp
    with open(path, 'rb') as f:
        data = f.read()
    resp = make_response(data)
    resp.mimetype = mimetype
    resp.set_etag(etag)
    resp.headers['Cache-Control'] = cache_control
    return resp

@app.route('/auth')
def auth_page():
    return serve_project_file('auth.html', 'text/html')

@app.route('/app.js')
def serve_js():
    return serve_project_file('app.js', 'application/javascript', cache_seconds=60)

@app.route('/app_utils.js')
def serve_utils_js():
    return serve_project_file('app_utils.js', 'application/javascript', cache_seconds=60)

@app.route('/')
@app.route('/admin')
def portal():
    if not current_user():
        return redirect('/auth')
    return serve_project_file('portal.html', 'text/html')


# ==================== Auth API ====================
@app.route('/api/departments/public')
def list_departments_public():
    return jsonify([
        {'id': d['id'], 'name': d['name'], 'description': d.get('description', '')}
        for d in load_depts_config()
    ])


@app.route('/healthz')
def healthz():
    db_exists = os.path.exists(sqlite_store.DB_PATH)
    return jsonify({
        'ok': True,
        'database_exists': db_exists,
        'runtime_ok': runtime_diagnostics().get('ok', False),
        'server_time': time.strftime('%Y-%m-%d %H:%M:%S'),
    })

@app.route('/api/auth/register', methods=['POST'])
def auth_register():
    body = request.json or {}
    email = (body.get('email') or '').strip().lower()
    password = body.get('password') or ''
    name = (body.get('name') or '').strip()
    dept_id = (body.get('dept_id') or '').strip()
    selected_role = (body.get('role') or '').strip()
    campus = (body.get('campus') or '').strip()
    district = (body.get('district') or '').strip()
    invite_code = (body.get('invite_code') or '').strip().upper()

    if not name:
        return jsonify({'error': '请填写姓名'}), 400
    if not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email):
        return jsonify({'error': '邮箱格式不正确'}), 400
    if len(password) < 6:
        return jsonify({'error': '密码至少 6 位'}), 400
    if not invite_code:
        return jsonify({'error': '请输入邀请码'}), 400

    def updater(users, invites):
        if any(u.get('email') == email for u in users):
            raise AuthUpdateError('邮箱已被注册', 400)
        rec = next((c for c in invites if c['code'] == invite_code), None)
        if not rec:
            raise AuthUpdateError('邀请码无效', 400)
        if rec.get('used_by'):
            raise AuthUpdateError('邀请码已被使用', 400)
        role = rec.get('role') or 'user'
        if selected_role and selected_role != role:
            raise AuthUpdateError('所选身份与权限码不匹配，请确认后重新选择', 400)
        if role in STAFF_ROLES:
            final_dept = dept_id or None
            if final_dept and not dept_exists(final_dept):
                raise AuthUpdateError('部门不存在', 400)
        else:
            if not dept_id:
                raise AuthUpdateError('请选择部门', 400)
            if not dept_exists(dept_id):
                raise AuthUpdateError('部门不存在', 400)
            final_dept = dept_id
        if role == 'store_manager':
            if not campus:
                raise AuthUpdateError('店长身份请选择校区', 400)

        auth_fields = {}
        set_user_password(auth_fields, password)
        new_users = list(users)
        new_users.append({
            'id': secrets.token_hex(8),
            'email': email,
            'name': name,
            'salt': auth_fields['salt'],
            'password_hash': auth_fields['password_hash'],
            'dept_id': final_dept,
            'role': role,
            'campus': campus if role == 'store_manager' else None,
            'district': None,
            'created_at': time.strftime('%Y-%m-%d %H:%M:%S'),
        })
        new_invites = [dict(c) for c in invites]
        rec_copy = next(c for c in new_invites if c['code'] == invite_code)
        used_at = time.strftime('%Y-%m-%d %H:%M:%S')
        rec_copy['used_by'] = email
        rec_copy['used_at'] = used_at
        return new_users, new_invites, {'ok': True}

    try:
        update_users_and_invites_atomic(updater)
    except AuthUpdateError as exc:
        return jsonify({'error': exc.message}), exc.status_code

    return jsonify({'ok': True})


def public_user_info(u, dept_names=None):
    dept_names = dept_names or {d['id']: d.get('name', d['id']) for d in load_depts_config()}
    return {
        'id': u.get('id'),
        'email': u.get('email'),
        'name': u.get('name', ''),
        'dept_id': u.get('dept_id'),
        'dept_name': dept_names.get(u.get('dept_id') or '', ''),
        'role': u.get('role') or 'user',
        'campus': u.get('campus') or '',
        'district': u.get('district') or '',
        'created_at': u.get('created_at', ''),
    }


def normalize_user_scope(role, dept_id='', campus=''):
    role = role or 'user'
    dept_id = (dept_id or '').strip() or None
    campus = (campus or '').strip() or None
    if role not in VALID_ROLES:
        return None, {'error': '角色不正确'}
    if role in STAFF_ROLES:
        if dept_id and not dept_exists(dept_id):
            return None, {'error': '部门不存在'}
        return {'dept_id': dept_id, 'campus': None, 'district': None}, None
    if not dept_id or not dept_exists(dept_id):
        return None, {'error': '请选择有效部门'}
    if role == 'store_manager':
        if not campus:
            return None, {'error': '店长账号必须绑定校区'}
        return {'dept_id': dept_id, 'campus': campus, 'district': None}, None
    return {'dept_id': dept_id, 'campus': None, 'district': None}, None

@app.route('/api/auth/login', methods=['POST'])
def auth_login():
    body = request.json or {}
    email = (body.get('email') or '').strip().lower()
    password = body.get('password') or ''
    client_ip = client_ip_for_rate_limit()
    if count_login_attempts(client_ip) >= 8:
        return jsonify({'error': '登录尝试过于频繁，请稍后再试'}), 429
    u = find_user_by_email(email)
    if not u or not verify_password(password, u['salt'], u['password_hash']):
        record_login_attempt(client_ip, success=False)
        return jsonify({'error': '邮箱或密码错误'}), 401
    record_login_attempt(client_ip, success=True)
    token = create_session(u['id'])
    request.session_token = token
    request.session_record = sqlite_store.get_session(token, ttl=SESSION_TTL)
    resp = make_response(jsonify({'ok': True}))
    resp.set_cookie(SESSION_COOKIE, token, max_age=SESSION_TTL, httponly=True, samesite='Lax', secure=COOKIE_SECURE)
    return resp

@app.route('/api/auth/logout', methods=['POST'])
def auth_logout():
    token = request.cookies.get(SESSION_COOKIE)
    if token:
        destroy_session(token)
    request.session_record = None
    resp = make_response(jsonify({'ok': True}))
    resp.set_cookie(SESSION_COOKIE, '', expires=0, secure=COOKIE_SECURE)
    resp.set_cookie(CSRF_COOKIE, '', expires=0, secure=COOKIE_SECURE)
    return resp

@app.route('/api/auth/me')
def auth_me():
    u = current_user()
    if not u:
        return jsonify({'error': 'unauthorized'}), 401
    return jsonify({
        'email': u['email'],
        'name': u.get('name', ''),
        'dept_id': u.get('dept_id'),
        'role': u.get('role'),
        'campus': u.get('campus'),
        'district': u.get('district'),
    })

@app.route('/api/auth/options')
def auth_options():
    campuses = set()
    districts = set()
    for dept in load_depts_config():
        dept_id = dept.get('id')
        if not dept_id:
            continue
        cfg = get_campus_config(dept_id)
        for district, names in cfg.get('districts', {}).items():
            districts.add(district)
            campuses.update(names or [])
        for term in load_terms(dept_id):
            term_id = term.get('id')
            if not term_id:
                continue
            for course in load_term_data(dept_id, term_id):
                if course.get('campus'):
                    campuses.add(course['campus'])
    return jsonify({
        'roles': [
            {'id': 'user', 'name': '普通用户'},
            {'id': 'jiaowu', 'name': '教务（跨部门）'},
            {'id': 'store_manager', 'name': '店长'},
            {'id': 'supervisor', 'name': '主管'},
            {'id': 'regional_manager', 'name': '大区经理'},
            {'id': 'director', 'name': '总监'},
            {'id': 'admin', 'name': '管理员'},
        ],
        'campuses': sorted(campuses),
        'districts': sorted(districts),
    })


# ==================== 今日排课心情墙 ====================
@app.route('/api/mood-board')
@require_auth
def mood_board_get():
    date_text = mood_today()
    board = normalize_mood_board(sqlite_store.get_document(mood_namespace(date_text), {'date': date_text, 'entries': {}}), date_text)
    return jsonify(mood_board_payload(board, request.user))


@app.route('/api/mood-board', methods=['POST'])
@require_auth
def mood_board_update():
    body = request.json or {}
    try:
        payload = update_today_mood_for_user(
            request.user,
            body.get('level'),
            body.get('note') or '',
        )
    except AuthUpdateError as exc:
        return jsonify({'error': exc.message}), exc.status_code
    return jsonify(payload)


@app.route('/api/mood-board', methods=['DELETE'])
@require_auth
def mood_board_delete():
    return jsonify(delete_today_mood_for_user(request.user))


# ==================== 邀请码管理（管理员） ====================
@app.route('/api/auth/invites', methods=['GET'])
@require_admin
def list_invites():
    invites = load_invites()
    depts = {d['id']: d['name'] for d in load_depts_config()}
    for c in invites:
        c['dept_name'] = depts.get(c.get('dept_id') or '', '')
    return jsonify(invites)

@app.route('/api/auth/invites', methods=['POST'])
@require_admin
def create_invites():
    body = request.json or {}
    role = body.get('role') or 'user'
    count = int(body.get('count') or 1)
    if role not in VALID_ROLES:
        return jsonify({'error': '角色不正确'}), 400
    count = max(1, min(50, count))

    def updater(invites):
        new_codes = []
        new_invites = list(invites)
        existing = {c['code'] for c in invites}
        for _ in range(count):
            code = gen_invite_code(12)
            while code in existing:
                code = gen_invite_code(12)
            existing.add(code)
            rec = {
                'code': code, 'dept_id': None, 'role': role,
                'campus': None,
                'district': None,
                'created_at': time.strftime('%Y-%m-%d %H:%M:%S'),
                'created_by': request.user['email'],
                'used_by': None, 'used_at': None,
            }
            new_invites.append(rec)
            new_codes.append(rec)
        return new_invites, new_codes

    new_codes = update_invites_atomic(updater)
    return jsonify({'ok': True, 'codes': new_codes})

@app.route('/api/auth/invites/<code>', methods=['DELETE'])
@require_admin
def delete_invite(code):
    code = (code or '').strip().upper()

    def updater(invites):
        before = len(invites)
        new_invites = [c for c in invites if not (c['code'] == code and not c.get('used_by'))]
        if len(new_invites) == before:
            raise AuthUpdateError('无法删除（已使用或不存在）', 400)
        return new_invites, {'ok': True}

    try:
        update_invites_atomic(updater)
    except AuthUpdateError as exc:
        return jsonify({'error': exc.message}), exc.status_code
    return jsonify({'ok': True})


@app.route('/api/admin/overview')
@require_admin
def admin_overview():
    depts = load_depts_config()
    users = load_users()
    invites = load_invites()
    dept_names = {d['id']: d.get('name', d['id']) for d in depts}

    dept_rows = []
    term_rows = []
    total_courses = 0
    teacher_names = set()
    campus_names = set()
    room_names = set()

    for dept in depts:
        dept_id = dept.get('id')
        if not dept_id:
            continue
        terms = load_terms(dept_id)
        dept_course_count = 0
        dept_teachers = set()
        dept_campuses = set()
        dept_rooms = set()
        active_terms = 0

        for term in terms:
            term_id = term.get('id')
            if not term_id:
                continue
            courses = load_term_data(dept_id, term_id)
            metadata = load_term_metadata(dept_id, term_id)
            workflow = load_workflow(dept_id, term_id)
            backup = latest_daily_backup_info(dept_id, term_id)
            teachers = {c.get('teacher') for c in courses if c.get('teacher')}
            campuses = {c.get('campus') for c in courses if c.get('campus')}
            rooms = {c.get('room') for c in courses if c.get('room')}
            dept_course_count += len(courses)
            dept_teachers.update(teachers)
            dept_campuses.update(campuses)
            dept_rooms.update(rooms)
            total_courses += len(courses)
            teacher_names.update(teachers)
            campus_names.update(campuses)
            room_names.update(rooms)
            if not term.get('archived'):
                active_terms += 1
            version = sqlite_store.get_term_version(dept_id, term_id)
            term_rows.append({
                'dept_id': dept_id,
                'dept_name': dept_names.get(dept_id, dept_id),
                'term_id': term_id,
                'term_name': term.get('name') or term_id,
                'description': term.get('description', ''),
                'archived': bool(term.get('archived')),
                'is_default': bool(term.get('is_default')),
                'courses_count': len(courses),
                'original_count': metadata.get('original_count', 0),
                'teachers_count': len(teachers),
                'campuses_count': len(campuses),
                'rooms_count': len(rooms),
                'workflow_status': workflow.get('status', 'draft'),
                'updated_at': metadata.get('updated_at', ''),
                'updated_by': metadata.get('updated_by', ''),
                'version': version.get('version'),
                'latest_backup_at': (backup or {}).get('saved_at', ''),
            })

        dept_rows.append({
            'id': dept_id,
            'name': dept_names.get(dept_id, dept_id),
            'description': dept.get('description', ''),
            'terms_count': len(terms),
            'active_terms_count': active_terms,
            'courses_count': dept_course_count,
            'teachers_count': len(dept_teachers),
            'campuses_count': len(dept_campuses),
            'rooms_count': len(dept_rooms),
        })

    role_counts = {}
    for user in users:
        role = user.get('role') or 'user'
        role_counts[role] = role_counts.get(role, 0) + 1

    invite_counts = {
        'total': len(invites),
        'unused': len([c for c in invites if not c.get('used_by')]),
        'used': len([c for c in invites if c.get('used_by')]),
    }
    db_exists = os.path.exists(sqlite_store.DB_PATH)
    return jsonify({
        'summary': {
            'departments_count': len(dept_rows),
            'terms_count': len(term_rows),
            'courses_count': total_courses,
            'teachers_count': len(teacher_names),
            'campuses_count': len(campus_names),
            'rooms_count': len(room_names),
            'users_count': len(users),
            'invite_unused_count': invite_counts['unused'],
        },
        'departments': dept_rows,
        'terms': sorted(term_rows, key=lambda x: (x['dept_name'], x['term_name'])),
        'users': {
            'total': len(users),
            'role_counts': role_counts,
            'items': [public_user_info(u, dept_names) for u in users],
        },
        'invites': invite_counts,
        'database': {
            'path': sqlite_store.DB_PATH,
            'exists': db_exists,
            'size': os.path.getsize(sqlite_store.DB_PATH) if db_exists else 0,
        },
        'server_time': time.strftime('%Y-%m-%d %H:%M:%S'),
    })


@app.route('/api/admin/users')
@require_admin
def admin_users():
    dept_names = {d['id']: d.get('name', d['id']) for d in load_depts_config()}
    return jsonify([public_user_info(u, dept_names) for u in load_users()])


@app.route('/api/admin/users/<user_id>', methods=['PATCH'])
@require_admin
def admin_update_user(user_id):
    body = request.json or {}
    role = (body.get('role') or '').strip()
    dept_id = (body.get('dept_id') or '').strip()
    campus = (body.get('campus') or '').strip()
    scope, err = normalize_user_scope(role, dept_id, campus)
    if err:
        return jsonify(err), 400

    def updater(users):
        new_users = [dict(u) for u in users]
        target = next((u for u in new_users if u.get('id') == user_id), None)
        if not target:
            raise AuthUpdateError('用户不存在', 404)
        target['role'] = role
        target['dept_id'] = scope['dept_id']
        target['campus'] = scope['campus']
        target['district'] = scope['district']
        target['updated_at'] = time.strftime('%Y-%m-%d %H:%M:%S')
        target['updated_by'] = request.user.get('email')
        return new_users, {'ok': True, 'user': public_user_info(target)}

    try:
        result = update_users_atomic(updater)
    except AuthUpdateError as exc:
        return jsonify({'error': exc.message}), exc.status_code
    return jsonify(result)


@app.route('/api/admin/users/<user_id>/reset-password', methods=['POST'])
@require_admin
def admin_reset_password(user_id):
    body = request.json or {}
    new_password = (body.get('password') or '').strip()
    if len(new_password) < 6:
        return jsonify({'error': '密码至少6位'}), 400

    def updater(users):
        new_users = [dict(u) for u in users]
        target = next((u for u in new_users if u.get('id') == user_id), None)
        if not target:
            raise AuthUpdateError('用户不存在', 404)
        set_user_password(target, new_password)
        target['updated_at'] = time.strftime('%Y-%m-%d %H:%M:%S')
        target['updated_by'] = request.user.get('email')
        return new_users, {'ok': True, 'user': public_user_info(target)}

    try:
        result = update_users_atomic(updater)
    except AuthUpdateError as exc:
        return jsonify({'error': exc.message}), exc.status_code
    sqlite_store.delete_sessions_for_user(user_id)
    return jsonify(result)


@app.route('/api/admin/users/<user_id>', methods=['DELETE'])
@require_admin
def admin_delete_user(user_id):
    if request.user.get('id') == user_id:
        return jsonify({'error': '不能删除当前登录账号'}), 400

    def updater(users):
        if not any(u.get('id') == user_id for u in users):
            raise AuthUpdateError('用户不存在', 404)
        return [u for u in users if u.get('id') != user_id], {'ok': True}

    try:
        update_users_atomic(updater)
    except AuthUpdateError as exc:
        return jsonify({'error': exc.message}), exc.status_code
    sqlite_store.delete_sessions_for_user(user_id)
    return jsonify({'ok': True})


# ==================== 部门管理 API ====================
@app.route('/api/departments')
@require_auth
def list_departments():
    depts = load_depts_config()
    u = request.user
    if u.get('role') != 'admin' and u.get('dept_id'):
        depts = [d for d in depts if d['id'] == u['dept_id']]
    include_counts = str(request.args.get('include_counts') or '').lower() in {'1', 'true', 'yes'}
    out = []
    for d in depts:
        terms = load_terms(d['id'])
        row = {**d, 'terms_count': len(terms)}
        if include_counts:
            row['count'] = sum(load_term_metadata(d['id'], t['id']).get('count', 0) for t in terms)
        out.append(row)
    return jsonify(out)

@app.route('/api/departments', methods=['POST'])
@require_admin
def create_department():
    body = request.json or {}
    dept_id = (body.get('id') or '').strip()
    name = (body.get('name') or '').strip()
    desc = (body.get('description') or '').strip()
    if not dept_id or not name:
        return jsonify({'error': '缺少ID或名称'}), 400
    if not re.match(r'^[a-z0-9_-]+$', dept_id):
        return jsonify({'error': 'ID格式不正确'}), 400

    def updater(depts):
        if any(d['id'] == dept_id for d in depts):
            raise ConfigUpdateError('部门ID已存在', 400)
        new_depts = list(depts)
        new_depts.append({'id': dept_id, 'name': name, 'description': desc})
        return new_depts, {'ok': True}

    try:
        update_json_document_atomic(DEPTS_CONFIG, [], updater)
    except ConfigUpdateError as exc:
        return jsonify({'error': exc.message}), exc.status_code
    get_dept_dir(dept_id)
    return jsonify({'ok': True})


# ==================== 学期管理 API ====================
@app.route('/dept/<dept_id>/api/terms')
def dept_list_terms(dept_id):
    _, err = check_dept_access(dept_id)
    if err: return err
    include_archived = request.args.get('include_archived') == '1'
    terms = load_terms(dept_id)
    if not include_archived:
        terms = [t for t in terms if not t.get('archived')]
    return jsonify(terms)

@app.route('/dept/<dept_id>/api/terms', methods=['POST'])
def dept_create_term(dept_id):
    user, err = check_dept_access(dept_id)
    if err: return err
    if user.get('role') not in STAFF_ROLES:
        return jsonify({'error': 'forbidden'}), 403
    body = request.json or {}
    term_id = (body.get('id') or '').strip()
    name = (body.get('name') or '').strip()
    desc = (body.get('description') or '').strip()
    if not name and dept_id == 'yiduiyi':
        name = yiduiyi_monthly_term_name()
    if not name:
        return jsonify({'error': '缺少批次名称'}), 400
    requested_term_id = term_id
    if requested_term_id and not re.match(r'^[a-z0-9_-]+$', requested_term_id):
        return jsonify({'error': 'ID格式不正确（小写字母/数字/下划线/横线）'}), 400
    if not dept_exists(dept_id):
        return jsonify({'error': '部门不存在'}), 400
    def updater(terms):
        final_term_id = requested_term_id or make_term_id_from_terms(dept_id, terms)
        if any(t['id'] == final_term_id for t in terms):
            raise ConfigUpdateError('批次ID已存在', 400)
        new_terms = list(terms)
        new_terms.append({
            'id': final_term_id, 'name': name, 'description': desc, 'archived': False, 'is_default': not terms,
            'created_at': time.strftime('%Y-%m-%d %H:%M:%S'),
        })
        return new_terms, {'ok': True, 'id': final_term_id}

    try:
        result = update_json_document_atomic(dept_terms_file(dept_id), [], updater)
    except ConfigUpdateError as exc:
        return jsonify({'error': exc.message}), exc.status_code
    term_id = result['id']
    get_term_dir(dept_id, term_id)
    save_json(term_data_file(dept_id, term_id), [])
    save_json(term_original_file(dept_id, term_id), [])
    save_term_metadata(dept_id, term_id, user=user)
    return jsonify({'ok': True, 'id': term_id})

@app.route('/dept/<dept_id>/api/terms/<term_id>', methods=['PATCH'])
def dept_update_term(dept_id, term_id):
    user, err = check_dept_access(dept_id)
    if err: return err
    if user.get('role') not in STAFF_ROLES:
        return jsonify({'error': 'forbidden'}), 403
    body = request.json or {}

    def updater(terms):
        new_terms = [dict(t) for t in terms]
        target = next((t for t in new_terms if t.get('id') == term_id), None)
        if not target:
            raise ConfigUpdateError('批次不存在', 404)
        if 'archived' in body:
            target['archived'] = bool(body.get('archived'))
        if body.get('is_default'):
            for t in new_terms:
                t['is_default'] = t.get('id') == term_id
        if 'name' in body and str(body.get('name') or '').strip():
            target['name'] = str(body.get('name')).strip()
        if 'description' in body:
            target['description'] = str(body.get('description') or '').strip()
        return new_terms, {'ok': True, 'term': target}

    try:
        result = update_json_document_atomic(dept_terms_file(dept_id), [], updater)
    except ConfigUpdateError as exc:
        return jsonify({'error': exc.message}), exc.status_code
    return jsonify(result)

@app.route('/dept/<dept_id>/api/terms/<term_id>', methods=['DELETE'])
def dept_delete_term(dept_id, term_id):
    user, err = check_dept_access(dept_id)
    if err: return err
    if user.get('role') not in STAFF_ROLES:
        return jsonify({'error': 'forbidden'}), 403

    def updater(terms):
        target = next((t for t in terms if t.get('id') == term_id), None)
        if not target:
            raise ConfigUpdateError('批次不存在', 404)
        if len(terms) <= 1:
            raise ConfigUpdateError('至少保留一个批次', 400)
        remaining = [dict(t) for t in terms if t.get('id') != term_id]
        if target.get('is_default') and remaining:
            next_default = next((t for t in remaining if not t.get('archived')), remaining[0])
            next_default['is_default'] = True
        return remaining, {'ok': True, 'deleted': dict(target)}

    try:
        result = update_json_document_atomic(dept_terms_file(dept_id), [], updater)
    except ConfigUpdateError as exc:
        return jsonify({'error': exc.message}), exc.status_code

    term_dir = os.path.join(get_dept_dir(dept_id), 'terms', term_id)
    archived_dir = ''
    if os.path.exists(term_dir):
        trash_root = os.path.join(get_dept_dir(dept_id), 'deleted_terms')
        os.makedirs(trash_root, exist_ok=True)
        archived_dir = os.path.join(trash_root, f'{term_id}_{time.strftime("%Y%m%d_%H%M%S")}')
        shutil.move(term_dir, archived_dir)
    return jsonify({'ok': True, 'deleted': result['deleted'], 'archived_dir': archived_dir})


# ==================== 资源管理 API ====================
def infer_teacher_campuses(dept_id):
    mapping = {}
    for term in load_terms(dept_id):
        for c in load_term_data(dept_id, term['id']):
            teacher = c.get('teacher')
            campus = c.get('campus')
            if teacher and campus:
                mapping.setdefault(teacher, set()).add(campus)
    return {k: sorted(v) for k, v in mapping.items()}

def workbook_sheet(file_storage, preferred='教师&教室名单'):
    try:
        import openpyxl
        import zipfile
        from openpyxl.utils.exceptions import InvalidFileException
    except ImportError:
        return None, {'error': 'openpyxl not installed', 'status': 500}
    try:
        wb = openpyxl.load_workbook(file_storage, data_only=True, read_only=True)
    except (InvalidFileException, OSError, ValueError, zipfile.BadZipFile):
        return None, {'error': 'Excel 文件无法读取', 'status': 400}
    if preferred in wb.sheetnames:
        return wb[preferred], None
    return wb[wb.sheetnames[0]], None

def sheet_rows_by_header(ws):
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [str(c).strip() if c is not None else '' for c in next(rows)]
    except StopIteration:
        return [], []
    out = []
    for row in rows:
        rec = {}
        for i, h in enumerate(headers):
            if h:
                rec[h] = row[i] if i < len(row) else None
        out.append(rec)
    return headers, out

@app.route('/dept/<dept_id>/api/resources/teachers')
def resources_teachers(dept_id):
    _, err = check_dept_access(dept_id)
    if err: return err
    campus = request.args.get('campus', '')
    subject = request.args.get('subject', '')
    teachers = load_teachers(dept_id)
    if campus:
        teachers = [t for t in teachers if campus in (t.get('campuses') or [t.get('campus')])]
    if subject:
        teachers = [t for t in teachers if t.get('subject') == subject]
    return jsonify(teachers)

@app.route('/dept/<dept_id>/api/resources/teachers/import', methods=['POST'])
@require_staff
def resources_teachers_import(dept_id):
    _, err = check_dept_access(dept_id)
    if err: return err
    if 'file' not in request.files:
        return jsonify({'error': 'no file'}), 400
    ws, err_info = workbook_sheet(request.files['file'])
    if err_info: return jsonify({'error': err_info['error']}), err_info['status']
    _, rows = sheet_rows_by_header(ws)
    campus_map = infer_teacher_campuses(dept_id)
    teachers = {}
    for row in rows:
        subject = str(row.get('科目') or row.get('学科') or '').strip()
        name = str(row.get('教师姓名') or row.get('教师') or row.get('姓名') or '').strip()
        if not name or not subject:
            continue
        teachers[name] = {'name': name, 'subject': subject, 'campuses': campus_map.get(name, [])}
    teacher_list = sorted(teachers.values(), key=lambda x: (x.get('subject', ''), x['name']))
    update_json_document_atomic(
        teachers_file(dept_id),
        [],
        lambda _current: (teacher_list, {'ok': True}),
    )
    return jsonify({'ok': True, 'count': len(teachers)})

@app.route('/dept/<dept_id>/api/resources/classrooms')
def resources_classrooms(dept_id):
    _, err = check_dept_access(dept_id)
    if err: return err
    campus = request.args.get('campus', '')
    room_type = request.args.get('type', '')
    rooms = load_classrooms(dept_id)
    if campus:
        rooms = [r for r in rooms if r.get('campus') == campus]
    if room_type:
        rooms = [r for r in rooms if r.get('type') == room_type]
    return jsonify(rooms)

@app.route('/dept/<dept_id>/api/resources/classrooms/import', methods=['POST'])
@require_staff
def resources_classrooms_import(dept_id):
    _, err = check_dept_access(dept_id)
    if err: return err
    if 'file' not in request.files:
        return jsonify({'error': 'no file'}), 400
    ws, err_info = workbook_sheet(request.files['file'])
    if err_info: return jsonify({'error': err_info['error']}), err_info['status']
    _, rows = sheet_rows_by_header(ws)
    rooms = []
    districts = {}
    for row in rows:
        campus = str(row.get('校区') or row.get('校区名称') or '').strip()
        name = str(row.get('教室名称') or row.get('教室') or '').strip()
        if not campus or not name:
            continue
        district = str(row.get('大区') or '').strip()
        if district:
            districts.setdefault(district, [])
            if campus not in districts[district]:
                districts[district].append(campus)
        capacity = row.get('座位数') or row.get('容量') or row.get('标准人数') or ''
        rooms.append({
            'name': name,
            'campus': campus,
            'capacity': capacity,
            'type': str(row.get('教室分类') or row.get('类型') or '').strip(),
        })
    if districts:
        def updater(values):
            config = values.get(campus_config_file(dept_id)) or get_campus_config(dept_id)
            next_config = {
                'districts': {**config.get('districts', {}), **districts},
                'campus_codes': config.get('campus_codes', {}),
            }
            return {
                classrooms_file(dept_id): rooms,
                campus_config_file(dept_id): next_config,
            }, {'ok': True}

        update_json_documents_atomic(
            {
                classrooms_file(dept_id): [],
                campus_config_file(dept_id): DEFAULT_CAMPUS_CONFIG,
            },
            updater,
        )
    else:
        update_json_document_atomic(
            classrooms_file(dept_id),
            [],
            lambda _current: (rooms, {'ok': True}),
        )
    return jsonify({'ok': True, 'count': len(rooms)})

@app.route('/dept/<dept_id>/api/resources/campus-config')
def resources_campus_config(dept_id):
    _, err = check_dept_access(dept_id)
    if err: return err
    return jsonify(get_campus_config(dept_id))

@app.route('/dept/<dept_id>/api/resources/campus-config', methods=['PUT'])
@require_admin
def resources_campus_config_update(dept_id):
    _, err = check_dept_access(dept_id)
    if err: return err
    body = request.json or {}
    config = {
        'districts': body.get('districts') or {},
        'campus_codes': body.get('campus_codes') or {},
    }
    update_json_document_atomic(
        campus_config_file(dept_id),
        DEFAULT_CAMPUS_CONFIG,
        lambda _current: (config, {'ok': True}),
    )
    return jsonify({'ok': True})


# ==================== 部门入口 / 学期入口 ====================
@app.route('/dept/<dept_id>/')
def dept_redirect(dept_id):
    u = current_user()
    if not u:
        return redirect('/auth')
    if not can_access_dept_page(u, dept_id):
        return redirect('/?denied=1')
    if not dept_exists(dept_id):
        return redirect('/?missing=1')
    default_term = get_default_term(dept_id)
    if not default_term:
        save_terms(dept_id, [{
            'id': DEFAULT_TERM_ID, 'name': DEFAULT_TERM_NAME,
            'description': '默认批次',
            'created_at': time.strftime('%Y-%m-%d %H:%M:%S'),
        }])
        get_term_dir(dept_id, DEFAULT_TERM_ID)
        save_json(term_data_file(dept_id, DEFAULT_TERM_ID), [])
        save_json(term_original_file(dept_id, DEFAULT_TERM_ID), [])
        save_term_metadata(dept_id, DEFAULT_TERM_ID, user=u)
        default_term = DEFAULT_TERM_ID
    return redirect(f'/dept/{dept_id}/{default_term}/')

@app.route('/dept/<dept_id>/<term_id>/')
def dept_term_index(dept_id, term_id):
    u = current_user()
    if not u:
        return redirect('/auth')
    if not can_access_dept_page(u, dept_id):
        return redirect('/?denied=1')
    if not dept_exists(dept_id):
        return redirect('/?missing=1')
    if not term_exists(dept_id, term_id):
        return redirect(f'/dept/{dept_id}/')
    return serve_project_file('index.html', 'text/html')

@app.route('/dept/<dept_id>/<term_id>/app.js')
def dept_term_js(dept_id, term_id):
    return serve_project_file('app.js', 'application/javascript', cache_seconds=60)

@app.route('/dept/<dept_id>/<term_id>/app_utils.js')
def dept_term_utils_js(dept_id, term_id):
    return serve_project_file('app_utils.js', 'application/javascript', cache_seconds=60)


# ==================== 学期数据 API ====================
def _check_dt(dept_id, term_id):
    u, err = check_dept_access(dept_id)
    if err: return None, err
    if not dept_exists(dept_id):
        return None, (jsonify({'error': '部门不存在'}), 404)
    if not term_exists(dept_id, term_id):
        return None, (jsonify({'error': '批次不存在'}), 404)
    return u, None

@app.route('/dept/<dept_id>/<term_id>/api/courses')
def t_get_courses(dept_id, term_id):
    _, err = _check_dt(dept_id, term_id)
    if err: return err
    own_version = sqlite_store.get_term_version(dept_id, term_id).get('version')
    related_signature = tuple(cross_dept_room_index_signature(term_id)) if dept_id in SHARED_CLASSROOM_DEPTS else ()
    etag = data_api_etag('courses', dept_id, term_id, own_version, related_signature)
    headers = {'X-Data-Version': f"sqlite:{own_version}"}
    cached = conditional_not_modified_response(etag, headers)
    if cached:
        return cached
    courses = load_term_data(dept_id, term_id)
    related = related_room_courses_for_dept(dept_id, term_id)
    return conditional_json_response({
        'courses': courses,
        'related_room_courses': related['courses'],
        'related_rooms': related['rooms'],
    }, etag, headers)

@app.route('/dept/<dept_id>/<term_id>/api/original')
def t_get_original(dept_id, term_id):
    _, err = _check_dt(dept_id, term_id)
    if err: return err
    version = sqlite_store.get_term_version(dept_id, term_id).get('version')
    etag = data_api_etag('original', dept_id, term_id, version)
    cached = conditional_not_modified_response(etag, {'X-Data-Version': f"sqlite:{version}"})
    if cached:
        return cached
    return conditional_json_response(load_term_original(dept_id, term_id), etag, {'X-Data-Version': f"sqlite:{version}"})

@app.route('/dept/<dept_id>/<term_id>/api/preview-code')
def t_preview_code(dept_id, term_id):
    _, err = _check_dt(dept_id, term_id)
    if err: return err
    from code_generator import generate_code, extract_target_fy
    courses = load_term_data(dept_id, term_id)
    code = generate_code(
        dept_id,
        request.args.get('season', ''),
        request.args.get('subject', ''),
        request.args.get('grade', ''),
        request.args.get('level', ''),
        request.args.get('campus', ''),
        [c.get('code', '') for c in courses],
        config=get_campus_config(dept_id),
        fy=extract_target_fy(term_id),
    )
    return jsonify({'code': code})

@app.route('/dept/<dept_id>/<term_id>/api/check-conflict')
def t_check_conflict(dept_id, term_id):
    _, err = _check_dt(dept_id, term_id)
    if err: return err
    courses = load_term_data(dept_id, term_id)
    result = check_course_conflict(
        dept_id, term_id, courses,
        request.args.get('teacher', ''), request.args.get('room', ''),
        request.args.get('campus', ''), request.args.get('season', ''),
        request.args.get('period', ''), request.args.get('slot', ''),
        request.args.get('day', ''),
    )
    return jsonify(result)

@app.route('/dept/<dept_id>/<term_id>/api/available-slots')
def t_available_slots(dept_id, term_id):
    _, err = _check_dt(dept_id, term_id)
    if err: return err
    campus = request.args.get('campus', '')
    season = request.args.get('season', '')
    period_filter = request.args.get('period', '')
    subject = request.args.get('subject', '')
    if not campus or not season:
        return jsonify({'error': 'campus 和 season 必填'}), 400
    courses = load_term_data(dept_id, term_id)
    related = related_room_courses_for_dept(dept_id, term_id)
    occupancy_courses = active_courses(courses) + active_courses(related['courses'])
    def add_occupancy(index, key, item):
        if not key:
            return
        bucket = index.setdefault(key, {'items': [], 'used': set()})
        bucket['items'].append(item)
        bucket['used'].add((item['period'], item.get('day', ''), item['slot']))

    teacher_occupancy = {}
    room_occupancy = {}
    for c in occupancy_courses:
        if c.get('season') != season:
            continue
        if period_filter and c.get('period') != period_filter:
            continue
        if not c.get('period') or not c.get('slot'):
            continue
        item = {
            'period': c.get('period'),
            'slot': c.get('slot'),
            'day': c.get('day', ''),
            'course_name': c.get('name', ''),
        }
        add_occupancy(teacher_occupancy, c.get('teacher'), item)
        if c.get('campus') == campus and c.get('room'):
            add_occupancy(room_occupancy, c.get('room'), item)
        related_key = c.get('room_key') if c.get('related') else None
        add_occupancy(room_occupancy, related_key, item)
    periods = sorted({c.get('period') for c in courses if c.get('season') == season and c.get('period')})
    if period_filter:
        periods = [period_filter]
    days = sorted({c.get('day') for c in courses if c.get('season') == season and c.get('day')}) or ['周五', '周六', '周日']
    slots = slot_order_for(dept_id)
    combos = [
        {'period': p, 'day': d, 'slot': s}
        for p in periods
        for d in days
        for s in schedulable_slots_for_period_day(dept_id, p, d)
    ]

    teachers = load_teachers(dept_id)
    if not teachers:
        teacher_map = {}
        for c in active_courses(courses):
            if c.get('teacher') and c.get('campus') == campus:
                teacher_map[c['teacher']] = {'name': c['teacher'], 'subject': c.get('subject', ''), 'campuses': [campus]}
        teachers = list(teacher_map.values())
    teachers = [
        t for t in teachers
        if campus in (t.get('campuses') or [t.get('campus')]) and (not subject or t.get('subject') == subject)
    ]
    classrooms = load_classrooms(dept_id)
    if not classrooms:
        room_map = {}
        for c in courses:
            if c.get('room') and c.get('campus') == campus:
                room_map[c['room']] = {'name': c['room'], 'campus': campus, 'capacity': c.get('capacity', ''), 'type': ''}
        classrooms = list(room_map.values())
    classrooms = [r for r in classrooms if r.get('campus') == campus]

    def empty_bucket():
        return {'items': [], 'used': set()}

    def merge_room_bucket(room):
        direct = room_occupancy.get(room) or empty_bucket()
        key = shared_room_key(campus, room)
        shared = room_occupancy.get(key) if key else None
        if not shared:
            return direct
        return {
            'items': direct['items'] + shared['items'],
            'used': direct['used'] | shared['used'],
        }

    def available_from(bucket):
        used = bucket['used']
        return [c for c in combos if (c['period'], c['day'], c['slot']) not in used]

    teacher_out = []
    for t in teachers:
        bucket = teacher_occupancy.get(t['name']) or empty_bucket()
        teacher_out.append({**t, 'occupied': bucket['items'], 'available': available_from(bucket)})
    room_out = []
    for r in classrooms:
        bucket = merge_room_bucket(r['name'])
        room_out.append({**r, 'occupied': bucket['items'], 'available': available_from(bucket)})
    return jsonify({'teachers': teacher_out, 'classrooms': room_out, 'periods': periods, 'days': days, 'slots': slots})

@app.route('/dept/<dept_id>/<term_id>/api/courses', methods=['POST'])
def t_create_course(dept_id, term_id):
    user, err = _check_dt(dept_id, term_id)
    if err: return err
    wf_err = check_workflow_permission(dept_id, term_id, user)
    if wf_err: return wf_err
    conflict = term_version_conflict(dept_id, term_id)
    if conflict: return conflict
    body = request.json or {}
    reason = str(body.get('reason') or '').strip()
    campus = (body.get('campus') or '').strip()
    if not can_create_course(user, campus, dept_id):
        return jsonify({'error': '无权在该校区新增课程'}), 403
    required = ['season', 'campus', 'subject', 'grade', 'level', 'teacher', 'room', 'period', 'slot', 'day']
    missing = [k for k in required if not (body.get(k) or '').strip()]
    if missing:
        return jsonify({'error': '缺少字段：' + '、'.join(missing)}), 400
    courses = load_term_data(dept_id, term_id)
    result = check_course_conflict(
        dept_id, term_id, courses, body.get('teacher'), body.get('room'), campus,
        body.get('season'), body.get('period'), body.get('slot'), body.get('day')
    )
    if result['teacher_conflict'] or result['room_conflict'] or result.get('shared_room_conflict'):
        return jsonify({'error': '新增课程存在冲突', 'conflict_type': 'mixed', **result}), 409
    from code_generator import generate_code, extract_target_fy
    new_id = max([int(c.get('id', -1)) for c in courses if str(c.get('id', '')).isdigit()] or [-1]) + 1
    code = generate_code(
        dept_id, body.get('season'), body.get('subject'), body.get('grade'),
        body.get('level'), campus, [c.get('code', '') for c in courses],
        config=get_campus_config(dept_id),
        fy=extract_target_fy(term_id),
    )
    slot = body.get('slot')
    course = {
        'id': new_id,
        'season': body.get('season'),
        'campus': campus,
        'code': code,
        'name': body.get('name') or f"{body.get('subject')}{body.get('grade')}{body.get('level')}插空班",
        'teacher': body.get('teacher'),
        'period': body.get('period'),
        'timeDesc': body.get('timeDesc') or f"{body.get('day')} {get_time_range(dept_id, slot)}",
        'timeRange': get_time_range(dept_id, slot),
        'slot': slot,
        'day': body.get('day'),
        'startDate': body.get('startDate', ''),
        'endDate': body.get('endDate', ''),
        'room': body.get('room'),
        'sessions': body.get('sessions') or '',
        'department': body.get('department', ''),
        'capacity': body.get('capacity') or '',
        'desc': body.get('desc') or '插空新增',
        'sourceCode': body.get('sourceCode', ''),
        'currentCount': body.get('currentCount', ''),
        'subject': body.get('subject'),
        'classType': body.get('level'),
        'created_by_action': 'insert_course',
        'created_at': time.strftime('%Y-%m-%d %H:%M:%S'),
        'created_by': user.get('email'),
    }
    old_courses = [dict(c) for c in courses]
    courses.append(course)
    save_err = save_term_data_or_conflict(dept_id, term_id, courses, user=user, old_courses=old_courses, action='新增插空排课', reason=reason)
    if save_err: return save_err
    log = load_changelog(dept_id, term_id)
    log.append({
        'time': time.strftime('%Y-%m-%d %H:%M:%S'),
        'user': user.get('name') or user.get('email'),
        'email': user.get('email'),
        'action': '新增插空排课',
        'reason': reason,
        'changes': [{
            'course_id': new_id, 'code': code, 'name': course['name'],
            'field': 'create', 'from': '', 'to': '新增插空课程',
        }],
    })
    save_changelog(dept_id, term_id, log[-500:])
    return jsonify_with_term_version({'ok': True, 'course': course, 'code': code}, dept_id, term_id)

@app.route('/dept/<dept_id>/<term_id>/api/version')
def t_get_version(dept_id, term_id):
    _, err = _check_dt(dept_id, term_id)
    if err: return err
    return jsonify(load_term_version_metadata(dept_id, term_id))

@app.route('/dept/<dept_id>/<term_id>/api/workflow')
def t_get_workflow(dept_id, term_id):
    _, err = _check_dt(dept_id, term_id)
    if err: return err
    return jsonify(load_workflow(dept_id, term_id))

@app.route('/dept/<dept_id>/<term_id>/api/presence', methods=['POST'])
def t_update_presence(dept_id, term_id):
    user, err = _check_dt(dept_id, term_id)
    if err: return err
    body = request.json or {}
    user_id = user.get('id') or user.get('email')
    users = sqlite_store.update_presence(
        dept_id,
        term_id,
        user_id,
        {
            'name': user.get('name') or user.get('email') or '',
            'email': user.get('email', ''),
            'role': user.get('role', ''),
            'campus': user.get('campus', ''),
            'cursor': str(body.get('cursor') or '')[:80],
            'activity': str(body.get('activity') or '')[:24],
            'tab': str(body.get('tab') or '')[:32],
            'course_id': str(body.get('course_id') or body.get('courseId') or '')[:80],
            'field': str(body.get('field') or '')[:40],
        },
        ttl=PRESENCE_TTL,
    )
    return jsonify({'users': users, 'ttl': PRESENCE_TTL})

@app.route('/dept/<dept_id>/<term_id>/api/workflow', methods=['PATCH'])
def t_update_workflow(dept_id, term_id):
    user, err = _check_dt(dept_id, term_id)
    if err: return err
    body = request.json or {}
    new_status = body.get('status')
    try:
        return jsonify(update_workflow_atomic(dept_id, term_id, user, new_status))
    except WorkflowUpdateError as exc:
        return jsonify({'error': exc.message}), exc.status_code

@app.route('/dept/<dept_id>/<term_id>/api/courses/<int:course_id>', methods=['PATCH'])
def t_update_course(dept_id, term_id, course_id):
    user, err = _check_dt(dept_id, term_id)
    if err: return err
    wf_err = check_workflow_permission(dept_id, term_id, user)
    if wf_err: return wf_err
    updates = request.json or {}
    reason = str(updates.get('reason') or '').strip()
    fields = course_patch_fields(updates)
    base_fields = course_patch_base_fields(updates)
    soft_allow_room_conflict = should_soft_allow_room_conflict(updates)
    version_problem = missing_data_version_problem(dept_id, term_id)
    if version_problem:
        payload, status = version_problem
        return jsonify(payload), status

    last_conflict = None
    for attempt in range(COURSE_UPDATE_RETRY_LIMIT):
        current_version = current_term_version_number(dept_id, term_id)
        courses = load_term_data(dept_id, term_id)
        idx = find_course_index(courses, course_id)
        if idx is None:
            return jsonify({'error': 'not found'}), 404
        if not can_edit_course(user, courses[idx], dept_id):
            return jsonify({'error': '无权编辑该课程'}), 403
        if not is_active_course(courses[idx]):
            return jsonify({'error': '已取消或已合并的班级需先恢复后再编辑'}), 400

        field_problem = field_level_conflict_problem(
            dept_id, term_id, courses[idx], fields, base_fields, current_version
        )
        if field_problem:
            payload, status = field_problem
            return jsonify(payload), status

        old_courses = [dict(c) for c in courses]
        room_occupancy_notices = []
        changed = False
        for k, v in fields.items():
            if courses[idx].get(k) != v:
                courses[idx][k] = v
                changed = True
        if changed:
            c = courses[idx]
            result = check_course_conflict(
                dept_id, term_id, courses, c.get('teacher'), c.get('room'), c.get('campus'),
                c.get('season'), c.get('period'), c.get('slot'), c.get('day', ''), exclude_id=c.get('id')
            )
            hard_room_conflict = has_room_occupancy_notice(result) and not soft_allow_room_conflict
            if result['teacher_conflict'] or hard_room_conflict:
                return jsonify({'error': '修改后存在冲突', 'conflict_type': 'mixed', **result}), 409
            if soft_allow_room_conflict and has_room_occupancy_notice(result):
                room_occupancy_notices.append({
                    'course': dict(c),
                    'notice': room_occupancy_notice_text(result),
                })
            action = '产能表拖拽调整' if soft_allow_room_conflict else '表格编辑'
            try:
                save_term_data(
                    dept_id,
                    term_id,
                    courses,
                    user=user,
                    old_courses=old_courses,
                    action=action,
                    reason=reason,
                    expected_version=current_version,
                )
            except TermVersionConflict as exc:
                last_conflict = exc
                if attempt + 1 < COURSE_UPDATE_RETRY_LIMIT:
                    continue
                return term_version_conflict_response(exc)
            append_room_occupancy_notices(dept_id, term_id, user, action, reason, room_occupancy_notices)
        return jsonify_with_term_version(courses[idx], dept_id, term_id)

    if last_conflict:
        return term_version_conflict_response(last_conflict)
    return jsonify({'error': '保存失败，请刷新后重试'}), 409

@app.route('/dept/<dept_id>/<term_id>/api/courses/batch', methods=['PATCH'])
def t_update_courses_batch(dept_id, term_id):
    user, err = _check_dt(dept_id, term_id)
    if err: return err
    wf_err = check_workflow_permission(dept_id, term_id, user)
    if wf_err: return wf_err
    body = request.json or {}
    reason = str(body.get('reason') or '').strip()
    items = body.get('updates') or []
    allow_partial = bool(body.get('partial_success'))
    soft_allow_room_conflict = should_soft_allow_room_conflict(body)
    if not isinstance(items, list) or not items:
        return jsonify({'error': '缺少更新内容'}), 400
    if len(items) > 20:
        return jsonify({'error': '单次最多更新 20 个课程'}), 400
    version_problem = missing_data_version_problem(dept_id, term_id)
    if version_problem:
        payload, status = version_problem
        return jsonify(payload), status

    last_conflict = None
    for attempt in range(COURSE_UPDATE_RETRY_LIMIT):
        current_version = current_term_version_number(dept_id, term_id)
        courses = load_term_data(dept_id, term_id)
        old_courses = [dict(c) for c in courses]
        changed = False
        updated_courses = []
        conflicts = []
        errors = []
        room_occupancy_notices = []

        if not allow_partial:
            touched_indexes = []
            touched_index_set = set()
            for item in items:
                course_id = item.get('id')
                fields = course_patch_fields(item.get('fields') or {})
                base_fields = course_patch_base_fields(item)
                try:
                    course_id_int = int(course_id) if course_id is not None else None
                except (TypeError, ValueError):
                    course_id_int = None
                idx = find_course_index(courses, course_id_int) if course_id_int is not None else None
                if idx is None:
                    return jsonify({'error': f'课程不存在：{course_id}'}), 404
                if not can_edit_course(user, courses[idx], dept_id):
                    return jsonify({'error': f'无权编辑课程：{course_id}'}), 403
                if not is_active_course(courses[idx]):
                    return jsonify({'error': f'已取消或已合并的班级需先恢复后再编辑：{course_id}'}), 400
                field_problem = field_level_conflict_problem(
                    dept_id, term_id, courses[idx], fields, base_fields, current_version
                )
                if field_problem:
                    payload, status = field_problem
                    return jsonify(payload), status
                item_changed = False
                for k, v in fields.items():
                    if courses[idx].get(k) != v:
                        courses[idx][k] = v
                        item_changed = True
                if idx not in touched_index_set:
                    touched_indexes.append(idx)
                    touched_index_set.add(idx)
                if item_changed:
                    changed = True
                updated_courses.append(courses[idx])
            if changed:
                for idx in touched_indexes:
                    c = courses[idx]
                    result = check_course_conflict(
                        dept_id, term_id, courses, c.get('teacher'), c.get('room'), c.get('campus'),
                        c.get('season'), c.get('period'), c.get('slot'), c.get('day', ''), exclude_id=c.get('id')
                    )
                    hard_room_conflict = has_room_occupancy_notice(result) and not soft_allow_room_conflict
                    if result['teacher_conflict'] or hard_room_conflict:
                        label = c.get('code') or c.get('name') or c.get('id')
                        return jsonify({'error': f'修改后存在冲突：{label}', 'conflict_type': 'mixed', **result}), 409
                    if soft_allow_room_conflict and has_room_occupancy_notice(result):
                        room_occupancy_notices.append({
                            'course': dict(c),
                            'notice': room_occupancy_notice_text(result),
                        })
                action = '产能表拖拽调整'
                try:
                    save_term_data(
                        dept_id,
                        term_id,
                        courses,
                        user=user,
                        old_courses=old_courses,
                        action=action,
                        reason=reason,
                        expected_version=current_version,
                    )
                except TermVersionConflict as exc:
                    last_conflict = exc
                    if attempt + 1 < COURSE_UPDATE_RETRY_LIMIT:
                        continue
                    return term_version_conflict_response(exc)
                append_room_occupancy_notices(dept_id, term_id, user, action, reason, room_occupancy_notices)
            return jsonify_with_term_version({
                'ok': True,
                'partial': False,
                'courses': updated_courses,
                'success': [c.get('id') for c in updated_courses],
                'conflicts': conflicts,
                'errors': errors,
            }, dept_id, term_id)

        for item in items:
            course_id = item.get('id')
            fields = course_patch_fields(item.get('fields') or {})
            base_fields = course_patch_base_fields(item)
            try:
                course_id_int = int(course_id) if course_id is not None else None
            except (TypeError, ValueError):
                course_id_int = None
            idx = find_course_index(courses, course_id_int) if course_id_int is not None else None
            if idx is None:
                errors.append({'id': course_id, 'error': f'课程不存在：{course_id}'})
                continue
            if not can_edit_course(user, courses[idx], dept_id):
                errors.append({'id': course_id, 'error': f'无权编辑课程：{course_id}'})
                continue
            if not is_active_course(courses[idx]):
                errors.append({'id': course_id, 'error': f'已取消或已合并的班级需先恢复后再编辑：{course_id}'})
                continue
            field_problem = field_level_conflict_problem(
                dept_id, term_id, courses[idx], fields, base_fields, current_version
            )
            if field_problem:
                payload, _status = field_problem
                conflicts.append({
                    'id': course_id,
                    'label': payload.get('course_label') or course_label_for_message(courses[idx], course_id),
                    **payload,
                })
                continue
            before = dict(courses[idx])
            item_changed = False
            for k, v in fields.items():
                if courses[idx].get(k) != v:
                    courses[idx][k] = v
                    item_changed = True
            if not item_changed:
                updated_courses.append(courses[idx])
                continue
            c = courses[idx]
            result = check_course_conflict(
                dept_id, term_id, courses, c.get('teacher'), c.get('room'), c.get('campus'),
                c.get('season'), c.get('period'), c.get('slot'), c.get('day', ''), exclude_id=c.get('id')
            )
            hard_room_conflict = has_room_occupancy_notice(result) and not soft_allow_room_conflict
            if result['teacher_conflict'] or hard_room_conflict:
                label = c.get('code') or c.get('name') or c.get('id')
                courses[idx] = before
                conflicts.append({
                    'id': course_id,
                    'label': label,
                    'error': f'修改后存在冲突：{label}',
                    'conflict_type': 'mixed',
                    **result,
                })
                continue
            if soft_allow_room_conflict and has_room_occupancy_notice(result):
                room_occupancy_notices.append({
                    'course': dict(c),
                    'notice': room_occupancy_notice_text(result),
                })
            changed = True
            updated_courses.append(courses[idx])
        if changed:
            action = '批量修改课程'
            try:
                save_term_data(
                    dept_id,
                    term_id,
                    courses,
                    user=user,
                    old_courses=old_courses,
                    action=action,
                    reason=reason,
                    expected_version=current_version,
                )
            except TermVersionConflict as exc:
                last_conflict = exc
                if attempt + 1 < COURSE_UPDATE_RETRY_LIMIT:
                    continue
                return term_version_conflict_response(exc)
            append_room_occupancy_notices(dept_id, term_id, user, action, reason, room_occupancy_notices)
        return jsonify_with_term_version({
            'ok': True,
            'partial': allow_partial,
            'courses': updated_courses,
            'success': [c.get('id') for c in updated_courses],
            'conflicts': conflicts,
            'errors': errors,
        }, dept_id, term_id)

    if last_conflict:
        return term_version_conflict_response(last_conflict)
    return jsonify({'error': '保存失败，请刷新后重试'}), 409

@app.route('/dept/<dept_id>/<term_id>/api/courses/<int:course_id>/cancel', methods=['POST'])
def t_cancel_course(dept_id, term_id, course_id):
    user, err = _check_dt(dept_id, term_id)
    if err: return err
    wf_err = check_workflow_permission(dept_id, term_id, user)
    if wf_err: return wf_err
    conflict = term_version_conflict(dept_id, term_id)
    if conflict: return conflict
    courses = load_term_data(dept_id, term_id)
    idx = find_course_index(courses, course_id)
    if idx is None:
        return jsonify({'error': 'not found'}), 404
    if not can_edit_course(user, courses[idx], dept_id):
        return jsonify({'error': '无权取消该课程'}), 403
    if course_lifecycle_status(courses[idx]) == 'cancelled':
        return jsonify_with_term_version({'ok': True, 'course': courses[idx]}, dept_id, term_id)
    if not is_active_course(courses[idx]):
        return jsonify({'error': '已合并班级不能直接取消，请先恢复后再处理'}), 400
    body = request.json or {}
    reason = str(body.get('reason') or '').strip()
    old_courses = [dict(c) for c in courses]
    mark_course_lifecycle(courses[idx], 'cancelled', user, reason=reason)
    save_err = save_term_data_or_conflict(dept_id, term_id, courses, user=user, old_courses=old_courses, action='取消班级', reason=reason)
    if save_err: return save_err
    return jsonify_with_term_version({'ok': True, 'course': courses[idx]}, dept_id, term_id)

@app.route('/dept/<dept_id>/<term_id>/api/courses/merge', methods=['POST'])
def t_merge_courses(dept_id, term_id):
    user, err = _check_dt(dept_id, term_id)
    if err: return err
    wf_err = check_workflow_permission(dept_id, term_id, user)
    if wf_err: return wf_err
    conflict = term_version_conflict(dept_id, term_id)
    if conflict: return conflict
    body = request.json or {}
    source_id = body.get('source_id')
    target_raw = body.get('target_id') if 'target_id' in body else body.get('target_code')
    target_ref = str(target_raw).strip() if target_raw is not None else ''
    reason = str(body.get('reason') or '').strip()
    if source_id is None or not target_ref:
        return jsonify({'error': 'source_id 和 target_id/target_code 必填'}), 400
    courses = load_term_data(dept_id, term_id)
    try:
        source_id_int = int(source_id)
    except (TypeError, ValueError):
        return jsonify({'error': 'source_id 必须是课程 ID'}), 400
    source_idx = find_course_index(courses, source_id_int)
    target_idx = find_course_index(courses, int(target_ref)) if target_ref.isdigit() else None
    if target_idx is None:
        target_idx = next((i for i, c in enumerate(courses) if str(c.get('code') or '').strip() == target_ref), None)
    if source_idx is None or target_idx is None:
        return jsonify({'error': '来源班或目标班不存在'}), 404
    if source_idx == target_idx:
        return jsonify({'error': '不能合并到同一个班'}), 400
    source = courses[source_idx]
    target = courses[target_idx]
    if not can_edit_course(user, source, dept_id) or not can_edit_course(user, target, dept_id):
        return jsonify({'error': '无权合并所选班级'}), 403
    if not is_active_course(source):
        return jsonify({'error': '来源班已取消或已合并'}), 400
    if not is_active_course(target):
        return jsonify({'error': '目标班已取消或已合并'}), 400
    old_courses = [dict(c) for c in courses]
    mark_course_lifecycle(source, 'merged', user, reason=reason)
    source['merged_into_id'] = target.get('id')
    source['merged_into_code'] = target.get('code', '')
    source['merged_into_name'] = target.get('name', '')
    source['merged_count_added'] = merge_course_counts(target, source)
    merge_sources = target.get('merge_sources') or []
    merge_sources.append({
        **public_course_ref(source),
        'merged_at': time.strftime('%Y-%m-%d %H:%M:%S'),
        'merged_by': user.get('email'),
        'reason': reason,
    })
    target['merge_sources'] = merge_sources[-20:]
    save_err = save_term_data_or_conflict(dept_id, term_id, courses, user=user, old_courses=old_courses, action='合并班级', reason=reason)
    if save_err: return save_err
    return jsonify_with_term_version({'ok': True, 'source': source, 'target': target}, dept_id, term_id)

@app.route('/dept/<dept_id>/<term_id>/api/courses/<int:course_id>/restore', methods=['POST'])
def t_restore_course_lifecycle(dept_id, term_id, course_id):
    user, err = _check_dt(dept_id, term_id)
    if err: return err
    wf_err = check_workflow_permission(dept_id, term_id, user)
    if wf_err: return wf_err
    conflict = term_version_conflict(dept_id, term_id)
    if conflict: return conflict
    courses = load_term_data(dept_id, term_id)
    idx = find_course_index(courses, course_id)
    if idx is None:
        return jsonify({'error': 'not found'}), 404
    if not can_edit_course(user, courses[idx], dept_id):
        return jsonify({'error': '无权恢复该课程'}), 403
    if is_active_course(courses[idx]):
        return jsonify_with_term_version({'ok': True, 'course': courses[idx]}, dept_id, term_id)
    old_courses = [dict(c) for c in courses]
    reason = str((request.json or {}).get('reason') or '').strip() if request.is_json else ''
    course = courses[idx]
    target = None
    if course_lifecycle_status(course) == 'merged':
        target_id = course.get('merged_into_id')
        target_code = str(course.get('merged_into_code') or '').strip()
        if isinstance(target_id, int):
            target_idx = find_course_index(courses, target_id)
        elif str(target_id or '').isdigit():
            target_idx = find_course_index(courses, int(target_id))
        else:
            target_idx = None
        if target_idx is None and target_code:
            target_idx = next((i for i, c in enumerate(courses) if str(c.get('code') or '').strip() == target_code), None)
        if target_idx is not None:
            target = courses[target_idx]
            target['merge_sources'] = [
                s for s in (target.get('merge_sources') or [])
                if s.get('id') != course.get('id') and s.get('code') != course.get('code')
            ]
            if course.get('merged_count_added'):
                unmerge_course_counts(target, course)
    mark_course_lifecycle(course, 'active', user, reason=reason)
    conflict_result = check_course_conflict(
        dept_id, term_id, courses,
        course.get('teacher'), course.get('room'), course.get('campus'),
        course.get('season'), course.get('period'), course.get('slot'), course.get('day', ''),
        exclude_id=course.get('id'),
    )
    if (
        conflict_result.get('teacher_conflict')
        or conflict_result.get('room_conflict')
        or conflict_result.get('shared_room_conflict')
    ):
        label = course.get('code') or course.get('name') or course.get('id')
        return jsonify({'error': f'恢复后存在冲突：{label}', 'conflict_type': 'mixed', **conflict_result}), 409
    save_err = save_term_data_or_conflict(dept_id, term_id, courses, user=user, old_courses=old_courses, action='恢复班级状态', reason=reason)
    if save_err: return save_err
    payload = {'ok': True, 'course': course}
    if target:
        payload['target'] = target
    return jsonify_with_term_version(payload, dept_id, term_id)

@app.route('/dept/<dept_id>/<term_id>/api/courses/<int:course_id>', methods=['DELETE'])
def t_delete_course(dept_id, term_id, course_id):
    user, err = _check_dt(dept_id, term_id)
    if err: return err
    wf_err = check_workflow_permission(dept_id, term_id, user)
    if wf_err: return wf_err
    conflict = term_version_conflict(dept_id, term_id)
    if conflict: return conflict
    courses = load_term_data(dept_id, term_id)
    idx = find_course_index(courses, course_id)
    if idx is None:
        return jsonify({'error': 'not found'}), 404
    course = courses[idx]
    if not can_edit_course(user, course, dept_id):
        return jsonify({'error': '无权删除该课程'}), 403
    if not is_active_course(course):
        return jsonify({'error': '已取消或已合并的班级需先恢复后再删除'}), 400
    if not is_inserted_course(course):
        return jsonify({'error': '只能删除插空新增的课程；导入课程请通过回滚或重新导入处理'}), 400
    old_courses = [dict(c) for c in courses]
    removed = courses.pop(idx)
    reason = str((request.json or {}).get('reason') or '').strip() if request.is_json else ''
    save_err = save_term_data_or_conflict(dept_id, term_id, courses, user=user, old_courses=old_courses, action='删除插空排课', reason=reason)
    if save_err: return save_err
    log = load_changelog(dept_id, term_id)
    log.append({
        'time': time.strftime('%Y-%m-%d %H:%M:%S'),
        'user': user.get('name') or user.get('email'),
        'email': user.get('email'),
        'action': '删除插空排课',
        'reason': reason,
        'changes': [{
            'course_id': removed.get('id'),
            'code': removed.get('code', ''),
            'name': removed.get('name', ''),
            'field': 'delete',
            'from': '插空新增课程',
            'to': '已删除',
        }],
    })
    save_changelog(dept_id, term_id, log[-500:])
    return jsonify_with_term_version({'ok': True, 'deleted': removed}, dept_id, term_id)

@app.route('/dept/<dept_id>/<term_id>/api/conflicts')
def t_conflicts(dept_id, term_id):
    _, err = _check_dt(dept_id, term_id)
    if err: return err
    include_suggestions = str(request.args.get('suggestions', '1')).lower() not in {'0', 'false', 'no'}
    cache_key = conflict_suggestions_cache_key(dept_id, term_id)
    status = load_conflict_status(dept_id, term_id)
    etag = data_api_etag('conflicts', include_suggestions, cache_key, stable_json_hash(status))
    cached_response = conditional_not_modified_response(etag)
    if cached_response:
        return cached_response
    if include_suggestions:
        cached = get_conflict_suggestions_cache(cache_key)
        if cached is not None:
            return conditional_json_response({
                'teacher': cached['teacher'],
                'room': cached['room'],
                'status': status,
                'suggestions_ready': True,
            }, etag)
    else:
        cached = get_conflict_summary_cache(cache_key)
        if cached is not None:
            return conditional_json_response({
                'teacher': cached['teacher'],
                'room': cached['room'],
                'status': status,
                'suggestions_ready': False,
            }, etag)
    payload = build_conflict_groups_payload(dept_id, term_id, include_suggestions)
    if include_suggestions:
        set_conflict_suggestions_cache(cache_key, payload)
    else:
        set_conflict_summary_cache(cache_key, payload)
    return conditional_json_response({
        **payload,
        'status': status,
        'suggestions_ready': include_suggestions,
    }, etag)

@app.route('/dept/<dept_id>/<term_id>/api/conflicts/status', methods=['PATCH'])
def t_conflict_status_update(dept_id, term_id):
    user, err = _check_dt(dept_id, term_id)
    if err: return err
    if user.get('role') not in EDIT_ROLES:
        return jsonify({'error': 'forbidden'}), 403
    body = request.json or {}
    key = str(body.get('key') or '').strip()
    status = str(body.get('status') or '').strip()
    if not key or status not in {'未处理', '处理中', '已确认'}:
        return jsonify({'error': '状态不正确'}), 400
    item = update_conflict_status_atomic(dept_id, term_id, key, status, user)
    return jsonify({'ok': True, 'item': item})

@app.route('/dept/<dept_id>/<term_id>/api/history')
def t_history(dept_id, term_id):
    _, err = _check_dt(dept_id, term_id)
    if err: return err
    d = term_history_dir(dept_id, term_id)
    if not os.path.exists(d): return jsonify([])
    files = []
    history_meta = load_history_meta(dept_id, term_id)
    changelog = load_changelog(dept_id, term_id)
    for idx, fn in enumerate(sorted(os.listdir(d), reverse=True)[:30], start=1):
        p = os.path.join(d, fn)
        if not fn.endswith('.json') or not os.path.isfile(p):
            continue
        mtime = os.path.getmtime(p)
        saved_at_full = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(mtime))
        meta = history_meta.get(fn, {})
        diff_summary = meta.get('diff_summary') or {}
        class_count = meta.get('class_count')
        added_count = diff_summary.get('added_count')
        removed_count = diff_summary.get('removed_count')
        changed_count = diff_summary.get('changed_count')
        if changed_count is None and 'fields' in diff_summary:
            changed_count = int(diff_summary.get('changed_count') or 0)
        summary_pending = class_count is None or added_count is None or removed_count is None or changed_count is None
        changed_total = None if summary_pending else int(added_count or 0) + int(removed_count or 0) + int(changed_count or 0)
        actor = meta.get('user', '')
        if not actor:
            actor = next((entry.get('user', '') for entry in changelog if entry.get('time', '') >= saved_at_full), '')
        files.append({
            'filename': fn,
            'title': meta.get('action') or f'自动备份 {idx}',
            'saved_at': time.strftime('%Y-%m-%d %H:%M', time.localtime(mtime)),
            'mtime': saved_at_full,
            'actor': actor or '系统自动保存',
            'action': meta.get('action', ''),
            'reason': meta.get('reason', ''),
            'backup_kind': meta.get('backup_kind', 'before_change' if meta.get('action') else 'snapshot'),
            'before_version': meta.get('before_version'),
            'after_version': meta.get('after_version'),
            'diff_summary': diff_summary,
            'summary_pending': summary_pending,
            'class_count': class_count,
            'changed_total': changed_total,
            'added_count': added_count,
            'removed_count': removed_count,
            'changed_count': changed_count,
        })
    return jsonify(files)

@app.route('/dept/<dept_id>/<term_id>/api/history/<path:filename>/diff')
def t_history_diff(dept_id, term_id, filename):
    _, err = _check_dt(dept_id, term_id)
    if err: return err
    p = safe_history_path(dept_id, term_id, filename)
    if not p or not os.path.exists(p):
        return jsonify({'error': 'not found'}), 404
    with open(p, 'r', encoding='utf-8') as f:
        old_courses = json.load(f)
    current_courses = load_term_data(dept_id, term_id)
    return jsonify({
        'filename': filename,
        'old_count': len(old_courses),
        'current_count': len(current_courses),
        'diff': summarize_course_changes(old_courses, current_courses),
    })

@app.route('/dept/<dept_id>/<term_id>/api/rollback/<filename>/preview')
def t_rollback_preview(dept_id, term_id, filename):
    _, err = _check_dt(dept_id, term_id)
    if err: return err
    p = safe_history_path(dept_id, term_id, filename)
    if not p or not os.path.exists(p):
        return jsonify({'error': 'not found'}), 404
    with open(p, 'r', encoding='utf-8') as f:
        target_courses = json.load(f)
    current_courses = load_term_data(dept_id, term_id)
    return jsonify({
        'filename': filename,
        'target_count': len(target_courses),
        'current_count': len(current_courses),
        'diff': summarize_course_changes(current_courses, target_courses),
    })

@app.route('/dept/<dept_id>/<term_id>/api/history/<path:filename>/download')
def t_history_download(dept_id, term_id, filename):
    _, err = _check_dt(dept_id, term_id)
    if err: return err
    p = safe_history_path(dept_id, term_id, filename)
    if not p or not os.path.exists(p):
        return jsonify({'error': 'not found'}), 404
    return send_file(p, as_attachment=True, download_name=filename, mimetype='application/json')

@app.route('/dept/<dept_id>/<term_id>/api/changelog')
def t_changelog(dept_id, term_id):
    _, err = _check_dt(dept_id, term_id)
    if err: return err
    log = load_changelog(dept_id, term_id)
    log.reverse()
    return jsonify(log[:100])

@app.route('/dept/<dept_id>/<term_id>/api/system/status')
def t_system_status(dept_id, term_id):
    user, err = _check_dt(dept_id, term_id)
    if err: return err
    if user.get('role') not in STAFF_ROLES:
        return jsonify({'error': 'forbidden'}), 403
    courses = load_term_data(dept_id, term_id)
    metadata = load_term_version_metadata(dept_id, term_id)
    return jsonify({
        'database': {
            'path': sqlite_store.DB_PATH,
            'exists': os.path.exists(sqlite_store.DB_PATH),
            'size': os.path.getsize(sqlite_store.DB_PATH) if os.path.exists(sqlite_store.DB_PATH) else 0,
        },
        'courses_count': len(courses),
        'teachers_count': len({c.get('teacher') for c in courses if c.get('teacher')}),
        'campuses_count': len({c.get('campus') for c in courses if c.get('campus')}),
        'metadata': metadata,
        'latest_daily_backup': latest_daily_backup_info(dept_id, term_id),
        'history_count': history_file_count(dept_id, term_id),
        'workflow': load_workflow(dept_id, term_id),
        'runtime': runtime_diagnostics(),
        'slow_request_threshold_ms': int(SLOW_REQUEST_SECONDS * 1000),
        'recent_slow_requests': recent_slow_request_snapshot(),
        'server_time': time.strftime('%Y-%m-%d %H:%M:%S'),
    })

@app.route('/dept/<dept_id>/<term_id>/api/capacity')
def t_capacity_api(dept_id, term_id):
    _, err = _check_dt(dept_id, term_id)
    if err: return err
    if dept_id != 'yiduiyi':
        return jsonify({'error': '热力图仅在一对一部门可用'}), 404
    lessons = load_term_data(dept_id, term_id)
    slot_order = ['早一', '早二', '下一', '下二', '晚上']
    from collections import defaultdict
    # teacher -> {date -> {slot -> [{hours, campus}]}}
    matrix = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    teacher_meta = {}
    dates = set()
    for l in lessons:
        t = l.get('teacher')
        d = l.get('date')
        s = l.get('slot') or slot_from_time_range(l.get('timeRange', '')) or '未知'
        try:
            h = float(l.get('hours') or 0)
        except (TypeError, ValueError):
            h = 0
        campus = l.get('campus', '')
        if not t or not d: continue
        dates.add(d)
        matrix[t][d][s].append({
            'hours': h,
            'campus': campus,
            'room': l.get('room', ''),
            'student': l.get('student', ''),
            'subject': l.get('subject', ''),
            'time': l.get('timeRange', ''),
        })
        if t not in teacher_meta:
            teacher_meta[t] = {'campus': set(), 'subject': set()}
        teacher_meta[t]['campus'].add(campus)
        teacher_meta[t]['subject'].add(l.get('subject', ''))
    dates = sorted(dates)
    teachers = []
    for t in sorted(matrix.keys()):
        total_hours = sum(h['hours'] for d in matrix[t].values() for s in d.values() for h in s)
        schedule = {}
        for d in dates:
            day_slots = {}
            for s in slot_order:
                entries = matrix[t][d][s]
                if entries:
                    total_h = sum(e['hours'] for e in entries)
                    campuses_in_slot = list(set(e['campus'] for e in entries if e['campus']))
                    day_slots[s] = {
                        'hours': total_h,
                        'campuses': campuses_in_slot,
                        'lessons': entries[:20],
                    }
            if day_slots:
                schedule[d] = day_slots
        campus_list = sorted(teacher_meta[t]['campus'] - {''})
        teachers.append({
            'name': t,
            'total_hours': total_hours,
            'campus_count': len(campus_list),
            'campuses': campus_list,
            'subjects': sorted(teacher_meta[t]['subject'] - {''}),
            'schedule': schedule,
        })
    teachers.sort(key=lambda x: -x['total_hours'])
    all_campuses = sorted(set(c for tm in teacher_meta.values() for c in tm['campus'] if c))
    return jsonify({'dates': dates, 'slots': slot_order, 'teachers': teachers, 'all_campuses': all_campuses})


def slot_from_time_range(time_range):
    text = str(time_range or '')
    if not text:
        return ''
    m = re.search(r'(\d{1,2}):(\d{2})\s*-\s*(\d{1,2}):(\d{2})', text)
    if not m:
        return ''
    start_hour = int(m.group(1))
    start_min = int(m.group(2))
    minutes = start_hour * 60 + start_min
    if minutes < 10 * 60 + 20:
        return '早一'
    if minutes < 13 * 60:
        return '早二'
    if minutes < 15 * 60 + 40:
        return '下一'
    if minutes < 18 * 60:
        return '下二'
    return '晚上'

@app.route('/dept/<dept_id>/<term_id>/api/rollback/<filename>', methods=['POST'])
def t_rollback(dept_id, term_id, filename):
    user, err = _check_dt(dept_id, term_id)
    if err: return err
    if user.get('role') not in STAFF_ROLES:
        return jsonify({'error': '仅教务/管理员可恢复历史备份'}), 403
    wf_err = check_workflow_permission(dept_id, term_id, user)
    if wf_err: return wf_err
    conflict = term_version_conflict(dept_id, term_id)
    if conflict: return conflict
    p = safe_history_path(dept_id, term_id, filename)
    if not p or not os.path.exists(p):
        return jsonify({'error': 'not found'}), 404
    old_courses = load_term_data(dept_id, term_id)
    with open(p, 'r', encoding='utf-8') as f:
        courses = json.load(f)
    reason = str((request.json or {}).get('reason') or '').strip() if request.is_json else ''
    save_err = save_term_data_or_conflict(dept_id, term_id, courses, user=user, old_courses=old_courses, action='恢复历史备份', reason=reason)
    if save_err: return save_err
    return jsonify_with_term_version({'ok': True, 'count': len(courses)}, dept_id, term_id)

def parse_import_file(file_storage):
    try:
        import openpyxl
        import zipfile
        from openpyxl.utils.exceptions import InvalidFileException
    except ImportError:
        return None, None, {'error': 'openpyxl not installed', 'status': 500}
    if request.content_length and request.content_length > MAX_IMPORT_BYTES:
        return None, None, {'error': '文件过大，请上传 20MB 以内的 xlsx 文件', 'status': 413}
    filename = (file_storage.filename or '').lower()
    if not filename.endswith('.xlsx'):
        return None, None, {'error': '请上传 xlsx 文件（暂不支持 xls）', 'status': 400}
    try:
        wb = openpyxl.load_workbook(file_storage, data_only=True, read_only=True)
    except (InvalidFileException, OSError, ValueError, zipfile.BadZipFile):
        return None, None, {'error': 'Excel 文件无法读取，请确认文件未损坏且格式为 xlsx', 'status': 400}
    ws = wb[wb.sheetnames[0]]
    try:
        ws.reset_dimensions()
    except Exception:
        pass
    try:
        headers = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    except StopIteration:
        return None, None, {'error': 'Excel 文件为空', 'status': 400}
    col_map = {}
    for i, h in enumerate(headers):
        if h:
            col_map[str(h).strip()] = i
    aliases = {
        '季度': ['季度', '季度（外）'],
        '班级名称（外）': ['班级名称（外）', '班级名称'],
        '校区名称': ['校区名称', '校区'],
        '科目': ['科目', '学科', '实际科目'],
        '年级': ['年级', '实际年级'],
        '授课教师': ['授课教师', '教师姓名', '教师'],
        '期数': ['期数'],
        '上课时间(外)': ['上课时间(外)', '上课时间'],
        '上课教室': ['上课教室', '教室名称', '教室'],
        '班级课次数': ['班级课次数', '课次数'],
        '标准人数': ['标准人数', '班容名称', '班容'],
        '当前人数(占名额)': ['当前人数(占名额)', '当前人数'],
        '描述': ['描述', '班级状态'],
        '管理项目': ['管理项目'],
    }

    def has_col(name):
        return any(alias in col_map for alias in aliases.get(name, [name]))

    required_headers = ['季度', '班级名称（外）', '校区名称']
    missing = [h for h in required_headers if not has_col(h)]
    if missing:
        return None, None, {'error': '缺少必要表头：' + '、'.join(missing), 'status': 400}
    useful_headers = [
        '班级编码', '授课教师', '期数', '上课时间(外)', '上课教室',
        '班级课次数', '标准人数', '当前人数(占名额)', '描述',
    ]
    missing_useful = [h for h in useful_headers if not has_col(h)]

    def get_col(row, *names):
        for name in names:
            for alias in aliases.get(name, [name]):
                if alias in col_map and col_map[alias] < len(row):
                    val = row[col_map[alias]]
                    if val is not None and str(val).strip():
                        return val
        return None

    slot_map = {
        '08:00-10:00': 'A', '10:20-12:20': 'B', '10:35-12:35': 'B',
        '13:20-15:20': 'C', '13:30-15:30': 'C',
        '15:50-17:50': 'D', '16:00-18:00': 'D', '18:30-20:30': 'E',
        '08:30-10:30': 'A', '10:40-12:40': 'B', '10:30-12:30': 'B',
        '14:00-16:00': 'C', '16:10-18:10': 'D',
    }
    subject_map = {'思维': '益智', '益智': '益智', '双语': '双语', '博文': '博文'}

    def normalize_campus_name(campus):
        text = str(campus or '').strip()
        if not text:
            return ''
        if text in CANONICAL_CAMPUSES:
            return CANONICAL_CAMPUSES[text]
        if text.endswith('教学区'):
            return text
        return CANONICAL_CAMPUSES.get(text.replace('教学区', ''), text + '教学区')

    def normalize_subject(raw, class_name):
        text = str(raw or '').strip()
        if text in subject_map:
            return subject_map[text]
        for s in ['双语', '博文', '益智']:
            if s in text or s in class_name:
                return s
        if '思维' in class_name:
            return '益智'
        for s in ['科学', '实践', 'KET', 'PET', 'YLE']:
            if s in text or s in class_name:
                return text or s
        return text

    def normalize_capacity(value):
        text = str(value or '').strip()
        m = re.search(r'\d+', text)
        return m.group(0) if m else text

    new_courses = []
    skipped_details = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        season = get_col(row, '季度')
        if not season:
            if len(skipped_details) < 20:
                skipped_details.append({'row': row_num, 'reason': '缺少季度'})
            continue
        teacher = get_col(row, '修改后教师', '授课教师') or ''
        original_teacher = get_col(row, '授课教师') or ''
        period = get_col(row, '修改后期数', '期数') or ''
        original_period = get_col(row, '期数') or ''
        slot_direct = get_col(row, '修改后时间段', '上课时间(内)') or ''
        time_str = str(get_col(row, '上课时间(外)') or '')
        name = get_col(row, '班级名称（外）') or ''
        campus = normalize_campus_name(get_col(row, '校区名称') or '')
        code = get_col(row, '班级编码') or ''
        room = get_col(row, '上课教室') or ''
        sessions = get_col(row, '班级课次数')
        department = get_col(row, '标准部门', '管理项目') or ''
        capacity = normalize_capacity(get_col(row, '标准人数'))
        desc = get_col(row, '描述') or ''
        source_code = get_col(row, '源班班号') or ''
        current_count = get_col(row, '当前人数(占名额)')
        grade = str(get_col(row, '年级') or '').strip()

        sdr = get_col(row, '开课日期')
        edr = get_col(row, '结课日期')
        start_date = sdr.strftime('%Y-%m-%d') if hasattr(sdr, 'strftime') else str(sdr or '')
        end_date = edr.strftime('%Y-%m-%d') if hasattr(edr, 'strftime') else str(edr or '')

        slot = ''
        slot_from_time = ''
        tm = re.search(r'(\d{2}:\d{2}-\d{2}:\d{2})', time_str)
        if tm:
            slot_from_time = slot_map.get(tm.group(1), '')
        if slot_direct and str(slot_direct).strip().upper() in ('A', 'B', 'C', 'D', 'E'):
            slot = str(slot_direct).strip().upper()
        else:
            slot = slot_from_time
        time_match = re.search(r'(\d{2}:\d{2}-\d{2}:\d{2})', time_str)
        time_range = time_match.group(1) if time_match else ''
        day = ''
        if '每周五' in time_str: day = '周五'
        elif '每周六' in time_str: day = '周六'
        elif '每周日' in time_str: day = '周日'
        elif '每天' in time_str: day = '每天'
        subject = normalize_subject(get_col(row, '科目'), str(name))

        new_courses.append({
            'id': len(new_courses),
            'season': season, 'campus': campus, 'code': code, 'name': name,
            'teacher': str(teacher) if teacher else '',
            'period': period, 'timeDesc': time_str, 'timeRange': time_range,
            'slot': slot, 'day': day,
            'startDate': start_date, 'endDate': end_date,
            'room': room, 'sessions': sessions,
            'department': department, 'capacity': capacity,
            'desc': desc, 'sourceCode': source_code,
            'currentCount': str(current_count) if current_count else '',
            'subject': subject,
            'grade': grade,
            '_originalTeacher': str(original_teacher) if original_teacher else '',
            '_originalPeriod': str(original_period) if original_period else '',
        })

    if not new_courses:
        return None, None, {'error': '未解析到有效数据，请检查表头是否包含"季度"列', 'status': 400}

    original_courses = []
    for c in new_courses:
        oc = dict(c)
        oc['teacher'] = c['_originalTeacher']
        oc['period'] = c['_originalPeriod']
        original_courses.append(oc)
    info = {
        'filename': file_storage.filename or '',
        'sheet': ws.title,
        'rows': ws.max_row - 1 if ws.max_row else 0,
        'parsed_count': len(new_courses),
        'skipped_rows': max(0, (ws.max_row - 1 if ws.max_row else 0) - len(new_courses)),
        'skipped_details': skipped_details,
        'missing_optional_headers': missing_useful,
        'seasons': sorted({str(c.get('season', '')) for c in new_courses if c.get('season')}),
        'subjects': sorted({str(c.get('subject', '')) for c in new_courses if c.get('subject')}),
        'campus_count': len({c.get('campus', '') for c in new_courses if c.get('campus')}),
        'teacher_count': len({c.get('teacher', '') for c in new_courses if c.get('teacher')}),
        'no_code_count': sum(1 for c in new_courses if not c.get('code')),
    }
    return new_courses, original_courses, info


def parse_yiduiyi_import_file(file_storage):
    try:
        import openpyxl
        import zipfile
        from openpyxl.utils.exceptions import InvalidFileException
    except ImportError:
        return None, None, {'error': 'openpyxl not installed', 'status': 500}
    if request.content_length and request.content_length > MAX_IMPORT_BYTES:
        return None, None, {'error': '文件过大，请上传 20MB 以内的 xlsx 文件', 'status': 413}
    filename = (file_storage.filename or '').lower()
    if not filename.endswith('.xlsx'):
        return None, None, {'error': '请上传 xlsx 文件（暂不支持 xls）', 'status': 400}
    try:
        wb = openpyxl.load_workbook(file_storage, data_only=True, read_only=True)
    except (InvalidFileException, OSError, ValueError, zipfile.BadZipFile):
        return None, None, {'error': 'Excel 文件无法读取，请确认文件未损坏且格式为 xlsx', 'status': 400}
    ws = wb[wb.sheetnames[0]]
    try:
        ws.reset_dimensions()
    except Exception:
        pass
    try:
        headers = [cell for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    except StopIteration:
        return None, None, {'error': 'Excel 文件为空', 'status': 400}
    col_map = {str(h).strip(): i for i, h in enumerate(headers) if h is not None and str(h).strip()}
    required_headers = ['校区名称', '课次时间', '小时数', '科目', '教师姓名']
    missing = [h for h in required_headers if h not in col_map]
    if missing:
        return None, None, {'error': '缺少一对一必要表头：' + '、'.join(missing), 'status': 400}

    def get_col(row, *names):
        for name in names:
            if name in col_map and col_map[name] < len(row):
                val = row[col_map[name]]
                if val is not None and str(val).strip():
                    return val
        return None

    lessons = []
    skipped_count = 0
    skipped_details = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        teacher = str(get_col(row, '教师姓名') or '').strip()
        time_text = str(get_col(row, '课次时间') or '').strip()
        campus = str(get_col(row, '校区名称') or '').strip()
        subject = str(get_col(row, '科目') or '').strip()
        if not teacher or not time_text:
            skipped_count += 1
            if len(skipped_details) < 20:
                skipped_details.append({'row': row_num, 'reason': '缺少教师姓名或课次时间'})
            continue
        date_match = re.search(r'(\d{4}-\d{2}-\d{2})', time_text)
        time_match = re.search(r'(\d{1,2}:\d{2}\s*-\s*\d{1,2}:\d{2})', time_text)
        if not date_match:
            skipped_count += 1
            if len(skipped_details) < 20:
                skipped_details.append({'row': row_num, 'reason': '课次时间缺少日期'})
            continue
        date = date_match.group(1)
        time_range = time_match.group(1).replace(' ', '') if time_match else ''
        hours_raw = get_col(row, '小时数')
        minutes_raw = get_col(row, '分钟数')
        try:
            hours = float(hours_raw)
        except (TypeError, ValueError):
            try:
                hours = float(minutes_raw) / 60
            except (TypeError, ValueError):
                hours = 0
        lesson = {
            'id': len(lessons),
            'type': 'yiduiyi_lesson',
            'date': date,
            'slot': slot_from_time_range(time_range),
            'timeRange': time_range,
            'timeDesc': time_text,
            'hours': hours,
            'minutes': int(float(minutes_raw or hours * 60 or 0)),
            'campus': campus,
            'room': str(get_col(row, '教室名称') or '').strip(),
            'teacher': teacher,
            'teacherCode': str(get_col(row, '教师编号') or '').strip(),
            'subject': subject,
            'grade': str(get_col(row, '年级') or '').strip(),
            'student': str(get_col(row, '学员姓名') or '').strip(),
            'studentCode': str(get_col(row, '学号') or '').strip(),
            'packageCode': str(get_col(row, '课时包编号') or '').strip(),
            'lessonIndex': str(get_col(row, '课次序号') or '').strip(),
            'attendance': str(get_col(row, '教师考勤') or '').strip(),
            'studentAttendance': str(get_col(row, '学员考勤') or '').strip(),
            'product': str(get_col(row, '产品分组') or '').strip(),
            'classType': str(get_col(row, '班容类型') or '').strip(),
            'weekday': str(get_col(row, '星期') or '').strip(),
        }
        lessons.append(lesson)
    if not lessons:
        return None, None, {'error': '未解析到一对一课次数据，请检查表头是否包含"课次时间"和"教师姓名"', 'status': 400}
    dates = sorted({l['date'] for l in lessons})
    info = {
        'filename': file_storage.filename or '',
        'sheet': ws.title,
        'rows': max(0, row_num - 1) if 'row_num' in locals() else 0,
        'parsed_count': len(lessons),
        'skipped_rows': skipped_count,
        'skipped_details': skipped_details,
        'missing_optional_headers': [],
        'seasons': dates,
        'campus_count': len({l.get('campus', '') for l in lessons if l.get('campus')}),
        'teacher_count': len({l.get('teacher', '') for l in lessons if l.get('teacher')}),
        'subject_count': len({l.get('subject', '') for l in lessons if l.get('subject')}),
        'total_hours': round(sum(float(l.get('hours') or 0) for l in lessons), 2),
        'date_range': f"{dates[0]} 至 {dates[-1]}" if dates else '',
        'import_type': 'yiduiyi_heatmap',
        'no_code_count': 0,
    }
    return lessons, [dict(l) for l in lessons], info


def parse_department_import_file(dept_id, file_storage):
    if dept_id == 'yiduiyi':
        return parse_yiduiyi_import_file(file_storage)
    return parse_import_file(file_storage)

@app.route('/dept/<dept_id>/<term_id>/api/import/preview', methods=['POST'])
def t_import_preview(dept_id, term_id):
    user, err = _check_dt(dept_id, term_id)
    if err: return err
    if user.get('role') not in STAFF_ROLES:
        return jsonify({'error': '仅教务/管理员可预检导入'}), 403
    if 'file' not in request.files:
        return jsonify({'error': 'no file'}), 400
    new_courses, _, info = parse_department_import_file(dept_id, request.files['file'])
    if isinstance(info, dict) and info.get('error'):
        return jsonify({'error': info['error']}), info.get('status', 400)
    current_courses = load_term_data(dept_id, term_id)
    if info.get('import_type') == 'yiduiyi_heatmap':
        warnings = []
        if info['skipped_rows']:
            warnings.append(f"有 {info['skipped_rows']} 行未解析，通常是缺少教师姓名或课次时间")
        return jsonify({
            'ok': True,
            'preview': {
                **info,
                'current_count': len(current_courses),
                'replace_count': len(new_courses),
                'overwritten_by_code': 0,
                'added_by_code': len(new_courses),
                'diff': {},
                'warnings': warnings,
            }
        })
    current_codes = {str(c.get('code')) for c in current_courses if c.get('code')}
    incoming_codes = {str(c.get('code')) for c in new_courses if c.get('code')}
    overwritten = len(current_codes & incoming_codes)
    added = len(incoming_codes - current_codes)
    diff = summarize_course_changes(current_courses, new_courses)
    warnings = []
    if info['skipped_rows']:
        warnings.append(f"有 {info['skipped_rows']} 行未解析，通常是缺少季度")
    if info['missing_optional_headers']:
        warnings.append('缺少可选表头：' + '、'.join(info['missing_optional_headers']))
    if info['no_code_count']:
        warnings.append(f"有 {info['no_code_count']} 个班级没有班级编码，无法按编码判断覆盖关系")
    return jsonify({
        'ok': True,
        'preview': {
            **info,
            'current_count': len(current_courses),
            'replace_count': len(new_courses),
            'overwritten_by_code': overwritten,
            'added_by_code': added,
            'diff': diff,
            'warnings': warnings,
        }
    })

@app.route('/dept/<dept_id>/<term_id>/api/import', methods=['POST'])
def t_import(dept_id, term_id):
    user, err = _check_dt(dept_id, term_id)
    if err: return err
    if user.get('role') not in STAFF_ROLES:
        return jsonify({'error': '仅教务/管理员可执行覆盖导入'}), 403
    wf_err = check_workflow_permission(dept_id, term_id, user)
    if wf_err: return wf_err
    conflict = term_version_conflict(dept_id, term_id)
    if conflict: return conflict
    if 'file' not in request.files:
        return jsonify({'error': 'no file'}), 400
    new_courses, original_courses, info = parse_department_import_file(dept_id, request.files['file'])
    if isinstance(info, dict) and info.get('error'):
        return jsonify({'error': info['error']}), info.get('status', 400)
    old_courses = load_term_data(dept_id, term_id)
    reason = str(request.form.get('reason') or '').strip()
    action = '一对一课次导入' if info.get('import_type') == 'yiduiyi_heatmap' else 'Excel覆盖导入'
    save_err = save_term_data_or_conflict(dept_id, term_id, new_courses, user=user, old_courses=old_courses, action=action, reason=reason)
    if save_err: return save_err
    save_json(term_original_file(dept_id, term_id), original_courses)
    save_term_metadata(dept_id, term_id, user=user)
    payload = {'ok': True, 'count': len(new_courses)}
    if info.get('import_type') == 'yiduiyi_heatmap':
        payload.update({
            'import_type': info.get('import_type'),
            'teacher_count': info.get('teacher_count', 0),
            'total_hours': info.get('total_hours', 0),
            'date_range': info.get('date_range', ''),
        })
    return jsonify_with_term_version(payload, dept_id, term_id)

@app.route('/dept/<dept_id>/<term_id>/api/import/generate/preview', methods=['POST'])
def t_import_generate_preview(dept_id, term_id):
    user, err = _check_dt(dept_id, term_id)
    if err: return err
    if user.get('role') not in STAFF_ROLES:
        return jsonify({'error': '仅教务/管理员可执行此操作'}), 403
    mode = request.form.get('mode', 'spring_to_summer_autumn')
    if mode not in GENERATE_MODES:
        return jsonify({'error': f'不支持的模式：{mode}'}), 400
    if 'file' not in request.files:
        return jsonify({'error': 'no file'}), 400
    new_courses, _, info = parse_import_file(request.files['file'])
    if isinstance(info, dict) and info.get('error'):
        return jsonify({'error': info['error']}), info.get('status', 400)
    _, preview = build_generated_courses(new_courses, dept_id, term_id, mode)
    return jsonify({'ok': True, 'preview': preview})

@app.route('/dept/<dept_id>/<term_id>/api/import/generate', methods=['POST'])
def t_import_generate(dept_id, term_id):
    user, err = _check_dt(dept_id, term_id)
    if err: return err
    if user.get('role') not in STAFF_ROLES:
        return jsonify({'error': '仅教务/管理员可执行此操作'}), 403
    wf_err = check_workflow_permission(dept_id, term_id, user)
    if wf_err: return wf_err
    conflict = term_version_conflict(dept_id, term_id)
    if conflict: return conflict
    mode = request.form.get('mode', 'spring_to_summer_autumn')
    if mode not in GENERATE_MODES:
        return jsonify({'error': f'不支持的模式：{mode}'}), 400
    if 'file' not in request.files:
        return jsonify({'error': 'no file'}), 400
    new_courses, _, info = parse_import_file(request.files['file'])
    if isinstance(info, dict) and info.get('error'):
        return jsonify({'error': info['error']}), info.get('status', 400)
    generated, preview = build_generated_courses(new_courses, dept_id, term_id, mode)
    old_courses = load_term_data(dept_id, term_id)
    reason = str(request.form.get('reason') or '').strip()
    save_err = save_term_data_or_conflict(dept_id, term_id, generated, user=user, old_courses=old_courses, action='跨季节生成导入', reason=reason)
    if save_err: return save_err
    save_json(term_original_file(dept_id, term_id), generated)
    save_term_metadata(dept_id, term_id, user=user)
    return jsonify_with_term_version({'ok': True, **preview, 'total': len(generated)}, dept_id, term_id)

@app.route('/dept/<dept_id>/<term_id>/api/backup/json')
def t_backup_json(dept_id, term_id):
    _, err = _check_dt(dept_id, term_id)
    if err: return err
    dept_name = next((d['name'] for d in load_depts_config() if d['id'] == dept_id), dept_id)
    term_name = next((t['name'] for t in load_terms(dept_id) if t['id'] == term_id), term_id)
    filename = f'{dept_name}_{term_name}_data_backup.json'
    return send_file(
        term_data_file(dept_id, term_id),
        as_attachment=True,
        download_name=filename,
        mimetype='application/json',
    )

@app.route('/dept/<dept_id>/<term_id>/api/export')
def t_export(dept_id, term_id):
    _, err = _check_dt(dept_id, term_id)
    if err: return err
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return jsonify({'error': 'openpyxl not installed'}), 500
    courses = load_term_data(dept_id, term_id)
    original = load_term_original(dept_id, term_id)
    orig_map = {c['id']: c for c in original}
    changed_only = request.args.get('changed_only') == '1'
    slot_time_map_gaozhi = {'A':'08:00-10:00','B':'10:20-12:20','C':'13:30-15:30','D':'15:50-17:50','E':'18:30-20:30'}
    slot_time_map_qingshao = {'A':'08:30-10:30','B':'10:40-12:40','C':'14:00-16:00','D':'16:10-18:10','E':'18:30-20:30'}
    slot_time_map = slot_time_map_qingshao if dept_id == 'qingshao' else slot_time_map_gaozhi

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '导出说明'
    headers = ['是否调整','季度','校区名称','班级编码','班级名称（外）',
               '原授课教师','调整后教师','原期数','调整后期数',
               '原时段','调整后时段','原班型','调整后班型',
               '上课时间(外)','开课日期','结课日期','上课教室','班级课次数',
               '标准部门','标准人数','描述','源班班号','当前人数(占名额)',
               '班级状态','合并至班级','状态原因','状态时间','状态操作人','合并来源']
    header_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=10)
    changed_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    inserted_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
    orig_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 62
    ws['A1'] = '排课导出说明'
    ws['A1'].font = Font(bold=True, color='1A237E', size=16)
    ws.merge_cells('A1:B1')
    guide_rows = [
        ('排课调整版', '完整明细，可筛选、核对、二次处理。黄色表示调整，绿色表示插空新增，班级状态列标记取消/合并。'),
        ('校区课表看板', '尽量还原网页里的校区课表，用于给老师、店长直接查看。'),
        ('教师产能看板', '按教师和时段展示课程占用，红色表示同一时段多课冲突。'),
        ('教室空挡看板', '按校区和教室展示占用；绿色为空闲，红色为本部门占用，蓝色为其他部门占用，橙色为多部门同时占用。'),
        ('冲突汇总', '列出教师冲突、教室冲突和跨部门教室冲突。'),
        ('修改记录', '显示谁在什么时候通过什么方式修改了什么。'),
    ]
    for idx, (name, desc) in enumerate(guide_rows, 3):
        ws.cell(idx, 1, name).font = Font(bold=True, color='1A237E')
        ws.cell(idx, 2, desc).alignment = Alignment(wrap_text=True, vertical='top')
        for col in (1, 2):
            ws.cell(idx, col).border = thin_border
    ws = wb.create_sheet('排课调整版')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    row = 2
    for c in courses:
        orig = orig_map.get(c['id'], {})
        teacher_changed = c.get('teacher','') != orig.get('teacher','')
        period_changed = c.get('period','') != orig.get('period','')
        slot_changed = c.get('slot','') != orig.get('slot','')
        inserted = is_inserted_course(c)
        orig_name = orig.get('name', c.get('name',''))
        orig_class_type_match = re.search(r'[A-C]', re.sub(r'IPARK','', orig_name))
        orig_class_type = orig_class_type_match.group(0) if orig_class_type_match else ''
        cur_class_type = c.get('classType', orig_class_type)
        class_type_changed = cur_class_type != orig_class_type
        lifecycle_changed = course_lifecycle_status(c) != course_lifecycle_status(orig)
        lifecycle_meta_changed = any(
            c.get(key, '') != orig.get(key, '')
            for key in ['lifecycle_reason', 'lifecycle_at', 'lifecycle_by']
        )
        count_changed = c.get('currentCount', '') != orig.get('currentCount', '')
        merged_into_changed = c.get('merged_into_code', '') != orig.get('merged_into_code', '')
        merge_sources_changed = (c.get('merge_sources') or []) != (orig.get('merge_sources') or [])
        has_change = (
            inserted or teacher_changed or period_changed or slot_changed or class_type_changed
            or lifecycle_changed or lifecycle_meta_changed or count_changed
            or merged_into_changed or merge_sources_changed
        )
        orig_slot = orig.get('slot','')
        cur_slot = c.get('slot','')
        orig_time_display = f"{orig_slot} ({slot_time_map.get(orig_slot,'')})" if orig_slot else ''
        cur_time_display = f"{cur_slot} ({slot_time_map.get(cur_slot,'')})" if cur_slot else ''
        values = [
            '新增' if inserted else ('是' if has_change else ''),
            c.get('season',''), c.get('campus',''), c.get('code',''), c.get('name',''),
            orig.get('teacher',''), c.get('teacher','') if teacher_changed else '',
            orig.get('period',''), c.get('period','') if period_changed else '',
            orig_time_display, cur_time_display if slot_changed else '',
            orig_class_type, cur_class_type if class_type_changed else '',
            c.get('timeDesc',''), c.get('startDate',''), c.get('endDate',''),
            c.get('room',''), c.get('sessions',''), c.get('department',''),
            c.get('capacity',''), c.get('desc',''), c.get('sourceCode',''),
            c.get('currentCount',''), lifecycle_status_label(c),
            c.get('merged_into_code',''), c.get('lifecycle_reason',''),
            c.get('lifecycle_at',''), c.get('lifecycle_by',''),
            format_merge_sources_for_export(c),
        ]
        changed_idx = set()
        if teacher_changed: changed_idx.add(7)
        if period_changed: changed_idx.add(9)
        if slot_changed: changed_idx.add(11)
        if class_type_changed: changed_idx.add(13)
        if count_changed: changed_idx.add(23)
        if lifecycle_changed: changed_idx.add(24)
        if merged_into_changed: changed_idx.add(25)
        if lifecycle_meta_changed: changed_idx.update({26, 27, 28})
        if merge_sources_changed: changed_idx.add(29)
        orig_idx = {6, 8, 10, 12}
        if inserted:
            values[5] = ''
            values[6] = c.get('teacher', '')
            values[7] = ''
            values[8] = c.get('period', '')
            values[9] = ''
            values[10] = cur_time_display
            values[11] = ''
            values[12] = cur_class_type
            changed_idx = {1, 7, 9, 11, 13}
        if changed_only and not has_change:
            continue
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.font = Font(size=10)
            if inserted and col in changed_idx:
                cell.fill = inserted_fill
                cell.font = Font(size=10, bold=True)
            elif col in changed_idx:
                cell.fill = changed_fill
                cell.font = Font(size=10, bold=True)
            elif col in orig_idx:
                cell.fill = orig_fill
            if col == 1 and has_change:
                cell.fill = inserted_fill if inserted else changed_fill
                cell.font = Font(size=10, bold=True)
                cell.alignment = Alignment(horizontal='center')
        row += 1

    col_widths = [8,6,18,14,24,10,10,8,8,16,16,6,6,28,12,12,20,8,12,8,16,14,12,12,16,22,18,24,50]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}1"
    ws.freeze_panes = 'A2'

    def write_table(sheet, table_headers, rows, widths=None):
        for col, h in enumerate(table_headers, 1):
            cell = sheet.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        for row_idx, row_values in enumerate(rows, 2):
            for col, val in enumerate(row_values, 1):
                cell = sheet.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border
                cell.font = Font(size=10)
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(start_color='FAFBFC', end_color='FAFBFC', fill_type='solid')
        if rows:
            sheet.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(table_headers))}{len(rows) + 1}"
        sheet.freeze_panes = 'A2'
        if widths:
            for col, width in enumerate(widths, 1):
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    def short_grade(course):
        name = course.get('name', '') if isinstance(course, dict) else (course or '')
        if dept_id == 'qingshao':
            raw_grade = str(course.get('grade', '') if isinstance(course, dict) else '').strip()
            grade_map = {
                '0': '大班', '1': '一年级', '2': '二年级', '3': '三年级',
                '4': '四年级', '5': '五年级', '6': '六年级',
            }
            if raw_grade in {'幼儿园大班', '一年级', '二年级', '三年级', '四年级', '五年级', '六年级'}:
                return '大班' if raw_grade == '幼儿园大班' else raw_grade
            m = re.search(r'([一二三四五六])年级', name)
            if m:
                return m.group(1) + '年级'
            if '大班' in name or '幼儿园' in name or raw_grade == 'S3':
                return '大班'
            m = re.search(r'([0-6])级', name)
            if m:
                return grade_map.get(m.group(1), '')
            code = str(course.get('code', '') if isinstance(course, dict) else '')
            if len(code) >= 4:
                return grade_map.get(code[3], '')
            return ''
        if '高一准备' in name:
            return '初一'
        if '高一预备' in name:
            return '初二'
        if '高一预科' in name:
            return '初三'
        m = re.search(r'(高[一二三]|[789]级)', name)
        if m:
            grade = m.group(1)
            return {'7级': '初一', '8级': '初二', '9级': '初三'}.get(grade, grade)
        return ''

    def short_campus(name):
        return (name or '').replace('教学区', '').replace('购物中心', '')

    def short_room(room, campus=''):
        text = (room or '').replace(campus or '', '').replace('教学区', '')
        text = re.sub(r'(素养|素质|学习机|教室)', '', text)
        return text.strip() or room or ''

    def class_letter(name):
        text = re.sub(r'IPARK', '', name or '')
        m = re.search(r'(?:双语|益智|实践|科学|博文)(?:素养)?([A-C])', text)
        if m:
            return m.group(1)
        m = re.search(r'素养([A-C])', text)
        return m.group(1) if m else ''

    board_title_fill = PatternFill(start_color='E8EAF6', end_color='E8EAF6', fill_type='solid')
    summer_fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
    autumn_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    free_fill = PatternFill(start_color='ECFDF5', end_color='ECFDF5', fill_type='solid')
    own_room_fill = PatternFill(start_color='FFF1F2', end_color='FFF1F2', fill_type='solid')
    related_room_fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
    mixed_room_fill = PatternFill(start_color='FFF7ED', end_color='FFF7ED', fill_type='solid')
    conflict_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')

    def style_range(sheet, min_row, max_row, min_col, max_col, fill=None, font=None, align=None):
        for rr in range(min_row, max_row + 1):
            for cc in range(min_col, max_col + 1):
                cell = sheet.cell(rr, cc)
                cell.border = thin_border
                if fill:
                    cell.fill = fill
                if font:
                    cell.font = font
                if align:
                    cell.alignment = align

    def write_board_title(sheet, row_num, col_count, title, fill=None):
        sheet.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=col_count)
        title_font = Font(bold=True, color='1A237E', size=12)
        title_align = Alignment(horizontal='center', vertical='center')
        cell = sheet.cell(row_num, 1, title)
        cell.fill = fill or board_title_fill
        cell.font = title_font
        cell.alignment = title_align
        style_range(sheet, row_num, row_num, 1, col_count, fill or board_title_fill, title_font, title_align)
        sheet.row_dimensions[row_num].height = 24

    def course_board_text(c):
        pieces = []
        subject = c.get('subject', '')
        letter = class_letter(c.get('name', ''))
        pieces.append((subject + letter).strip())
        if c.get('teacher'):
            pieces.append(c.get('teacher'))
        if c.get('code'):
            pieces.append(c.get('code'))
        room = short_room(c.get('room', ''), c.get('campus', ''))
        if room:
            pieces.append(room)
        if c.get('currentCount'):
            pieces.append(f"{c.get('currentCount')}人")
        return '\n'.join(x for x in pieces if x)

    def write_schedule_board():
        sheet = wb.create_sheet('校区课表看板')
        sheet.sheet_view.showGridLines = False
        sheet.column_dimensions['A'].width = 16
        for col in range(2, 9):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22
        grouped = {}
        for c in active_courses(courses):
            if not c.get('campus') or not c.get('slot'):
                continue
            grade = short_grade(c) or '未分年级'
            key = (c.get('campus', ''), grade, c.get('season', ''), c.get('period', ''))
            grouped.setdefault(key, []).append(c)
        row_num = 1
        slot_order = ['A', 'B', 'C', 'D', 'E']
        for key in sorted(grouped):
            campus, grade, season, period = key
            group = grouped[key]
            by_slot = {s: sorted([c for c in group if c.get('slot') == s], key=lambda x: (class_letter(x.get('name', '')), x.get('subject', ''), x.get('name', ''))) for s in slot_order}
            max_cols = max(1, min(7, max((len(v) for v in by_slot.values()), default=1)))
            col_count = 1 + max_cols
            write_board_title(sheet, row_num, col_count, f"{short_campus(campus)} · {grade} · {season} {period}", summer_fill if season == '暑假' else autumn_fill)
            row_num += 1
            headers2 = ['时间段'] + [f'课程{i}' for i in range(1, max_cols + 1)]
            for col, value in enumerate(headers2, 1):
                cell = sheet.cell(row_num, col, value)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            row_num += 1
            for slot in slot_order:
                slot_courses = by_slot[slot]
                if not slot_courses:
                    continue
                sheet.cell(row_num, 1, f"{slot}\n{slot_time_map.get(slot, '')}")
                sheet.cell(row_num, 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                sheet.cell(row_num, 1).fill = orig_fill
                sheet.cell(row_num, 1).border = thin_border
                for idx in range(max_cols):
                    cell = sheet.cell(row_num, 2 + idx)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    if idx < len(slot_courses):
                        cell.value = course_board_text(slot_courses[idx])
                        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                    else:
                        cell.value = ''
                        cell.fill = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')
                sheet.row_dimensions[row_num].height = 56
                row_num += 1
            row_num += 1
        sheet.freeze_panes = 'A1'

    def write_capacity_board():
        sheet = wb.create_sheet('教师产能看板')
        sheet.sheet_view.showGridLines = False
        periods = [
            ('暑假', '1期', ['A', 'B', 'C', 'D', 'E']),
            ('暑假', '2期', ['A', 'B', 'C', 'D', 'E']),
            ('暑假', '3期', ['A', 'B', 'C', 'D', 'E']),
            ('秋季', '周五', ['E']),
            ('秋季', '周六', ['A', 'B', 'C', 'D', 'E']),
            ('秋季', '周日', ['A', 'B']),
        ]
        headers2 = ['科目', '教师', '年级']
        slot_keys = []
        for season, period, slots in periods:
            for slot in slots:
                headers2.append(f"{season[:1]}{period}{slot}")
                slot_keys.append((season, period, slot))
        for col, value in enumerate(headers2, 1):
            cell = sheet.cell(1, col, value)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 11 if col <= 3 else 16
        teacher_data = {}
        for c in active_courses(courses):
            teacher = c.get('teacher') or '未排教师'
            key = (c.get('subject', ''), teacher, short_grade(c))
            item = teacher_data.setdefault(key, {k: [] for k in slot_keys})
            if (c.get('season'), c.get('period'), c.get('slot')) in item:
                item[(c.get('season'), c.get('period'), c.get('slot'))].append(c)
        row_num = 2
        for subject, teacher, grade in sorted(teacher_data):
            sheet.cell(row_num, 1, subject)
            sheet.cell(row_num, 2, teacher)
            sheet.cell(row_num, 3, grade)
            for col in range(1, 4):
                sheet.cell(row_num, col).fill = orig_fill
                sheet.cell(row_num, col).font = Font(bold=(col == 2), size=10)
                sheet.cell(row_num, col).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row_num, col).border = thin_border
            for idx, sk in enumerate(slot_keys, 4):
                cls = teacher_data[(subject, teacher, grade)][sk]
                cell = sheet.cell(row_num, idx)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                if len(cls) > 1:
                    cell.fill = conflict_fill
                    cell.value = '\n---\n'.join(course_board_text(c) for c in cls[:3])
                elif cls:
                    cell.fill = summer_fill if sk[0] == '暑假' else autumn_fill
                    cell.value = course_board_text(cls[0])
                else:
                    cell.fill = free_fill
                    cell.value = '空'
                    cell.font = Font(color='9E9E9E')
            sheet.row_dimensions[row_num].height = 58
            row_num += 1
        sheet.freeze_panes = 'D2'

    def write_classroom_board():
        sheet = wb.create_sheet('教室空挡看板')
        sheet.sheet_view.showGridLines = False
        sheet.column_dimensions['A'].width = 16
        sheet.column_dimensions['B'].width = 12
        slot_order = ['A', 'B', 'C', 'D', 'E']
        campus_rooms = {}
        season_periods = {}
        occupied_by_slot = {}
        for c in classroom_courses:
            campus = c.get('campus')
            room = c.get('room')
            season = c.get('season')
            period = c.get('period')
            slot = c.get('slot')
            if campus and room:
                campus_rooms.setdefault(campus, set()).add(room)
            if season and period:
                season_periods.setdefault(season, set()).add(period)
            if campus and room and season and period and slot:
                occupied_by_slot.setdefault((campus, room, season, period, slot), []).append(c)
        seasons = sorted({c.get('season') for c in classroom_courses if c.get('season')})
        row_num = 1
        for campus in sorted(campus_rooms):
            rooms = sorted(campus_rooms.get(campus) or [])
            if not rooms:
                continue
            col_count = 2 + len(rooms)
            write_board_title(sheet, row_num, col_count, f"{short_campus(campus)} 教室空挡看板")
            row_num += 1
            headers3 = ['季度/期数', '时段'] + [short_room(r, campus) for r in rooms]
            for col, value in enumerate(headers3, 1):
                cell = sheet.cell(row_num, col, value)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border
                if col >= 3:
                    sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
            row_num += 1
            for season in seasons:
                periods = ['1期', '2期', '3期'] if season == '暑假' else ['周五', '周六', '周日'] if season == '秋季' else sorted(season_periods.get(season) or [])
                for period in periods:
                    for slot in slot_order:
                        sheet.cell(row_num, 1, f"{season} {period}")
                        sheet.cell(row_num, 2, f"{slot}\n{slot_time_map.get(slot, '')}")
                        for col in (1, 2):
                            sheet.cell(row_num, col).fill = orig_fill
                            sheet.cell(row_num, col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            sheet.cell(row_num, col).border = thin_border
                        for idx, room in enumerate(rooms, 3):
                            cs = occupied_by_slot.get((campus, room, season, period, slot), [])
                            cell = sheet.cell(row_num, idx)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            if cs:
                                has_own = any(not c.get('related') for c in cs)
                                has_related = any(c.get('related') for c in cs)
                                cell.fill = mixed_room_fill if has_own and has_related else related_room_fill if has_related else own_room_fill
                                cell.value = '\n---\n'.join(f"{c.get('dept_label', DEPT_LABELS.get(dept_id, dept_id))}\n{c.get('name','')}\n{c.get('teacher','')}" for c in cs[:3])
                            else:
                                cell.fill = free_fill
                                cell.value = '空闲'
                                cell.font = Font(color='2E7D32')
                        sheet.row_dimensions[row_num].height = 64
                        row_num += 1
                    row_num += 1
            row_num += 2

    write_schedule_board()

    overview_rows = []
    for c in sorted(active_courses(courses), key=lambda x: (
        x.get('campus', ''), short_grade(x), x.get('season', ''),
        x.get('period', ''), x.get('slot', ''), x.get('subject', ''), x.get('name', '')
    )):
        overview_rows.append([
            c.get('campus', ''), short_grade(c), c.get('season', ''),
            c.get('period', ''), c.get('slot', ''), slot_time_map.get(c.get('slot', ''), ''),
            c.get('subject', ''), c.get('name', ''), c.get('teacher', ''),
            c.get('room', ''), c.get('code', ''), c.get('currentCount', ''),
        ])
    ws_overview = wb.create_sheet('校区课表数据')
    write_table(
        ws_overview,
        ['校区', '年级', '季度', '期数', '时段', '上课时间', '科目', '班级名称', '教师', '教室', '班级编码', '当前人数'],
        overview_rows,
        [18, 8, 8, 8, 6, 14, 10, 28, 12, 24, 16, 10],
    )

    teacher_map = {}
    for c in active_courses(courses):
        teacher = c.get('teacher') or '未排教师'
        key = (teacher, c.get('subject', ''), short_grade(c))
        item = teacher_map.setdefault(key, {
            'teacher': teacher, 'subject': c.get('subject', ''), 'grade': short_grade(c),
            'campuses': set(), 'summer': 0, 'autumn': 0, 'total': 0,
        })
        item['campuses'].add(c.get('campus', ''))
        item['total'] += 1
        if c.get('season') == '暑假':
            item['summer'] += 1
        elif c.get('season') == '秋季':
            item['autumn'] += 1
    capacity_rows = [[
        v['teacher'], v['subject'], v['grade'], v['summer'], v['autumn'], v['total'],
        '、'.join(sorted(x for x in v['campuses'] if x)),
    ] for v in sorted(teacher_map.values(), key=lambda x: (x['subject'], x['teacher'], x['grade']))]
    write_capacity_board()

    ws_capacity = wb.create_sheet('教师产能数据')
    write_table(
        ws_capacity,
        ['教师', '科目', '年级', '暑假班级数', '秋季班级数', '总班级数', '涉及校区'],
        capacity_rows,
        [12, 10, 8, 12, 12, 10, 40],
    )
    for row_cells in ws_capacity.iter_rows(min_row=2):
        total = row_cells[5].value or 0
        if total >= 8:
            for cell in row_cells:
                cell.fill = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')

    related = related_room_courses_for_dept(dept_id, term_id)
    classroom_courses = active_courses(courses) + active_courses(related['courses'])
    classroom_rows = []
    for c in sorted(classroom_courses, key=lambda x: (
        x.get('campus', ''), x.get('room', ''), x.get('season', ''), x.get('period', ''), x.get('slot', '')
    )):
        owner = DEPT_LABELS.get(room_owner_dept(c.get('room', '')), room_owner_dept(c.get('room', '')) or '')
        classroom_rows.append([
            c.get('campus', ''), c.get('room', ''), owner,
            '其他部门占用' if c.get('related') else '本部门占用',
            c.get('dept_label', DEPT_LABELS.get(dept_id, dept_id)),
            c.get('season', ''), c.get('period', ''), c.get('slot', ''),
            slot_time_map.get(c.get('slot', ''), ''), c.get('name', ''),
            c.get('teacher', ''), c.get('code', ''),
        ])
    write_classroom_board()

    ws_rooms = wb.create_sheet('教室占用数据')
    write_table(
        ws_rooms,
        ['校区', '教室', '教室归属', '占用来源', '部门', '季度', '期数', '时段', '上课时间', '班级名称', '教师', '班级编码'],
        classroom_rows,
        [18, 24, 14, 14, 12, 8, 8, 6, 14, 28, 12, 16],
    )
    room_fills = {
        '本部门占用': PatternFill(start_color='FFF1F2', end_color='FFF1F2', fill_type='solid'),
        '其他部门占用': PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid'),
    }
    for row_cells in ws_rooms.iter_rows(min_row=2):
        fill = room_fills.get(row_cells[3].value)
        if fill:
            for cell in row_cells:
                cell.fill = fill

    teacher_conflicts = {}
    room_conflicts = {}
    for c in active_courses(courses):
        if c.get('teacher') and c.get('slot'):
            key = (c.get('teacher'), c.get('season'), c.get('period'), c.get('slot'), c.get('day', ''))
            teacher_conflicts.setdefault(key, []).append(c)
    for c in classroom_courses:
        if c.get('room') and c.get('campus') and c.get('slot'):
            key = (course_shared_room_key(c) or f"{c.get('room')}|{c.get('campus')}", c.get('season'), c.get('period'), c.get('slot'), c.get('day', ''))
            room_conflicts.setdefault(key, []).append(c)
    conflict_rows = []
    for key, group in teacher_conflicts.items():
        if len(group) > 1:
            conflict_rows.append(['教师时间冲突', key[0], key[1], key[2], key[3], key[4], '\n'.join(f"{g.get('code','')} {g.get('name','')} {g.get('campus','')}" for g in group)])
    for key, group in room_conflicts.items():
        if len(group) > 1:
            label = group[0].get('room', '')
            conflict_rows.append(['教室时间冲突', label, key[1], key[2], key[3], key[4], '\n'.join(f"{g.get('dept_label', DEPT_LABELS.get(dept_id, dept_id))} {g.get('code','')} {g.get('name','')}" for g in group)])
    ws_conflicts = wb.create_sheet('冲突汇总')
    write_table(
        ws_conflicts,
        ['冲突类型', '对象', '季度', '期数', '时段', '上课日', '涉及班级'],
        conflict_rows,
        [14, 24, 8, 8, 6, 8, 60],
    )

    log_rows = []
    for entry in reversed(load_changelog(dept_id, term_id)[-200:]):
        for change in entry.get('changes', []):
            log_rows.append([
                entry.get('time', ''), entry.get('user', ''), entry.get('action', '排课调整'), change.get('code', ''),
                change.get('name', ''), FIELD_LABELS_PY.get(change.get('field', ''), change.get('field', '')),
                change.get('from', ''), change.get('to', ''),
            ])
    ws_log = wb.create_sheet('修改记录')
    write_table(
        ws_log,
        ['修改时间', '操作人', '操作方式', '班级编码', '班级名称', '修改内容', '原来', '改为'],
        log_rows,
        [18, 14, 16, 16, 28, 12, 18, 18],
    )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    dept_name = next((d['name'] for d in load_depts_config() if d['id'] == dept_id), dept_id)
    term_name = next((t['name'] for t in load_terms(dept_id) if t['id'] == term_id), term_id)
    return send_file(
        output, as_attachment=True,
        download_name=f'{dept_name}_{term_name}_排课调整版.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5100, debug=False)
